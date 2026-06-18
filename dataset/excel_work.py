"""Выгружает датасет в Excel и загружает данные обратно из Excel."""

import os
import shutil
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from config import constant
from data_work import storage
from dataset.dataset_records import update_dataset_record


SHEET_NAME = "dataset"


def get_excel_headers() -> list:
    """Возвращает заголовки существующего Excel-файла."""
    workbook = load_workbook(constant.EDIT_EXCEL, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            return []
        return [str(value).strip() if value is not None else "" for value in header]
    finally:
        workbook.close()


def is_excel_schema_actual() -> bool:
    """Проверяет, совпадают ли колонки Excel с текущей схемой тегов."""
    return get_excel_headers() == constant.CSV_FIELDS


def move_excel_to_backup() -> bool:
    """Переносит устаревший Excel в backup перед пересозданием."""
    backup_dir = os.path.join(constant.DIR_TXT, "tags_backup")
    date_name = datetime.now().strftime('%d-%m-%Y %H-%M-%S-%f')
    new_name = date_name + " " + os.path.basename(constant.EDIT_EXCEL)

    try:
        os.makedirs(backup_dir, exist_ok=True)
        shutil.move(constant.EDIT_EXCEL, os.path.join(backup_dir, new_name))
        return True
    except PermissionError:
        print(f'Не удалось переместить устаревший Excel: {constant.EDIT_EXCEL}')
        print('Закрой файл в Excel и попробуй открыть датасет снова.')
        return False


def build_row(movie: dict) -> list:
    """Собирает строку Excel из записи фильма."""
    row = []
    for feature in constant.MAIN_INFO:
        row.append(movie["main_info"][feature])
    for feature in constant.RAW_SCORES:
        row.append(movie["raw_scores"][feature])
    for feature in constant.TAGS_VIBE:
        row.append(movie[constant.TAGS_VIBE_SECTION][feature])
    for feature in constant.GENRE:
        row.append(movie.get(constant.GENRE_SECTION, {}).get(feature, 0))
    return row


def apply_header_column_widths(worksheet, padding: int = 4) -> None:
    """Выставляет ширину столбцов по заголовкам с небольшим запасом."""
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    for column_index, value in enumerate(header, start=1):
        title = str(value or "")
        worksheet.column_dimensions[get_column_letter(column_index)].width = len(title) + padding


def export_dataset_to_excel(overwrite: bool = False) -> bool:
    """Выгружает датасет в Excel."""
    data = storage.load_dataset()
    meta = storage.load_meta()
    os.makedirs(constant.DIR_TXT, exist_ok=True)

    if overwrite is False and os.path.exists(constant.EDIT_EXCEL):
        try:
            if is_excel_schema_actual():
                print(f'Excel для редактирования уже существует: {constant.EDIT_EXCEL}')
                print('Открываю существующий файл без перезаписи.')
                return True
        except PermissionError:
            print(f'Не удалось прочитать Excel: {constant.EDIT_EXCEL}')
            print('Закрой файл в Excel и попробуй открыть датасет снова.')
            return False

        print('Схема тегов изменилась. Пересоздаю Excel по актуальным колонкам.')
        if move_excel_to_backup() is False:
            return False

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(constant.CSV_FIELDS)

    for movie in data.values():
        worksheet.append(build_row(movie))

    if len(data) == 0 and len(meta) > 0:
        for movie in meta.values():
            row = []
            for feature in constant.MAIN_INFO:
                row.append(movie["main_info"][feature])
            for feature in constant.RAW_SCORES:
                row.append(movie["raw_scores"][feature])
            row.extend([""] * len(constant.TAGS_VIBE))
            row.extend([""] * len(constant.GENRE))
            worksheet.append(row)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    apply_header_column_widths(worksheet)

    try:
        workbook.save(constant.EDIT_EXCEL)
    except PermissionError:
        print(f'Не удалось открыть Excel для записи: {constant.EDIT_EXCEL}')
        print('Закрой файл в Excel и попробуй снова.')
        return False

    print(f'Excel для редактирования сохранен: {constant.EDIT_EXCEL}')
    return True


def load_movies_from_excel() -> list:
    """Загружает фильмы из Excel."""
    try:
        workbook = load_workbook(constant.EDIT_EXCEL, data_only=True, read_only=True)
    except PermissionError:
        print(f'Не удалось прочитать Excel: {constant.EDIT_EXCEL}')
        print('Закрой файл в Excel и попробуй снова.')
        return None

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)

        if header is None:
            print('Excel-файл пуст!')
            return None

        fieldnames = [str(value).strip() if value is not None else "" for value in header]
        if fieldnames != constant.CSV_FIELDS:
            print('Ошибка Excel! Заголовки не совпадают с ожидаемыми')
            print('Ожидались:', constant.CSV_FIELDS)
            print('Получены:', fieldnames)
            return None

        movies = []
        for row_number, values in enumerate(rows, start=2):
            if all(value is None or str(value).strip() == "" for value in values):
                continue

            row = {
                field: "" if value is None else str(value)
                for field, value in zip(fieldnames, values)
            }
            movie = storage.build_movie_from_row(row, row_number)
            if movie is None:
                return None
            movies.append(movie)

        if len(movies) == 0:
            print('В Excel нет записей для импорта.')
            return None

        return movies
    finally:
        workbook.close()


def print_excel_import_forbidden_message() -> None:
    print('Excel-импорт остановлен.')
    print('Через Excel нельзя добавлять, удалять или переименовывать записи.')
    print('Для добавления используй пункт “Добавить запись”.')
    print('Для переименования используй пункт “Переименовать запись”.')


def build_patch_payload(movie: dict) -> dict:
    """Собирает безопасный patch из строки Excel без изменения title/key."""
    return {
        "main_info": {
            "user_score": movie["main_info"]["user_score"],
            "year": movie["main_info"]["year"],
        },
        "raw_scores": movie["raw_scores"],
        constant.TAGS_VIBE_SECTION: movie.get(constant.TAGS_VIBE_SECTION, {}),
        constant.GENRE_SECTION: movie.get(constant.GENRE_SECTION, {}),
    }


def validate_excel_titles(movies: list, dataset: dict) -> bool:
    """Проверяет, что Excel не добавляет, не удаляет и не переименовывает записи."""
    dataset_titles = list(dataset.keys())
    excel_titles = [movie["main_info"]["title"] for movie in movies]

    if len(excel_titles) != len(set(excel_titles)):
        print_excel_import_forbidden_message()
        print('В Excel есть повторяющиеся title.')
        return False

    unknown_in_dataset = sorted(set(excel_titles) - set(dataset_titles))
    if len(unknown_in_dataset) > 0:
        print('Excel-импорт остановлен.')
        print(f'Найдена неизвестная или переименованная запись: {unknown_in_dataset[0]}')
        print('Через Excel нельзя добавлять или переименовывать записи.')
        print('Для переименования используй пункт “Переименовать запись”.')
        print('Для добавления новой записи используй обычное добавление записи.')
        return False

    missing_in_excel = sorted(set(dataset_titles) - set(excel_titles))
    if len(missing_in_excel) > 0:
        print('Excel-импорт остановлен.')
        print(f'В Excel отсутствует запись из dataset: {missing_in_excel[0]}')
        print('Полная пересборка dataset через Excel временно запрещена.')
        print('Excel сейчас должен содержать тот же набор title, что и dataset.')
        return False

    return True


def replace_dataset_from_excel() -> bool:
    """Обновляет существующие записи dataset безопасным patch-импортом из Excel."""
    if os.path.exists(constant.EDIT_EXCEL) is False:
        print(f'Файл для редактирования не найден: {constant.EDIT_EXCEL}')
        return False

    movies = load_movies_from_excel()
    if movies is None:
        return False

    dataset = storage.load_dataset()
    if validate_excel_titles(movies, dataset) is False:
        return False

    if storage.is_file_writable(constant.FILE_NAME) is False:
        print(f'Не удалось открыть файл для записи: {constant.FILE_NAME}')
        print('Закрой файл в другой программе или проверь права доступа и попробуй снова.')
        return False

    storage.create_backup()
    updated_count = 0
    unchanged_count = 0

    for movie in movies:
        title = movie["main_info"]["title"]
        patch_payload = build_patch_payload(movie)
        result = update_dataset_record(title, patch_payload, source_name="Excel import")
        if result.ok is False:
            storage.save_dataset(dataset)
            print('Excel-импорт остановлен.')
            print(f'Проблемная строка: {title}')
            print(f'reason: {result.reason}')
            print(result.message)
            print('Dataset восстановлен к состоянию до импорта Excel.')
            return False
        if result.reason == "nothing_changed":
            unchanged_count += 1
        else:
            updated_count += 1

    print('Excel-импорт завершён.')
    print(f'Обновлено записей: {updated_count}')
    print(f'Пропущено без изменений: {unchanged_count}')
    print('Ошибок: 0')
    return True
