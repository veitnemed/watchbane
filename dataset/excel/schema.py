"""Excel sheet name and schema checks."""

import os
import shutil
from datetime import datetime

from openpyxl import load_workbook

from config import constant

SHEET_NAME = "dataset"


def get_excel_headers() -> list:
    """Возвращает заголовки существующего Excel-файла."""
    workbook = load_workbook(constant.EDIT_EXCEL, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            return []
        return [str(value).strip() if value is not None else "" for value in header]
    finally:
        workbook.close()


def is_excel_schema_actual() -> bool:
    """Проверяет, совпадают ли колонки Excel с текущей схемой тегов."""
    return get_excel_headers() == constant.CSV_FIELDS


def move_excel_to_backup() -> bool:
    """Переносит устаревший Excel в backup перед пересозданием."""
    backup_dir = os.path.join(constant.DIR_TXT, "tags_backup")
    date_name = datetime.now().strftime("%d-%m-%Y %H-%M-%S-%f")
    new_name = date_name + " " + os.path.basename(constant.EDIT_EXCEL)

    try:
        os.makedirs(backup_dir, exist_ok=True)
        shutil.move(constant.EDIT_EXCEL, os.path.join(backup_dir, new_name))
        return True
    except PermissionError:
        print(f"Не удалось переместить устаревший Excel: {constant.EDIT_EXCEL}")
        print("Закрой файл в Excel и попробуй открыть датасет снова.")
        return False
