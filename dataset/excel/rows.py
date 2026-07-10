"""Excel row builders and worksheet formatting."""

from openpyxl.utils import get_column_letter

from config import constant


def build_row(movie: dict) -> list:
    """Собирает строку Excel из записи фильма."""
    row = []
    for feature in constant.MAIN_INFO:
        row.append(movie["main_info"][feature])
    for feature in constant.RAW_SCORES:
        row.append(movie["raw_scores"][feature])
    for feature in constant.GENRE:
        row.append(movie.get(constant.GENRE_SECTION, {}).get(feature, 0))
    return row


def apply_header_column_widths(worksheet, padding: int = 4) -> None:
    """Выставляет ширину столбцов по заголовкам с небольшим запасом."""
    header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
    for column_index, value in enumerate(header, start=1):
        title = str(value or "")
        worksheet.column_dimensions[get_column_letter(column_index)].width = len(title) + padding
