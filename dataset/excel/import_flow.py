"""Import watched dataset updates from Excel."""

import os

from openpyxl import load_workbook

from config import constant
from dataset import storage_movie
from dataset.dataset_records import update_dataset_record
from storage import data as storage_data
from storage import files as storage_files


def print_excel_import_forbidden_message() -> None:
    print("Excel-импорт остановлен.")
    print("Через Excel нельзя добавлять, удалять или переименовывать записи.")
    print("Для добавления используй пункт “Добавить запись”.")
    print("Для переименования используй пункт “Переименовать запись”.")


def build_patch_payload(movie: dict) -> dict:
    """Собирает безопасный patch из строки Excel без изменения title/key."""
    return {
        "main_info": {
            "user_score": movie["main_info"]["user_score"],
            "year": movie["main_info"]["year"],
        },
        "raw_scores": movie["raw_scores"],
        constant.TAGS_VIBE_SECTION: movie.get(constant.TAGS_VIBE_SECTION, {}),
        constant.GENRE_SECTION: movie.get(constant.GENRE_SECTION, {}),
    }


def validate_excel_titles(movies: list, dataset: dict) -> bool:
    """Проверяет, что Excel не добавляет, не удаляет и не переименовывает записи."""
    dataset_titles = list(dataset.keys())
    excel_titles = [movie["main_info"]["title"] for movie in movies]

    if len(excel_titles) != len(set(excel_titles)):
        print_excel_import_forbidden_message()
        print("В Excel есть повторяющиеся title.")
        return False

    unknown_in_dataset = sorted(set(excel_titles) - set(dataset_titles))
    if len(unknown_in_dataset) > 0:
        print("Excel-импорт остановлен.")
        print(f"Найдена неизвестная или переименованная запись: {unknown_in_dataset[0]}")
        print("Через Excel нельзя добавлять или переименовывать записи.")
        print("Для переименования используй пункт “Переименовать запись”.")
        print("Для добавления новой записи используй обычное добавление записи.")
        return False

    missing_in_excel = sorted(set(dataset_titles) - set(excel_titles))
    if len(missing_in_excel) > 0:
        print("Excel-импорт остановлен.")
        print(f"В Excel отсутствует запись из dataset: {missing_in_excel[0]}")
        print("Полная пересборка dataset через Excel временно запрещена.")
        print("Excel сейчас должен содержать тот же набор title, что и dataset.")
        return False

    return True


def load_movies_from_excel() -> list:
    """Загружает фильмы из Excel."""
    try:
        workbook = load_workbook(constant.EDIT_EXCEL, data_only=True, read_only=True)
    except PermissionError:
        print(f"Не удалось прочитать Excel: {constant.EDIT_EXCEL}")
        print("Закрой файл в Excel и попробуй снова.")
        return None

    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)

        if header is None:
            print("Excel-файл пуст!")
            return None

        fieldnames = [str(value).strip() if value is not None else "" for value in header]
        if fieldnames != constant.CSV_FIELDS:
            print("Ошибка Excel! Заголовки не совпадают с ожидаемыми")
            print("Ожидались:", constant.CSV_FIELDS)
            print("Получены:", fieldnames)
            return None

        movies = []
        for row_number, values in enumerate(rows, start=2):
            if all(value is None or str(value).strip() == "" for value in values):
                continue

            row = {
                field: "" if value is None else str(value)
                for field, value in zip(fieldnames, values)
            }
            movie = storage_movie.build_movie_from_row(row, row_number)
            if movie is None:
                return None
            movies.append(movie)

        if len(movies) == 0:
            print("В Excel нет записей для импорта.")
            return None

        return movies
    finally:
        workbook.close()


def replace_dataset_from_excel() -> bool:
    """Обновляет существующие записи dataset безопасным patch-импортом из Excel."""
    if os.path.exists(constant.EDIT_EXCEL) is False:
        print(f"Файл для редактирования не найден: {constant.EDIT_EXCEL}")
        return False

    movies = load_movies_from_excel()
    if movies is None:
        return False

    dataset = storage_data.load_dataset()
    if validate_excel_titles(movies, dataset) is False:
        return False

    storage_files.create_backup()
    updated_count = 0
    unchanged_count = 0

    for movie in movies:
        title = movie["main_info"]["title"]
        patch_payload = build_patch_payload(movie)
        result = update_dataset_record(title, patch_payload, source_name="Excel import")
        if result.ok is False:
            storage_data.save_dataset(dataset)
            print("Excel-импорт остановлен.")
            print(f"Проблемная строка: {title}")
            print(f"reason: {result.reason}")
            print(result.message)
            print("Dataset восстановлен к состоянию до импорта Excel.")
            return False
        if result.reason == "nothing_changed":
            unchanged_count += 1
        else:
            updated_count += 1

    print("Excel-импорт завершён.")
    print(f"Обновлено записей: {updated_count}")
    print(f"Пропущено без изменений: {unchanged_count}")
    print("Ошибок: 0")
    return True
