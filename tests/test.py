"""Проверяет основные сценарии работы проекта на временных данных."""

import contextlib
import copy
import io
import inspect
import json
import tempfile
from pathlib import Path
import sys
from unittest.mock import patch
from urllib.error import HTTPError

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from config import constant
from config import scheme
from common import format_score
from model import model
from model import linear_regression_train
from model import noise_experiment
from model import train_report
from model import feature_ablation
from model import genre_markup_efficiency
from storage import data as storage_data
from storage import files as storage_files
from storage import normalize as storage_normalize
from dataset import storage_movie
from dataset import excel_work
from candidates import candidate_pool
from candidates import service as candidate_service
from candidates import genres as pool_genres
from candidates import import_tmdb as tmdb_import
from candidates import schema as candidate_schema
from candidates import genre_schema
from candidates import country_schema
from candidates import tmdb_candidate_pool
from candidates import kp_enrichment
from candidates import kp_tmdb_build_debug
from candidates import tmdb_country_options
from candidates import tmdb_genre_options
from candidates import to_dataset as candidate_to_dataset
from apis import imdb_sql as sql_search
from dataset import title_resolve
from scripts.dublecate import instrumenty_povtorov as pool_duplicate_tools
from ui.console import interface_funcs
from ui.console import request as request_ui
from apis import kp_api as api
from apis import tmdb_api
from common import valid
from web import export as web_export


def show_check(text: str, result: bool) -> None:
    """Печатает результат проверки."""
    print(f"{text}: {result}")


def assert_check(text: str, result: bool) -> None:
    """Проверяет условие и останавливает тест при ошибке."""
    show_check(text, result)
    assert result, text


def make_movie(title="Test Movie", user_score=8.0, raw_score=8.0) -> dict:
    """Создает тестовый объект фильма."""
    tags_vibe = {feature: 0 for feature in constant.TAGS_VIBE}
    genre_tags = {feature: 0 for feature in constant.GENRE}
    for feature in ["has_crime", "has_psychology", "has_romantic_pursuit"]:
        if feature in tags_vibe:
            tags_vibe[feature] = 1
    for feature in ["has_drama", "has_crime"]:
        if feature in genre_tags:
            genre_tags[feature] = 1

    return {
        "main_info": {
            "title": title,
            "user_score": user_score,
            "year": 2024
        },
        "raw_scores": {
            "kp_score": raw_score,
            "kp_votes": 120000,
            "imdb_score": raw_score,
            "imdb_votes": 1200
        },
        "computed_scores": format_score.raw_to_struct(
            {
                "kp_score": raw_score,
                "kp_votes": 120000,
                "imdb_score": raw_score,
                "imdb_votes": 1200,
            },
            {
                "title": title,
                "user_score": user_score,
                "year": 2024,
            },
        ),
        constant.TAGS_VIBE_SECTION: tags_vibe,
        constant.GENRE_SECTION: genre_tags
    }


def test_build_watched_movie_card() -> None:
    """Проверяет compact card для будущего web-экспорта."""
    print("\n0w.1) Проверяем build_watched_movie_card")
    movie = make_movie(title="Русская драма", user_score=9.5, raw_score=8.9)
    movie["main_info"]["year"] = 2008
    original = copy.deepcopy(movie)

    card = web_export.build_watched_movie_card(movie)

    assert_check("Card возвращает title", card["title"] == "Русская драма")
    assert_check("Card возвращает year", card["year"] == 2008)
    assert_check("Card возвращает user_score", card["user_score"] == 9.5)
    assert_check("Card возвращает kp_score", card["kp_score"] == 8.9)
    assert_check("Card возвращает imdb_score", card["imdb_score"] == 8.9)
    assert_check("Card не падает без poster", card["poster_url"] is None and card["poster_path"] is None)
    assert_check("Card не падает без overview", card["overview"] is None)
    assert_check("Card ставит runtime_status", card["runtime_status"] == "watched")
    assert_check("Card не мутирует исходную запись", movie == original)


def test_export_watched_movies_json() -> None:
    """Проверяет JSON payload и UTF-8 export для будущего web UI."""
    print("\n0w.2) Проверяем export_watched_movies_json")
    data = {
        "Русская драма": make_movie(title="Русская драма", user_score=9.5, raw_score=8.9),
        "Second Movie": make_movie(title="Second Movie", user_score=7.0, raw_score=7.5),
    }

    with tempfile.TemporaryDirectory() as temp_root:
        path = Path(temp_root) / "nested" / "watched_movies.json"
        result_path = web_export.export_watched_movies_json(data, path=path)
        file_exists = path.exists()
        raw_text = path.read_text(encoding="utf-8")
        payload = json.loads(raw_text)

    assert_check("Export возвращает путь", result_path == path)
    assert_check("Export создаёт JSON", file_exists)
    assert_check("report_type watched_movies", payload["report_type"] == "watched_movies")
    assert_check("created_at записан", isinstance(payload.get("created_at"), str) and payload["created_at"] != "")
    assert_check("count записан", payload["count"] == 2)
    assert_check("items записаны", len(payload["items"]) == 2)
    assert_check("Русские строки сохраняются без unicode escape", "Русская драма" in raw_text and "\\u0420" not in raw_text)


def test_watched_movies_export_is_read_only() -> None:
    """Проверяет, что web export не трогает dataset/pool/model и сетевые API."""
    print("\n0w.3) Проверяем read-only защиту watched web export")
    data = {"Read Only": make_movie(title="Read Only", user_score=8.0, raw_score=8.1)}

    with tempfile.TemporaryDirectory() as temp_root:
        path = Path(temp_root) / "watched_movies.json"
        with contextlib.ExitStack() as stack:
            save_dataset_mock = stack.enter_context(
                patch("storage.data.save_dataset", side_effect=AssertionError("save_dataset не должен вызываться"))
            )
            save_pool_mock = stack.enter_context(
                patch("candidates.candidate_pool.save_candidate_pool", side_effect=AssertionError("save_candidate_pool не должен вызываться"))
            )
            save_weights_mock = stack.enter_context(
                patch("storage.data.save_weights", side_effect=AssertionError("save_weights не должен вызываться"))
            )
            save_metrics_mock = stack.enter_context(
                patch("storage.data.save_model_metrics", side_effect=AssertionError("save_model_metrics не должен вызываться"))
            )
            kp_api_mock = stack.enter_context(
                patch("apis.kp_api.find_series_raw", side_effect=AssertionError("KP API не должен вызываться"))
            )
            tmdb_search_mock = stack.enter_context(
                patch("apis.tmdb_api.search_tv_by_name", side_effect=AssertionError("TMDb API не должен вызываться"))
            )
            tmdb_details_mock = stack.enter_context(
                patch("apis.tmdb_api.get_tv_details", side_effect=AssertionError("TMDb API не должен вызываться"))
            )

            web_export.export_watched_movies_json(data, path=path)

    assert_check("save_dataset не вызван", save_dataset_mock.call_count == 0)
    assert_check("save_candidate_pool не вызван", save_pool_mock.call_count == 0)
    assert_check("save_weights не вызван", save_weights_mock.call_count == 0)
    assert_check("save_model_metrics не вызван", save_metrics_mock.call_count == 0)
    assert_check("KP API не вызван", kp_api_mock.call_count == 0)
    assert_check("TMDb search не вызван", tmdb_search_mock.call_count == 0)
    assert_check("TMDb details не вызван", tmdb_details_mock.call_count == 0)


def setup_temp_project():
    """Готовит временный проект для тестов."""
    temp_dir = tempfile.TemporaryDirectory()
    root = Path(temp_dir.name)

    old_paths = {
        "DATA_DIR": constant.DATA_DIR,
        "FILE_NAME": constant.FILE_NAME,
        "WEIGHTS_JSON": constant.WEIGHTS_JSON,
        "CRITERIA_POOL_JSON": constant.CRITERIA_POOL_JSON,
        "CANDIDATE_POOL_JSON": constant.CANDIDATE_POOL_JSON,
        "RATING_ORDER_DRAFTS_DIR": constant.RATING_ORDER_DRAFTS_DIR,
        "MODEL_METRICS_JSON": constant.MODEL_METRICS_JSON,
        "BACKUP_DIR": constant.BACKUP_DIR,
        "DIR_META": constant.DIR_META,
        "META_JSON": constant.META_JSON,
        "DIR_TXT": constant.DIR_TXT,
        "EDIT_EXCEL": constant.EDIT_EXCEL,
    }

    constant.DATA_DIR = str(root / "data")
    constant.FILE_NAME = str(root / "data" / "dataset.json")
    constant.WEIGHTS_JSON = str(root / "data" / "weights.json")
    constant.CRITERIA_POOL_JSON = str(root / "data" / "candidate_criteria.json")
    constant.CANDIDATE_POOL_JSON = str(root / "data" / "candidate_pool.json")
    constant.RATING_ORDER_DRAFTS_DIR = str(root / "data" / "rating_order_drafts")
    constant.MODEL_METRICS_JSON = str(root / "config" / "model_metrics.json")
    constant.BACKUP_DIR = str(root / "backup") + "/"
    constant.DIR_META = str(root / "meta")
    constant.META_JSON = str(root / "meta" / "meta_data.json")
    constant.DIR_TXT = str(root / "txt")
    constant.EDIT_EXCEL = str(root / "txt" / "edit_dataset.xlsx")

    storage_data.init_dataset()
    storage_data.init_meta()
    storage_data.init_weights()
    storage_data.init_model_metrics()

    return temp_dir, old_paths


def restore_project_paths(temp_dir, old_paths: dict) -> None:
    """Возвращает реальные пути после тестов."""
    for name, value in old_paths.items():
        setattr(constant, name, value)
    temp_dir.cleanup()


def test_files_created() -> None:
    """Проверяет создание рабочих файлов."""
    print("\n1) Проверяем создание файлов")
    assert_check("Файл dataset.json открывается", Path(constant.FILE_NAME).exists())
    assert_check("Файл meta_data.json открывается", Path(constant.META_JSON).exists())
    assert_check("Файл weights.json открывается", Path(constant.WEIGHTS_JSON).exists())
    assert_check("Dataset изначально пустой", storage_data.load_dataset() == {})
    assert_check("Meta изначально пустая", storage_data.load_meta() == {})


def test_add_new_movie() -> None:
    """Проверяет добавление фильма."""
    print("\n2) Проверяем добавление новой записи")
    movie = make_movie()

    assert_check("Запись добавляется", storage_movie.add_movie(movie))

    dataset = storage_data.load_dataset()
    meta = storage_data.load_meta()

    assert_check("В dataset появилась 1 запись", len(dataset) == 1)
    assert_check("Title сохранен в dataset", dataset["Test Movie"]["main_info"]["title"] == "Test Movie")
    assert_check("Title сохранен в meta", "Test Movie" in meta)
    assert_check("В dataset нет постоянного поля features", "features" not in dataset["Test Movie"])
    assert_check("Features собраны полностью", set(model.get_features(dataset["Test Movie"])) == set(constant.FEATURES))
    assert_check("Raw-поля в dataset совпадают с meta", dataset["Test Movie"]["raw_scores"] == meta["Test Movie"]["raw_scores"])


def test_meta_overrides_raw_scores() -> None:
    """Проверяет приоритет raw-данных из meta."""
    print("\n3) Проверяем приоритет meta")
    storage_data.clean_dataset()

    first_movie = make_movie(title="Known Movie", raw_score=9.0)
    assert_check("Первая запись создает meta", storage_movie.add_movie(first_movie))
    storage_data.clean_dataset()

    second_movie = make_movie(title="Known Movie", raw_score=1.0)
    changed_tag = constant.TAGS_VIBE[0] if constant.TAGS_VIBE else None
    if changed_tag is not None:
        second_movie[constant.TAGS_VIBE_SECTION][changed_tag] = 1

    assert_check("Повторная запись с тем же title добавляется в пустой dataset", storage_movie.add_movie(second_movie))

    saved_movie = storage_data.load_dataset()["Known Movie"]
    assert_check("Raw kp_score взят из meta, а не из нового ввода", saved_movie["raw_scores"]["kp_score"] == 9.0)
    if changed_tag is not None:
        assert_check("Tag взят из нового ввода", saved_movie[constant.TAGS_VIBE_SECTION][changed_tag] == 1)


def test_duplicate_rejected() -> None:
    """Проверяет запрет дублей."""
    print("\n4) Проверяем запрет дублей")
    storage_data.clean_dataset()
    movie = make_movie(title="Duplicate Movie")

    assert_check("Первая запись добавляется", storage_movie.add_movie(movie))

    with contextlib.redirect_stdout(io.StringIO()):
        second_result = storage_movie.add_movie(movie)

    assert_check("Повторная запись не добавляется", second_result.ok is False)
    assert_check("Причина отказа - дубль title", second_result.reason == "duplicate_title")
    assert_check("В dataset осталась 1 запись", len(storage_data.load_dataset()) == 1)


def test_feature_formatting() -> None:
    """Проверяет расчет признаков."""
    print("\n5) Проверяем расчет признаков")
    raw_scores = make_movie()["raw_scores"]
    movie = make_movie()
    computed = format_score.raw_to_struct(raw_scores, movie["main_info"])
    movie["computed_scores"] = computed

    assert_check("Computed содержит все ожидаемые ключи", set(computed) == set(constant.COMPUTED_SCORES))
    movie_features = model.get_features(movie)
    assert_check("Bias входит в признаки модели", movie_features[constant.BIAS_FEATURE] == 1.0)
    assert_check("kp_score переносится без изменений", computed["kp_score"] == raw_scores["kp_score"])
    assert_check("imdb_score переносится без изменений", computed["imdb_score"] == raw_scores["imdb_score"])
    assert_check("imdb_popularity находится от 0 до 10", 0 <= computed["imdb_popularity"] <= 10)
    assert_check("CSV-схема содержит genre-поля", all(feature in constant.CSV_FIELDS for feature in constant.GENRE))


def test_excel_row_supports_genre() -> None:
    """Проверяет, что Excel-строка включает genre-поля."""
    print("\n5.1) Проверяем Excel-схему с жанрами")
    movie = make_movie()
    row = excel_work.build_row(movie)

    assert_check("Число значений в строке совпадает с CSV-схемой", len(row) == len(constant.CSV_FIELDS))
    if len(constant.GENRE) > 0:
        genre_feature = constant.GENRE[0]
        genre_index = constant.CSV_FIELDS.index(genre_feature)
        assert_check("Genre-поле попадает в Excel-строку", row[genre_index] == movie[constant.GENRE_SECTION][genre_feature])


def test_excel_title_guard() -> None:
    """Проверяет, что Excel не может добавлять, удалять или переименовывать записи."""
    print("\n5.2) Проверяем защиту title при Excel-импорте")
    dataset = {
        "Excel A": make_movie("Excel A"),
        "Excel B": make_movie("Excel B"),
    }
    same_titles = [make_movie("Excel A"), make_movie("Excel B")]
    renamed_title = [make_movie("Excel A renamed"), make_movie("Excel B")]
    missing_title = [make_movie("Excel A")]

    assert_check("Excel с тем же набором title проходит", excel_work.validate_excel_titles(same_titles, dataset))
    assert_check(
        "Excel с переименованной записью останавливается",
        excel_work.validate_excel_titles(renamed_title, dataset) is False,
    )
    assert_check(
        "Excel с удалённой строкой останавливается",
        excel_work.validate_excel_titles(missing_title, dataset) is False,
    )


def test_excel_export_overwrite_refreshes_dataset() -> None:
    """Проверяет, что повторный export пересоздаёт Excel по актуальному dataset."""
    print("\n5.3) Проверяем пересоздание Excel по актуальному dataset")
    from openpyxl import load_workbook

    storage_data.save_dataset({
        "Old Title": make_movie("Old Title"),
    })
    assert_check("Первый export Excel успешен", excel_work.export_dataset_to_excel(overwrite=True))

    storage_data.save_dataset({
        "Old Title": make_movie("Old Title"),
        "Кухня": make_movie("Кухня"),
    })
    assert_check("Повторный export Excel успешен", excel_work.export_dataset_to_excel(overwrite=True))

    workbook = load_workbook(constant.EDIT_EXCEL, data_only=True, read_only=True)
    try:
        worksheet = workbook.active
        titles = [
            str(row[0]).strip()
            for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True)
            if row[0] is not None
        ]
    finally:
        workbook.close()

    assert_check("Новая запись появляется в Excel после переэкспорта", "Кухня" in titles)
    assert_check("Excel содержит актуальный набор title", set(titles) == {"Old Title", "Кухня"})


def test_validation() -> None:
    """Проверяет валидаторы."""
    print("\n6) Проверяем валидацию")
    assert_check("Корректный title проходит", valid.is_correct_title("Good Title"))
    assert_check("Пустой title не проходит", valid.is_correct_title("") is False)
    assert_check("Оценка 8.5 проходит", valid.is_correct_score("8.5"))
    assert_check("Оценка 11 не проходит", valid.is_correct_score("11") is False)
    assert_check("Год 2024 проходит", valid.is_correct_year("2024"))
    assert_check("Год 1999 не проходит", valid.is_correct_year("1999") is False)
    assert_check("100 голосов проходит", valid.is_correct_votes("100"))
    assert_check("-1 голосов не проходит", valid.is_correct_votes("-1") is False)
    assert_check("Нечисловой пункт меню не проходит", valid.is_correct_select_menu(6, "abc") is False)
    assert_check("Пункт меню 0 проходит", valid.is_correct_select_menu(6, "0"))
    assert_check("Пункт меню 7 не проходит при max=6", valid.is_correct_select_menu(6, "7") is False)


def test_title_punctuation_validation() -> None:
    """Проверяет, что названия с обычной пунктуацией проходят validator."""
    print("\n6.1) Проверяем title validator с пунктуацией")
    assert_check("Ева, рожай! проходит", valid.is_correct_title("Ева, рожай!"))
    assert_check("Слово пацана. Кровь на асфальте проходит", valid.is_correct_title("Слово пацана. Кровь на асфальте"))
    assert_check("Что? Где? Когда? проходит", valid.is_correct_title("Что? Где? Когда?"))
    assert_check("Метод-2 проходит", valid.is_correct_title("Метод-2"))
    assert_check("Пищеблок: Экстра проходит", valid.is_correct_title("Пищеблок: Экстра"))
    assert_check("Строка из пробелов не проходит", valid.is_correct_title("   ") is False)
    assert_check("Управляющий символ не проходит", valid.is_correct_title("Bad\nTitle") is False)


def test_scores_menu_structure_and_safe_update() -> None:
    """Проверяет UX меню оценок и безопасное обновление user_score."""
    print("\n6.2) Проверяем меню оценок dataset")
    from ui.console import ui as ui_module

    storage_data.save_dataset({
        "High": make_movie("High", user_score=9.0, raw_score=9.0),
        "Low": make_movie("Low", user_score=4.0, raw_score=4.0),
        "Middle": make_movie("Middle", user_score=6.5, raw_score=6.5),
    })

    data_menu_output = io.StringIO()
    with contextlib.redirect_stdout(data_menu_output):
        ui_module.show_data_menu(3, 0)
    train_menu_output = io.StringIO()
    with contextlib.redirect_stdout(train_menu_output):
        ui_module.show_train_menu(3, 0)

    assert_check("Пункт уточнения порядка находится в данных", "Уточнить порядок оценок" in data_menu_output.getvalue())
    assert_check("Попарное сравнение убрано из обучения", "Попарное сравнение оценок" not in train_menu_output.getvalue())

    before_linear = storage_data.load_dataset()
    scores_output = io.StringIO()
    with patch("builtins.input", side_effect=["8"]):
        with contextlib.redirect_stdout(scores_output):
            interface_funcs.show_all_movies()
    output_text = scores_output.getvalue()

    assert_check("Просмотр отсортирован по user_score снизу вверх", output_text.index("Low") < output_text.index("Middle") < output_text.index("High"))
    assert_check("После просмотра есть пункт линейного распределения", "Линейное распределение оценок" in output_text)
    assert_check("После просмотра есть пункт изменения user_score", "Изменить оценку user_score" in output_text)
    assert_check("После просмотра есть пункт изменения названия", "Изменить название" in output_text)
    assert_check("Выход из подменю не меняет dataset", storage_data.load_dataset() == before_linear)

    with patch("builtins.input", side_effect=["6", "1", "5.5"]):
        with contextlib.redirect_stdout(io.StringIO()):
            interface_funcs.show_all_movies()
    updated = storage_data.load_dataset()
    assert_check("Изменение user_score проходит через update-service", updated["Low"]["main_info"]["user_score"] == 5.5)


def test_linear_distribution_draft() -> None:
    """Проверяет draft линейного распределения оценок без применения к dataset."""
    print("\n6.3) Проверяем draft линейного распределения оценок")

    rows = [
        {"title": "A", "score": 5.0},
        {"title": "B", "score": 7.0},
        {"title": "C", "score": 10.0},
    ]
    proposed = interface_funcs.build_linear_distribution_items(rows)
    assert_check("3 оценки распределяются линейно", [item["proposed_score"] for item in proposed] == [5.0, 7.5, 10.0])
    assert_check("Delta считается от old_score", [item["delta"] for item in proposed] == [0.0, 0.5, 0.0])

    single = interface_funcs.build_linear_distribution_items([{"title": "Only", "score": 8.0}])
    assert_check("Одна запись не падает", len(single) == 1)
    assert_check("Одна запись сохраняет old_score", single[0]["proposed_score"] == 8.0)

    storage_data.save_dataset({
        "A": make_movie("A", user_score=5.0, raw_score=5.0),
        "B": make_movie("B", user_score=7.0, raw_score=7.0),
        "C": make_movie("C", user_score=10.0, raw_score=10.0),
    })
    before_dataset = storage_data.load_dataset()
    draft_output = io.StringIO()
    with patch("builtins.input", side_effect=["5"]):
        with contextlib.redirect_stdout(draft_output):
            interface_funcs.show_all_movies()

    draft_files = sorted(Path(constant.RATING_ORDER_DRAFTS_DIR).glob("rating_order_draft_*.json"))
    assert_check("Draft JSON создан во временной папке", len(draft_files) == 1)
    draft = json.loads(draft_files[0].read_text(encoding="utf-8"))
    assert_check("Draft содержит метод linear_distribution", draft["method"] == "linear_distribution")
    assert_check("Draft содержит proposed_score", draft["items"][1]["proposed_score"] == 7.5)
    assert_check("Dataset не меняется при создании draft", storage_data.load_dataset() == before_dataset)
    assert_check("Preview показывает top изменений", "Top-10 изменений по модулю delta" in draft_output.getvalue())


def write_rating_order_draft(name: str, items: list[dict]) -> Path:
    """Создаёт draft распределения оценок для тестов."""
    draft_dir = Path(constant.RATING_ORDER_DRAFTS_DIR)
    draft_dir.mkdir(parents=True, exist_ok=True)
    draft_path = draft_dir / name
    draft = {
        "created_at": "2026-06-19T12:00:00",
        "method": "linear_distribution",
        "min_score": min(item["old_score"] for item in items),
        "max_score": max(item["old_score"] for item in items),
        "count": len(items),
        "items": items,
    }
    draft_path.write_text(json.dumps(draft, ensure_ascii=False, indent=4), encoding="utf-8")
    return draft_path


def clear_rating_order_drafts() -> None:
    """Очищает временные draft-файлы тестового проекта."""
    draft_dir = Path(constant.RATING_ORDER_DRAFTS_DIR)
    if draft_dir.exists() is False:
        return
    for draft_file in draft_dir.glob("rating_order_draft_*.json"):
        draft_file.unlink()


def test_apply_rating_order_draft() -> None:
    """Проверяет безопасное применение draft распределения оценок."""
    print("\n6.4) Проверяем применение draft распределения оценок")

    storage_data.save_dataset({
        "A": make_movie("A", user_score=5.0, raw_score=5.0),
        "B": make_movie("B", user_score=7.0, raw_score=7.0),
        "C": make_movie("C", user_score=10.0, raw_score=10.0),
    })
    clear_rating_order_drafts()
    storage_data.save_model_metrics({"loo_mae": 0.4321})
    before_weights = storage_data.load_weights()
    before_metrics = storage_data.load_model_metrics()
    write_rating_order_draft(
        "rating_order_draft_2026-06-19_12-00-00.json",
        [
            {"position": 1, "title": "A", "old_score": 5.0, "proposed_score": 5.0, "delta": 0.0},
            {"position": 2, "title": "B", "old_score": 7.0, "proposed_score": 7.5, "delta": 0.5},
            {"position": 3, "title": "C", "old_score": 10.0, "proposed_score": 10.0, "delta": 0.0},
        ],
    )

    with patch("builtins.input", side_effect=["1"]):
        with patch("ui.console.interface_funcs.calculate_rating_order_loo_mae", side_effect=[1.0, 0.9]) as loo_mock:
            with patch("ui.console.interface_funcs.update_dataset_record", wraps=interface_funcs.update_dataset_record) as update_mock:
                with contextlib.redirect_stdout(io.StringIO()):
                    applied = interface_funcs.apply_rating_order_draft_flow(input_func=lambda _text: "y")

    updated = storage_data.load_dataset()
    assert_check("Draft с совпадающим old_score применяется", applied is True)
    assert_check("update_dataset_record вызван для изменённой оценки", update_mock.call_count == 1)
    assert_check("LOO comparison вызывается для текущего и draft dataset", loo_mock.call_count == 2)
    assert_check("user_score обновлён после применения", updated["B"]["main_info"]["user_score"] == 7.5)
    assert_check("Веса не меняются при применении draft", storage_data.load_weights() == before_weights)
    after_metrics = storage_data.load_model_metrics()
    assert_check("model_metrics сохраняет прежний LOO при применении draft", after_metrics["loo_mae"] == before_metrics["loo_mae"])
    assert_check("model_metrics помечается устаревшим при изменении user_score", after_metrics["is_stale"] is True)
    assert_check("model_metrics хранит причину устаревания", after_metrics["stale_reason"] == "user_score_changed")


def test_apply_rating_order_draft_stops_on_stale_or_missing_data() -> None:
    """Проверяет остановку применения draft при рассинхроне с dataset."""
    print("\n6.5) Проверяем защиту применения draft")

    storage_data.save_dataset({
        "A": make_movie("A", user_score=5.0, raw_score=5.0),
        "B": make_movie("B", user_score=7.0, raw_score=7.0),
    })
    clear_rating_order_drafts()
    stale = {
        "method": "linear_distribution",
        "items": [
            {"title": "A", "old_score": 4.0, "proposed_score": 5.0},
        ],
    }
    missing = {
        "method": "linear_distribution",
        "items": [
            {"title": "Missing", "old_score": 7.0, "proposed_score": 8.0},
        ],
    }
    ok_stale, message_stale, _ = interface_funcs.validate_rating_order_draft(stale, storage_data.load_dataset())
    ok_missing, message_missing, _ = interface_funcs.validate_rating_order_draft(missing, storage_data.load_dataset())

    assert_check("Draft с несовпадающим old_score останавливается", ok_stale is False)
    assert_check("Stale draft сообщает о смене dataset", message_stale == "Dataset изменился после создания draft. Создайте новый draft.")
    assert_check("Draft с отсутствующим title останавливается", ok_missing is False)
    assert_check("Missing title сообщает об отсутствующей записи", "отсутствует запись" in message_missing)


def test_apply_rating_order_draft_cancel_keeps_dataset() -> None:
    """Проверяет, что отмена подтверждения не меняет dataset."""
    print("\n6.6) Проверяем отмену применения draft")

    storage_data.save_dataset({
        "A": make_movie("A", user_score=5.0, raw_score=5.0),
        "B": make_movie("B", user_score=7.0, raw_score=7.0),
    })
    clear_rating_order_drafts()
    before_dataset = storage_data.load_dataset()
    write_rating_order_draft(
        "rating_order_draft_2026-06-19_13-00-00.json",
        [
            {"position": 1, "title": "A", "old_score": 5.0, "proposed_score": 5.0, "delta": 0.0},
            {"position": 2, "title": "B", "old_score": 7.0, "proposed_score": 8.0, "delta": 1.0},
        ],
    )

    with patch("builtins.input", side_effect=["1"]):
        with patch("ui.console.interface_funcs.calculate_rating_order_loo_mae", side_effect=[1.0, 1.1]):
            with contextlib.redirect_stdout(io.StringIO()):
                applied = interface_funcs.apply_rating_order_draft_flow(input_func=lambda _text: "")

    assert_check("При отмене draft не применяется", applied is False)
    assert_check("Dataset не меняется при отмене", storage_data.load_dataset() == before_dataset)


def test_tag_compatibility() -> None:
    """Проверяет совместимость тегов."""
    expected_tags = constant.TAGS_VIBE
    old_tags = {
        "has_crime": 1,
        "has_psyhology": 1,
        "has_comedy": 0,
        "has_mystic": 1,
        "has_romantic_tension": 1,
    }
    normalized = storage_normalize.normalize_tags_vibe(old_tags)

    assert_check("Tag order matches the new scheme", constant.TAGS_VIBE == expected_tags)
    assert_check("Legacy tags are removed after migration", list(normalized) == expected_tags)
    assert_check("Missing active tags default to zero", all(feature in normalized for feature in constant.TAGS_VIBE))
    if constant.TAGS_VIBE:
        changed_tag = constant.TAGS_VIBE[0]
        assert_check("Invalid binary tag is rejected", storage_normalize.is_valid_tags_vibe({**normalized, changed_tag: 2}) is False)


def test_training() -> None:
    """Проверяет обучение модели."""
    print("\n7) Проверяем обучение модели")
    if linear_regression_train.is_method_available("mae_scipy") is False:
        print("SKIP: scipy minimize (MAE) недоступен в текущем окружении.")
        return
    storage_data.clean_dataset()

    movies = [
        make_movie("A", user_score=8.0, raw_score=8.0),
        make_movie("B", user_score=5.0, raw_score=5.0),
        make_movie("C", user_score=9.0, raw_score=9.0),
    ]
    for movie in movies:
        assert_check(f"Запись {movie['main_info']['title']} добавляется", storage_movie.add_movie(movie))

    data = storage_data.load_dataset()
    start_error = model.mean_absolute_error(data, constant.DEFAULT_WEIGHTS)
    new_weights = linear_regression_train.fit_linear_weights(
        data=data,
        method="mae_scipy",
        start_weights=constant.DEFAULT_WEIGHTS,
        alpha=0.1,
        l1_ratio=0.5,
        max_iter=500,
    )
    new_error = model.mean_absolute_error(data, new_weights)

    show_check("MAE до обучения", round(start_error, 4))
    show_check("MAE после обучения", round(new_error, 4))
    assert_check("Новые веса содержат все признаки", set(new_weights) == set(constant.FEATURES))
    assert_check("MAE не увеличилась", new_error <= start_error)


def test_noise_experiment_uses_loo_metrics() -> None:
    """Проверяет, что noise benchmark использует LOO-метрики и не вызывает обычный MAE."""
    print("\n8) Проверяем noise benchmark на LOO MAE")
    if linear_regression_train.is_method_available(linear_regression_train.BENCHMARK_METHOD) is False:
        print("SKIP: Ridge benchmark недоступен в текущем окружении.")
        return

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
        "C": make_movie("C", user_score=9.0, raw_score=8.7),
    }

    with patch("model.noise_experiment.model.mean_absolute_error", side_effect=AssertionError("MAE не должен вызываться")):
        result = noise_experiment.run_noise_experiment(
            data=data,
            weights=constant.DEFAULT_WEIGHTS.copy(),
            delta=0.5,
            runs=3,
            seed=123,
        )

    assert_check("Benchmark возвращает LOO baseline", "original_loo_mae_before" in result)
    assert_check("Benchmark возвращает средний LOO на шуме", "avg_noisy_loo_mae" in result)
    assert_check("Benchmark возвращает LOO на исходных данных после шума", "avg_original_loo_mae_after_noise_training" in result)
    assert_check("Вернулось 3 прогона benchmark", len(result["trials"]) == 3)

    progress_events = []
    grid = noise_experiment.run_noise_sensitivity_grid(
        data=data,
        weights=constant.DEFAULT_WEIGHTS.copy(),
        deltas=(0.25, 0.5),
        runs=2,
        seed=7,
        progress_callback=lambda *args: progress_events.append(args),
    )
    assert_check("Noise grid возвращает 2 delta", len(grid["results_by_delta"]) == 2)
    assert_check("Noise grid сохраняет runs", grid["runs"] == 2)
    assert_check("Noise grid пишет progress events", len(progress_events) > 0)


def test_feature_ablation_helpers() -> None:
    """Проверяет read-only helper-функции feature ablation."""
    print("\n8a) Проверяем helper-функции feature ablation")

    features = {"kp_score": 8.1, "imdb_score": "7.4"}
    features_before = copy.deepcopy(features)
    subset = feature_ablation.select_feature_subset(
        features,
        ["kp_score", "missing_feature", "imdb_score"],
    )
    assert_check("Subset выбирает нужные признаки", subset == [8.1, 0.0, 7.4])
    assert_check("Subset не мутирует source dict", features == features_before)

    values = [8.1, 7.4]
    values_before = values.copy()
    biased = feature_ablation.with_bias(values)
    assert_check("Bias добавлен первым значением", biased == [1.0, 8.1, 7.4])
    assert_check("with_bias не мутирует source list", values == values_before)

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
    }
    data_before = copy.deepcopy(data)
    x_data, y_data = feature_ablation.build_subset_xy(data, ["kp_score", "has_drama"])
    assert_check("Subset X содержит 2 строки", len(x_data) == 2)
    assert_check("Subset y содержит 2 значения", y_data == [8.0, 6.0])
    assert_check("Subset X содержит bias", all(row[0] == 1.0 for row in x_data))
    assert_check("Subset X содержит выбранные признаки", x_data[0][1:] == [8.0, 10.0])
    assert_check("build_subset_xy не мутирует dataset", data == data_before)


def test_feature_ablation_alpha_selection() -> None:
    """Проверяет выбор alpha по LOO MAE для feature ablation."""
    print("\n8b) Проверяем подбор alpha для feature ablation")

    def fake_loo_result(_data, _features, alpha=None):
        mae_by_alpha = {
            0.1: 0.7,
            1.0: 0.3,
            10.0: 0.5,
        }
        return {"mae": mae_by_alpha[alpha], "count": 3}

    with patch("model.feature_ablation.calculate_subset_ridge_loo_mae", side_effect=fake_loo_result):
        result = feature_ablation.select_best_alpha_by_loo(
            data={},
            feature_names=["kp_score"],
            alpha_grid=[0.1, 1.0, 10.0],
        )

    assert_check("Alpha selection возвращает best_alpha", result["best_alpha"] == 1.0)
    assert_check("Alpha selection возвращает best_mae", result["best_mae"] == 0.3)
    assert_check("Alpha selection возвращает alpha_results", len(result["alpha_results"]) == 3)

    def fake_tied_loo_result(_data, _features, alpha=None):
        mae_by_alpha = {
            0.1: 0.5,
            1.0: 0.4,
            10.0: 0.4,
        }
        return {"mae": mae_by_alpha[alpha], "count": 3}

    with patch("model.feature_ablation.calculate_subset_ridge_loo_mae", side_effect=fake_tied_loo_result):
        tied_result = feature_ablation.select_best_alpha_by_loo(
            data={},
            feature_names=["kp_score"],
            alpha_grid=[0.1, 1.0, 10.0],
        )

    assert_check("Tie-break выбирает больший alpha", tied_result["best_alpha"] == 10.0)


def test_feature_ablation_subset_weights() -> None:
    """Проверяет финальные диагностические веса subset-модели."""
    print("\n8c) Проверяем диагностические веса feature ablation")
    if linear_regression_train.is_method_available(linear_regression_train.BENCHMARK_METHOD) is False:
        print("SKIP: Ridge benchmark недоступен в текущем окружении.")
        return

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
        "C": make_movie("C", user_score=9.0, raw_score=8.7),
    }
    data_before = copy.deepcopy(data)
    weights = feature_ablation.fit_subset_ridge_weights(
        data,
        ["kp_score", "has_drama"],
        alpha=1.0,
    )

    assert_check("Subset weights возвращают bias", "bias" in weights)
    assert_check("Subset weights содержат только выбранные признаки", set(weights) == {"bias", "kp_score", "has_drama"})
    assert_check("Subset weights не содержат imdb_score вне subset", "imdb_score" not in weights)
    assert_check("fit_subset_ridge_weights не мутирует dataset", data == data_before)


def test_feature_ablation_error_rows() -> None:
    """Проверяет топ ошибок baseline и LOO subset-моделей."""
    print("\n8d) Проверяем строки ошибок feature ablation")

    data = {
        "A": make_movie("A", user_score=9.0, raw_score=6.0),
        "B": make_movie("B", user_score=5.0, raw_score=7.0),
        "C": make_movie("C", user_score=8.0, raw_score=8.0),
    }
    baseline = feature_ablation.calculate_imdb_baseline_mae(data)
    baseline_errors = baseline["errors"]
    required_fields = {"title", "year", "user_score", "predicted_score", "error", "variant", "contributions"}
    contribution_fields = {"feature", "value", "weight", "contribution"}

    assert_check("Baseline result содержит errors", len(baseline_errors) == 3)
    assert_check(
        "Baseline errors отсортированы по убыванию",
        [row["error"] for row in baseline_errors] == sorted([row["error"] for row in baseline_errors], reverse=True),
    )
    assert_check("Baseline error содержит обязательные поля", required_fields.issubset(baseline_errors[0]))
    assert_check("Baseline error содержит вклад raw-score", baseline_errors[0]["contributions"][0]["feature"] == "imdb_score")
    assert_check("Baseline contribution содержит обязательные поля", contribution_fields.issubset(baseline_errors[0]["contributions"][0]))

    class FakeEstimator:
        def fit(self, train_x, train_y):
            self.train_count = len(train_y)
            self.coef_ = [10.0, 2.0]
            return self

        def predict(self, rows):
            return [100.0 + self.train_count]

    with (
        patch("model.feature_ablation.linear_regression_train.is_method_available", return_value=True),
        patch("model.feature_ablation.linear_regression_train.build_estimator", return_value=FakeEstimator()),
    ):
        loo_result = feature_ablation.calculate_subset_ridge_loo_mae(
            data,
            ["kp_score"],
            alpha=1.0,
            variant="public_only_model",
        )

    model_errors = loo_result["errors"]
    assert_check("Model result содержит errors", len(model_errors) == 3)
    assert_check("Model error содержит обязательные поля", required_fields.issubset(model_errors[0]))
    assert_check(
        "Model error содержит не больше 4 вкладов",
        all(len(row["contributions"]) <= 4 for row in model_errors),
    )
    assert_check(
        "Model contribution содержит обязательные поля",
        all(contribution_fields.issubset(item) for item in model_errors[0]["contributions"]),
    )
    assert_check(
        "Model contributions отсортированы по модулю вклада",
        all(
            [abs(item["contribution"]) for item in row["contributions"]]
            == sorted([abs(item["contribution"]) for item in row["contributions"]], reverse=True)
            for row in model_errors
        ),
    )
    assert_check(
        "Model contributions используют LOO estimator weights",
        any(item["feature"] == "bias" and item["contribution"] == 10.0 for item in model_errors[0]["contributions"]),
    )
    assert_check(
        "Model errors используют LOO prediction, а не full-train prediction",
        all(row["predicted_score"] == 102.0 for row in model_errors),
    )
    assert_check(
        "Model errors отсортированы по убыванию",
        [row["error"] for row in model_errors] == sorted([row["error"] for row in model_errors], reverse=True),
    )


def test_feature_ablation_collect_report_is_read_only() -> None:
    """Проверяет сбор feature ablation report и отсутствие write-вызовов."""
    print("\n8e) Проверяем read-only feature ablation report")

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
        "C": make_movie("C", user_score=9.0, raw_score=8.7),
    }
    data_before = copy.deepcopy(data)

    with contextlib.ExitStack() as stack:
        save_weights_mock = stack.enter_context(
            patch("storage.data.save_weights", side_effect=AssertionError("save_weights не должен вызываться"))
        )
        save_metrics_mock = stack.enter_context(
            patch("storage.data.save_model_metrics", side_effect=AssertionError("save_model_metrics не должен вызываться"))
        )
        set_loo_mock = stack.enter_context(
            patch("storage.data.set_saved_loo_mae", side_effect=AssertionError("set_saved_loo_mae не должен вызываться"))
        )
        save_dataset_mock = stack.enter_context(
            patch("storage.data.save_dataset", side_effect=AssertionError("save_dataset не должен вызываться"))
        )
        save_pool_mock = stack.enter_context(
            patch("candidates.candidate_pool.save_candidate_pool", side_effect=AssertionError("save_candidate_pool не должен вызываться"))
        )
        loo_training_mock = stack.enter_context(
            patch("model.linear_regression_train.run_loo_training", side_effect=AssertionError("run_loo_training не должен вызываться"))
        )
        train_model_mock = stack.enter_context(
            patch("ui.console.train_menu.train_linear_model", side_effect=AssertionError("train_linear_model не должен вызываться"))
        )
        results = feature_ablation.collect_feature_ablation_report(data)

    assert_check("Ablation report возвращает 5 результатов", len(results) == 5)
    variants = [result["variant"] for result in results]
    assert_check(
        "Ablation variants идут в ожидаемом порядке",
        variants == [
            "imdb_baseline",
            "kp_baseline",
            "public_only_model",
            "genres_only_model",
            "public_plus_genres_model",
        ],
    )
    assert_check("Есть baseline variants", sum(result["kind"] == "baseline" for result in results) == 2)
    assert_check("Есть model variants", sum(result["kind"] == "model" for result in results) == 3)
    baseline_results = [result for result in results if result["kind"] == "baseline"]
    model_results = [result for result in results if result["kind"] == "model"]
    assert_check("Baseline results без weights", all("weights" not in result for result in baseline_results))
    assert_check("Baseline results имеют errors", all("errors" in result and len(result["errors"]) > 0 for result in baseline_results))
    assert_check("Model results имеют best_alpha", all("best_alpha" in result for result in model_results))
    assert_check("Model results имеют weights", all("weights" in result and "bias" in result["weights"] for result in model_results))
    assert_check("Model results имеют alpha_results", all(len(result.get("alpha_results", [])) > 0 for result in model_results))
    assert_check("Model results имеют errors", all("errors" in result and len(result["errors"]) > 0 for result in model_results))
    assert_check(
        "Ablation errors отсортированы по убыванию",
        all(
            [row["error"] for row in result["errors"]]
            == sorted([row["error"] for row in result["errors"]], reverse=True)
            for result in results
        ),
    )
    assert_check(
        "Ablation errors имеют обязательные поля",
        all(
            {"title", "year", "user_score", "predicted_score", "error", "contributions"}.issubset(row)
            for result in results
            for row in result["errors"]
        ),
    )
    assert_check(
        "Ablation errors имеют топ-4 вкладов",
        all(len(row["contributions"]) <= 4 for result in results for row in result["errors"]),
    )
    assert_check("Ablation report не мутирует dataset", data == data_before)
    assert_check("save_weights не вызван", save_weights_mock.call_count == 0)
    assert_check("save_model_metrics не вызван", save_metrics_mock.call_count == 0)
    assert_check("set_saved_loo_mae не вызван", set_loo_mock.call_count == 0)
    assert_check("save_dataset не вызван", save_dataset_mock.call_count == 0)
    assert_check("save_candidate_pool не вызван", save_pool_mock.call_count == 0)
    assert_check("run_loo_training не вызван", loo_training_mock.call_count == 0)
    assert_check("train_linear_model не вызван", train_model_mock.call_count == 0)

    lines = feature_ablation.format_feature_ablation_report(results)
    assert_check("Formatted report начинается с русским заголовком", lines[0] == "Отчёт диагностики признаков")
    assert_check("Formatted report выводит Alpha", "Alpha" in lines[2])
    assert_check("Formatted report выводит LOO MAE / MAE", "LOO MAE / MAE" in lines[2])
    assert_check("Formatted report выводит блоки весов", any(line.startswith("Веса:") for line in lines))
    assert_check("Formatted report выводит топ ошибок IMDb", "Топ-5 ошибок: Базовый IMDb" in lines)
    assert_check("Formatted report выводит топ ошибок KP", "Топ-5 ошибок: Базовый KP" in lines)
    assert_check("Formatted report выводит топ ошибок public", "Топ-5 ошибок: Модель только public" in lines)
    assert_check("Formatted report выводит топ ошибок жанров", "Топ-5 ошибок: Модель только жанры" in lines)
    assert_check("Formatted report выводит топ ошибок public+жанры", "Топ-5 ошибок: Public + жанры" in lines)
    assert_check("Formatted report выводит топ-4 вкладов", any("Топ-4 вклада:" in line for line in lines))
    assert_check("Formatted report поясняет диагностические веса", any("веса диагностические" in line for line in lines))
    assert_check("Formatted report содержит лучший результат", any(line.startswith("Лучший результат:") for line in lines))
    assert_check("Formatted report содержит вывод", any(line.startswith("Вывод:") for line in lines))


def test_feature_ablation_console_report() -> None:
    """Проверяет консольный показ feature ablation report без write-вызовов."""
    print("\n8f) Проверяем консольный feature ablation report")

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
    }
    fake_results = [{"variant": "imdb_baseline", "kind": "baseline", "mae": 0.5, "count": 2}]
    fake_lines = [
        "Отчёт диагностики признаков",
        "",
        "Базовый IMDb                база            2        -     0.50",
        "Веса: Модель только public",
        "Лучший результат: Базовый IMDb",
    ]

    with contextlib.ExitStack() as stack:
        collect_mock = stack.enter_context(
            patch("ui.console.interface_funcs.feature_ablation.collect_feature_ablation_report", return_value=fake_results)
        )
        format_mock = stack.enter_context(
            patch("ui.console.interface_funcs.feature_ablation.format_feature_ablation_report", return_value=fake_lines)
        )
        press_enter_mock = stack.enter_context(patch("ui.console.interface_funcs.ui.press_enter"))
        stack.enter_context(patch("ui.console.interface_funcs.ui.clean_terminal"))
        save_weights_mock = stack.enter_context(
            patch("storage.data.save_weights", side_effect=AssertionError("save_weights не должен вызываться"))
        )
        save_metrics_mock = stack.enter_context(
            patch("storage.data.save_model_metrics", side_effect=AssertionError("save_model_metrics не должен вызываться"))
        )
        set_loo_mock = stack.enter_context(
            patch("storage.data.set_saved_loo_mae", side_effect=AssertionError("set_saved_loo_mae не должен вызываться"))
        )
        save_dataset_mock = stack.enter_context(
            patch("storage.data.save_dataset", side_effect=AssertionError("save_dataset не должен вызываться"))
        )
        save_pool_mock = stack.enter_context(
            patch("candidates.candidate_pool.save_candidate_pool", side_effect=AssertionError("save_candidate_pool не должен вызываться"))
        )
        loo_training_mock = stack.enter_context(
            patch("model.linear_regression_train.run_loo_training", side_effect=AssertionError("run_loo_training не должен вызываться"))
        )
        train_model_mock = stack.enter_context(
            patch("ui.console.train_menu.train_linear_model", side_effect=AssertionError("train_linear_model не должен вызываться"))
        )

        output = io.StringIO()
        with contextlib.redirect_stdout(output):
            interface_funcs.show_feature_ablation_report(data)

    collect_mock.assert_called_once_with(data)
    format_mock.assert_called_once_with(fake_results)
    assert_check("Console report печатает русский заголовок", "Отчёт диагностики признаков" in output.getvalue())
    assert_check("Console report печатает строки отчёта", "Базовый IMDb" in output.getvalue())
    assert_check("Console report ждёт Enter", press_enter_mock.call_count == 1)
    assert_check("Console report не вызывает save_weights", save_weights_mock.call_count == 0)
    assert_check("Console report не вызывает save_model_metrics", save_metrics_mock.call_count == 0)
    assert_check("Console report не вызывает set_saved_loo_mae", set_loo_mock.call_count == 0)
    assert_check("Console report не вызывает save_dataset", save_dataset_mock.call_count == 0)
    assert_check("Console report не вызывает save_candidate_pool", save_pool_mock.call_count == 0)
    assert_check("Console report не вызывает run_loo_training", loo_training_mock.call_count == 0)
    assert_check("Console report не вызывает train_linear_model", train_model_mock.call_count == 0)


def test_genre_markup_efficiency_coverage_and_conclusion() -> None:
    """Проверяет coverage и выводы для эффективности жанровой разметки."""
    print("\n8g) Проверяем coverage жанровой эффективности")

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
        "C": make_movie("C", user_score=9.0, raw_score=8.7),
        "D": make_movie("D", user_score=None, raw_score=8.0),
    }
    for movie in data.values():
        movie[constant.GENRE_SECTION]["has_drama"] = 0
    data["A"][constant.GENRE_SECTION]["has_drama"] = 1
    data["B"][constant.GENRE_SECTION]["has_drama"] = 0.2
    data["D"][constant.GENRE_SECTION]["has_drama"] = 1
    data_before = copy.deepcopy(data)

    coverage = genre_markup_efficiency.collect_genre_coverage(data, "has_drama")
    assert_check("Coverage возвращает genre", coverage["genre"] == "has_drama")
    assert_check("Coverage считает count по > 0", coverage["count"] == 2)
    assert_check("Coverage пропускает записи без user_score", coverage["total_count"] == 3)
    assert_check("Coverage считает percent", abs(coverage["coverage_percent"] - 66.6667) < 0.01)
    assert_check("Coverage не мутирует dataset", data == data_before)

    assert_check(
        "Conclusion count < 5 = мало данных",
        genre_markup_efficiency.build_genre_efficiency_conclusion(0.5, 4) == "мало данных",
    )
    assert_check(
        "Conclusion delta > 0.01 = помогает",
        genre_markup_efficiency.build_genre_efficiency_conclusion(0.02, 5) == "помогает",
    )
    assert_check(
        "Conclusion delta < -0.01 = ухудшает",
        genre_markup_efficiency.build_genre_efficiency_conclusion(-0.02, 5) == "ухудшает",
    )
    assert_check(
        "Conclusion маленький delta = почти нет эффекта",
        genre_markup_efficiency.build_genre_efficiency_conclusion(0.005, 5) == "почти нет эффекта",
    )


def test_genre_markup_efficiency_report_is_read_only() -> None:
    """Проверяет read-only сбор и форматирование эффективности жанров."""
    print("\n8h) Проверяем read-only genre markup efficiency report")

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
        "C": make_movie("C", user_score=9.0, raw_score=8.7),
        "D": make_movie("D", user_score=7.0, raw_score=7.1),
        "E": make_movie("E", user_score=5.0, raw_score=5.5),
    }
    for movie in data.values():
        for genre_feature in feature_ablation.GENRE_FEATURES:
            movie[constant.GENRE_SECTION][genre_feature] = 0
    data["A"][constant.GENRE_SECTION]["has_drama"] = 1
    data["B"][constant.GENRE_SECTION]["has_drama"] = 1
    data["C"][constant.GENRE_SECTION]["has_crime"] = 1
    data_before = copy.deepcopy(data)

    def fake_alpha_selection(_data, features, alpha_grid=None, variant=None):
        selected_features = list(features)
        if selected_features == feature_ablation.PUBLIC_FEATURES:
            mae = 1.0
            alpha = 10.0
        else:
            genre_feature = selected_features[-1]
            mae_by_genre = {
                "has_drama": 0.8,
                "has_crime": 1.2,
            }
            mae = mae_by_genre.get(genre_feature, 0.95)
            alpha = 30.0
        return {
            "best_alpha": alpha,
            "best_mae": mae,
            "alpha_results": [{"alpha": alpha, "mae": mae}],
            "best_errors": [{
                "title": "A",
                "year": 2024,
                "user_score": 8.0,
                "predicted_score": 7.5,
                "error": 0.5,
                "variant": variant,
                "contributions": [],
            }],
        }

    def fake_weights(_data, features, alpha):
        weights = {"bias": 0.1}
        for feature in features:
            weights[feature] = {
                "has_drama": 0.3,
                "has_crime": -0.2,
            }.get(feature, 0.05)
        return weights

    with contextlib.ExitStack() as stack:
        stack.enter_context(
            patch("model.genre_markup_efficiency.feature_ablation.select_best_alpha_by_loo", side_effect=fake_alpha_selection)
        )
        stack.enter_context(
            patch("model.genre_markup_efficiency.feature_ablation.fit_subset_ridge_weights", side_effect=fake_weights)
        )
        save_weights_mock = stack.enter_context(
            patch("storage.data.save_weights", side_effect=AssertionError("save_weights не должен вызываться"))
        )
        save_metrics_mock = stack.enter_context(
            patch("storage.data.save_model_metrics", side_effect=AssertionError("save_model_metrics не должен вызываться"))
        )
        set_loo_mock = stack.enter_context(
            patch("storage.data.set_saved_loo_mae", side_effect=AssertionError("set_saved_loo_mae не должен вызываться"))
        )
        save_dataset_mock = stack.enter_context(
            patch("storage.data.save_dataset", side_effect=AssertionError("save_dataset не должен вызываться"))
        )
        save_pool_mock = stack.enter_context(
            patch("candidates.candidate_pool.save_candidate_pool", side_effect=AssertionError("save_candidate_pool не должен вызываться"))
        )
        loo_training_mock = stack.enter_context(
            patch("model.linear_regression_train.run_loo_training", side_effect=AssertionError("run_loo_training не должен вызываться"))
        )
        train_model_mock = stack.enter_context(
            patch("ui.console.train_menu.train_linear_model", side_effect=AssertionError("train_linear_model не должен вызываться"))
        )

        report = genre_markup_efficiency.collect_genre_markup_efficiency_report(data)

    assert_check("Genre efficiency возвращает report_type", report["report_type"] == "genre_markup_efficiency")
    assert_check("Genre efficiency возвращает base_result", "base_result" in report)
    assert_check("Genre efficiency возвращает genre_results", len(report["genre_results"]) == len(feature_ablation.GENRE_FEATURES))

    required_fields = {
        "count",
        "coverage_percent",
        "base_loo_mae",
        "genre_loo_mae",
        "delta",
        "best_alpha",
        "genre_weight",
        "conclusion",
    }
    assert_check(
        "Genre result содержит обязательные поля",
        all(required_fields.issubset(result) for result in report["genre_results"]),
    )
    deltas = [result["delta"] for result in report["genre_results"]]
    assert_check("Genre results отсортированы по delta", deltas == sorted(deltas, reverse=True))
    assert_check("Delta считается как base - genre", abs(report["genre_results"][0]["delta"] - 0.2) < 1e-9)
    assert_check("Genre weight берётся из диагностических весов", report["genre_results"][0]["genre_weight"] == 0.3)
    assert_check("Genre efficiency не мутирует dataset", data == data_before)
    assert_check("Genre efficiency не вызывает save_weights", save_weights_mock.call_count == 0)
    assert_check("Genre efficiency не вызывает save_model_metrics", save_metrics_mock.call_count == 0)
    assert_check("Genre efficiency не вызывает set_saved_loo_mae", set_loo_mock.call_count == 0)
    assert_check("Genre efficiency не вызывает save_dataset", save_dataset_mock.call_count == 0)
    assert_check("Genre efficiency не вызывает save_candidate_pool", save_pool_mock.call_count == 0)
    assert_check("Genre efficiency не вызывает run_loo_training", loo_training_mock.call_count == 0)
    assert_check("Genre efficiency не вызывает train_linear_model", train_model_mock.call_count == 0)

    lines = genre_markup_efficiency.format_genre_markup_efficiency_report(report)
    assert_check("Genre efficiency formatter выводит заголовок", lines[0] == "Эффективность жанровой разметки")
    header = next(line for line in lines if "Жанр" in line and "Кол-во" in line)
    assert_check("Genre efficiency formatter выводит Кол-во", "Кол-во" in header)
    assert_check("Genre efficiency formatter выводит Доля", "Доля" in header)
    assert_check("Genre efficiency formatter выводит LOO MAE", "LOO MAE" in header)
    assert_check("Genre efficiency formatter выводит Delta", "Delta" in header)
    assert_check("Genre efficiency formatter выводит Вес", "Вес" in header)
    assert_check("Genre efficiency formatter выводит Вывод", "Вывод" in header)
    assert_check(
        "Genre efficiency formatter выводит диагностическое примечание",
        any("не сохраняются как рабочая модель" in line for line in lines),
    )


def test_genre_markup_efficiency_console_report() -> None:
    """Проверяет консольный показ genre markup efficiency report без write-вызовов."""
    print("\n8i) Проверяем консольный genre markup efficiency report")

    data = {
        "A": make_movie("A", user_score=8.0, raw_score=8.0),
        "B": make_movie("B", user_score=6.0, raw_score=6.2),
    }
    fake_report = {
        "report_type": "genre_markup_efficiency",
        "base_result": {"loo_mae": 0.5},
        "genre_results": [],
    }
    fake_lines = [
        "Эффективность жанровой разметки",
        "",
        "Жанр              Кол-во   Доля    Alpha   LOO MAE   Delta    Вес     Вывод",
        "Примечание: это диагностический отчёт. Веса и alpha не сохраняются как рабочая модель.",
    ]

    with contextlib.ExitStack() as stack:
        collect_mock = stack.enter_context(
            patch(
                "ui.console.interface_funcs.genre_markup_efficiency.collect_genre_markup_efficiency_report",
                return_value=fake_report,
            )
        )
        format_mock = stack.enter_context(
            patch(
                "ui.console.interface_funcs.genre_markup_efficiency.format_genre_markup_efficiency_report",
                return_value=fake_lines,
            )
        )
        press_enter_mock = stack.enter_context(patch("ui.console.interface_funcs.ui.press_enter"))
        stack.enter_context(patch("ui.console.interface_funcs.ui.clean_terminal"))
        save_weights_mock = stack.enter_context(
            patch("storage.data.save_weights", side_effect=AssertionError("save_weights не должен вызываться"))
        )
        save_metrics_mock = stack.enter_context(
            patch("storage.data.save_model_metrics", side_effect=AssertionError("save_model_metrics не должен вызываться"))
        )
        set_loo_mock = stack.enter_context(
            patch("storage.data.set_saved_loo_mae", side_effect=AssertionError("set_saved_loo_mae не должен вызываться"))
        )
        save_dataset_mock = stack.enter_context(
            patch("storage.data.save_dataset", side_effect=AssertionError("save_dataset не должен вызываться"))
        )
        save_pool_mock = stack.enter_context(
            patch("candidates.candidate_pool.save_candidate_pool", side_effect=AssertionError("save_candidate_pool не должен вызываться"))
        )
        loo_training_mock = stack.enter_context(
            patch("model.linear_regression_train.run_loo_training", side_effect=AssertionError("run_loo_training не должен вызываться"))
        )
        train_model_mock = stack.enter_context(
            patch("ui.console.train_menu.train_linear_model", side_effect=AssertionError("train_linear_model не должен вызываться"))
        )

        output = io.StringIO()
        with contextlib.redirect_stdout(output):
            interface_funcs.show_genre_markup_efficiency_report(data)

    collect_mock.assert_called_once_with(data)
    format_mock.assert_called_once_with(fake_report)
    assert_check("Genre console report печатает заголовок", "Эффективность жанровой разметки" in output.getvalue())
    assert_check("Genre console report печатает строки отчёта", "Кол-во" in output.getvalue())
    assert_check("Genre console report ждёт Enter", press_enter_mock.call_count == 1)
    assert_check("Genre console report не вызывает save_weights", save_weights_mock.call_count == 0)
    assert_check("Genre console report не вызывает save_model_metrics", save_metrics_mock.call_count == 0)
    assert_check("Genre console report не вызывает set_saved_loo_mae", set_loo_mock.call_count == 0)
    assert_check("Genre console report не вызывает save_dataset", save_dataset_mock.call_count == 0)
    assert_check("Genre console report не вызывает save_candidate_pool", save_pool_mock.call_count == 0)
    assert_check("Genre console report не вызывает run_loo_training", loo_training_mock.call_count == 0)
    assert_check("Genre console report не вызывает train_linear_model", train_model_mock.call_count == 0)


def test_biggest_error_cards_in_report() -> None:
    """Проверяет карточки «САМЫЕ БОЛЬШИЕ ОШИБКИ» и fallback описаний."""
    print("\n8j) Проверяем карточки ошибок в train_report")

    long_text = "Описание " + ("очень длинное " * 80)
    assert_check(
        "Описание режется до 500 символов",
        len(train_report._truncate_description(long_text)) <= 500
        and train_report._truncate_description(long_text).endswith("..."),
    )

    meta_description = train_report.resolve_movie_description(
        title="Test Show",
        year=2024,
        meta_obj={"description": "Meta description text"},
        pool_by_identity={},
    )
    assert_check("Описание берётся из meta", meta_description == "Meta description text")

    pool_candidate = {
        "title": "Pool Show",
        "year": 2020,
        "overview": "Pool overview text",
    }
    from candidates import keys as candidate_keys

    pool_by_identity = {candidate_keys.title_identity_key(pool_candidate): pool_candidate}
    pool_description = train_report.resolve_movie_description(
        title="Pool Show",
        year=2020,
        meta_obj={},
        pool_by_identity=pool_by_identity,
    )
    assert_check("Описание берётся из pool", pool_description == "Pool overview text")

    def fake_tmdb_cache(_tmdb_id):
        return "TMDb cached overview"

    tmdb_description = train_report.resolve_movie_description(
        title="Cached Show",
        year=2019,
        meta_obj={"tmdb_id": 12345},
        pool_by_identity={},
        tmdb_cache_reader=fake_tmdb_cache,
    )
    assert_check("Описание берётся из TMDb cache", tmdb_description == "TMDb cached overview")

    missing_description = train_report.resolve_movie_description(
        title="No Info",
        year=2018,
        meta_obj={},
        pool_by_identity={},
        tmdb_cache_reader=lambda _tmdb_id: None,
    )
    assert_check("Без источников — нет описания", missing_description == "нет описания")

    error_row = {
        "title": "Card Show",
        "user_score": 8.5,
        "prediction": 6.2,
        "error": -2.3,
        "top_impacts": [(1.45, 1.45, "kp_score"), (0.82, 0.82, "has_drama")],
    }
    card_lines = train_report.format_biggest_error_card_lines(error_row, "Short description")
    card_text = "\n".join(card_lines)
    assert_check("Карточка содержит user_score", "Card Show (8.50)" in card_text)
    assert_check("Карточка содержит predict и ошибку", "Оценка: 6.20 (ошибка: -2.30)" in card_text)
    assert_check("Карточка содержит вклады", "Вклады:" in card_text and "kp_score" in card_text)
    assert_check("Карточка содержит описание", "Short description" in card_text)

    data = {
        "A": make_movie("Alpha", user_score=9.0, raw_score=6.0),
        "B": make_movie("Beta", user_score=5.0, raw_score=7.5),
    }
    error_section = "\n".join(train_report.build_error_lines(data, constant.DEFAULT_WEIGHTS))
    assert_check("Раздел содержит блок САМЫЕ БОЛЬШИЕ ОШИБКИ", "САМЫЕ БОЛЬШИЕ ОШИБКИ" in error_section)
    assert_check("Раздел сохраняет список всех записей", "Все записи по убыванию |error|" in error_section)


def test_backup_restore() -> None:
    """Проверяет восстановление датасета из backup."""
    print("\n9) Проверяем восстановление backup")
    storage_data.clean_dataset()

    first_movie = make_movie("Backup A", user_score=7.0, raw_score=7.0)
    second_movie = make_movie("Backup B", user_score=9.0, raw_score=9.0)

    assert_check("Первая запись добавляется", storage_movie.add_movie(first_movie))
    storage_files.create_backup()
    backup_path = storage_files.get_latest_backups(1)[0]

    assert_check("Вторая запись добавляется", storage_movie.add_movie(second_movie))
    assert_check("В датасете временно 2 записи", len(storage_data.load_dataset()) == 2)

    records_count = storage_files.restore_backup(backup_path)
    data = storage_data.load_dataset()

    assert_check("Backup восстановил 1 запись", records_count == 1)
    assert_check("Первая запись осталась", "Backup A" in data)
    assert_check("Вторая запись исчезла после восстановления", "Backup B" not in data)


def test_api_fallback_to_secondary_token() -> None:
    """Проверяет переход на резервный API-ключ после HTTP 403."""
    print("\n10) Проверяем fallback на второй API-ключ")

    old_secondary = api.SECONDARY_TOKEN
    api.SECONDARY_TOKEN = "secondary-token"

    class FakeResponse:
        def __init__(self, payload: str):
            self.payload = payload.encode("utf-8")
            self.status = 200

        def read(self):
            return self.payload

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

    calls = []

    def fake_opener(request, timeout=20):
        calls.append(request.headers.get("X-api-key"))
        if len(calls) == 1:
            raise HTTPError(request.full_url, 403, "Forbidden", hdrs=None, fp=None)
        return FakeResponse('{"docs": [{"id": 1}]}')

    try:
        result = api.fetch_json(
            "https://example.test/api",
            token="primary-token",
            opener=fake_opener,
            retries=0,
        )
    finally:
        api.SECONDARY_TOKEN = old_secondary

    assert_check("Первый запрос ушёл с основным ключом", calls[0] == "primary-token")
    assert_check("Второй запрос ушёл с резервным ключом", calls[1] == "secondary-token")
    assert_check("Ответ успешно получен через резервный ключ", result["ok"] is True)


def test_sql_title_search() -> None:
    """Проверяет поиск тайтла в локальной SQLite-базе."""
    print("\n11) Проверяем поиск тайтла в SQL")

    result = sql_search.search_title_in_sql("Триггер", "Россия")

    assert_check("SQL-поиск возвращает ok", result["ok"] is True)
    assert_check("Вернулся словарь с данными", isinstance(result["data"], dict))
    assert_check("Есть название", bool(result["data"].get("title")))
    assert_check("Есть год", result["data"].get("year") is not None)
    assert_check("Есть жанры", isinstance(result["data"].get("genres"), list))
    assert_check("Есть рейтинг", result["data"].get("imdb_rating") is not None)
    assert_check("Есть голоса", result["data"].get("imdb_votes") is not None)
    assert_check("Есть режиссеры или актеры", bool(result["data"].get("credits", {}).get("directors") or result["data"].get("credits", {}).get("actors")))


def test_sql_title_search_aliases() -> None:
    """Проверяет сложные и опечатанные названия для SQL-поиска."""
    print("\n11.1) Проверяем алиасы и опечатки SQL-поиска")

    checks = [
        ("Чернобыль зона отчуждения", "Chernobyl: Zone of Exclusion"),
        ("Haappy End", "Happy End"),
        ("Индентификация", "Identification"),
        ("Фишшер", "Fisher"),
        ("Я знаю кто тебя убил", "YA znayu, kto tebya ubil"),
    ]

    for query, expected_title in checks:
        result = sql_search.search_title_in_sql(query, "Россия")
        assert_check(f"SQL-поиск находит {query}", result["ok"] is True)
        assert_check(
            f"SQL-поиск приводит {query} к ожидаемому тайтлу",
            result["data"].get("title") == expected_title
        )

    alias_result = sql_search.search_title_in_sql("Haappy End", "Россия")
    assert_check(
        "В match сохраняется информация о ручном alias",
        bool(alias_result["data"].get("match", {}).get("alias_applied"))
    )


def test_build_api_defaults_from_raw_movie() -> None:
    """Проверяет сбор defaults из сырого ответа API."""
    print("\n11.2) Проверяем build_api_defaults на сыром API-объекте")

    movie = {
        "name": "Тестовый сериал",
        "alternativeName": "Test Series",
        "year": 2025,
        "rating": {"kp": 7.4, "imdb": 6.9},
        "votes": {"kp": 12345, "imdb": 678},
        "genres": [{"name": "драма"}, {"name": "триллер"}],
    }

    defaults = title_resolve.build_api_defaults(movie)

    assert_check("Название берётся из raw API name", defaults["main_info"]["title"] == "Тестовый сериал")
    assert_check("Год берётся из raw API", defaults["main_info"]["year"] == 2025)
    assert_check("KP рейтинг извлекается из rating.kp", defaults["raw_scores"]["kp_score"] == 7.4)
    assert_check("IMDb рейтинг извлекается из rating.imdb", defaults["raw_scores"]["imdb_score"] == 6.9)
    assert_check("Жанры размечаются в defaults", sum(defaults["genre"].values()) >= 2)


def test_merge_defaults_prefers_api_and_keeps_sql() -> None:
    """Проверяет объединение SQL и API defaults."""
    print("\n11.3) Проверяем merge_defaults для SQL + API")

    sql_defaults = {
        "main_info": {"title": "Trigger", "user_score": None, "year": 2018},
        "raw_scores": {"kp_score": None, "kp_votes": None, "imdb_score": 7.7, "imdb_votes": 4321},
        "tags_vibe": {},
        "genre": {"has_drama": 1, "has_thriller": 1},
    }
    api_defaults = {
        "main_info": {"title": "Триггер", "user_score": None, "year": 2018},
        "raw_scores": {"kp_score": 8.0, "kp_votes": 55555, "imdb_score": 7.6, "imdb_votes": 4000},
        "tags_vibe": {},
        "genre": {"has_drama": 1, "has_detective": 1},
    }

    merged = title_resolve.merge_defaults(sql_defaults, api_defaults)

    assert_check("Название берётся из API как более пользовательское", merged["main_info"]["title"] == "Триггер")
    assert_check("KP приходит из API", merged["raw_scores"]["kp_score"] == 8.0)
    assert_check("IMDb остаётся заполненным", merged["raw_scores"]["imdb_score"] == 7.6)
    assert_check("Жанр из SQL сохраняется", merged["genre"]["has_thriller"] == 1)
    assert_check("Жанр из API добавляется", merged["genre"]["has_detective"] == 1)


def test_build_genre_defaults_ignores_unknown() -> None:
    """Проверяет, что неизвестные жанры не создают новые feature в обычном потоке."""
    print("\n11.4) Проверяем, что новые жанры не автодобавляются в модель")

    defaults = title_resolve.build_genre_defaults(["комедия", "future-micro-genre-x", "драма"])

    assert_check("Известный жанр комедия размечается", defaults.get("has_comedy") == 1)
    assert_check("Известный жанр драма размечается", defaults.get("has_drama") == 1)
    assert_check("Неизвестный жанр не включается в defaults", "has_future_micro_genre_x" not in defaults)


def test_build_genre_defaults_maps_mystery_to_detective() -> None:
    """Manual genre defaults map Mystery to has_detective via shared raw mapper."""
    print("\n11.4b) Проверяем Mystery -> has_detective в build_genre_defaults")

    defaults = title_resolve.build_genre_defaults(["Mystery", "детектив"])

    assert_check("Mystery -> has_detective", defaults.get("has_detective") == 1)
    assert_check("has_mystery не создаётся", "has_mystery" not in defaults)
    assert_check(
        "полный набор constant.GENRE",
        set(defaults.keys()) == set(constant.GENRE),
    )


def test_split_known_genres_uses_shared_raw_mapper() -> None:
    """split_known_genres aligns confirm UI hints with shared raw genre mapper."""
    print("\n11.4c) Проверяем split_known_genres через shared mapper")

    known, unknown = title_resolve.split_known_genres(["Mystery", "TotallyUnknownGenreXYZ", "history"])

    assert_check("Mystery считается known", "Mystery" in known)
    assert_check("unknown raw остаётся unknown", "TotallyUnknownGenreXYZ" in unknown)
    assert_check("history без has_* остаётся unknown", "history" in unknown)
    assert_check("Mystery не попадает в unknown", "Mystery" not in unknown)


def test_manual_add_defaults_when_lookup_fails() -> None:
    """Проверяет ручной fallback, если SQL/API ничего не нашли."""
    print("\n11.5) Проверяем ручной fallback добавления")
    resolved = {
        "title": "Manual Missing Title",
        "country": "Россия",
        "sql_result": {"ok": False, "details": "not_found"},
        "sql_data": None,
        "api_data": None,
        "api_error": {"ok": False, "error": "network_error", "details": "timeout"},
        "defaults": None,
        "found": False,
    }

    output = io.StringIO()
    with patch("ui.console.request.title_resolve.resolve_title_data_for_add", return_value=resolved):
        with patch("builtins.input", return_value="y"):
            with contextlib.redirect_stdout(output):
                defaults, _, _ = request_ui.resolve_title_for_training("Manual Missing Title", confirm_genres=True)

    assert_check("Fallback возвращает defaults", defaults is not None)
    assert_check("Title берётся из ручного ввода", defaults["main_info"]["title"] == "Manual Missing Title")
    assert_check("KP score остаётся пустым", defaults["raw_scores"]["kp_score"] is None)
    assert_check("Vibe defaults заполнены по схеме", set(defaults["tags_vibe"]) == set(constant.TAGS_VIBE))
    assert_check("Genre defaults заполнены по схеме", set(defaults["genre"]) == set(constant.GENRE))
    assert_check("Печатается режим ручной разметки", "Режим: ручная разметка" in output.getvalue())


def test_add_defaults_rejects_sql_api_identity_mismatch() -> None:
    """Проверяет, что SQL IMDb не подмешивается к другому API-объекту."""
    print("\n11.6) Проверяем identity gate при SQL/API mismatch")
    sql_data = {
        "title": "Mad",
        "original_title": "Mad",
        "year": 2010,
        "genres": ["Animation", "Comedy"],
        "imdb_rating": 6.0,
        "imdb_votes": 3480,
    }
    api_data = {
        "name": "Псих",
        "year": 2020,
        "rating": {"kp": 7.372, "imdb": 7.0},
        "votes": {"kp": 173943, "imdb": 584},
        "genres": [{"name": "драма"}],
    }

    built = title_resolve.build_add_defaults_by_priority("Псих", sql_data, api_data, None)
    defaults = built["defaults"]
    sql_status = title_resolve.get_sql_status(sql_data, built["sql_identity"])

    assert_check("SQL candidate отклонён", built["sql_identity"]["accepted"] is False)
    assert_check("Причина reject зафиксирована", built["sql_identity"]["reason"] == "identity_mismatch")
    assert_check("SQL статус показывает reject", sql_status == "найдено, но отклонено (identity_mismatch)")
    assert_check("Title остаётся из KP API", defaults["main_info"]["title"] == "Псих")
    assert_check("Год остаётся из KP API", defaults["main_info"]["year"] == 2020)
    assert_check("IMDb rating берётся из KP API", defaults["raw_scores"]["imdb_score"] == 7.0)
    assert_check("IMDb votes берутся из KP API", defaults["raw_scores"]["imdb_votes"] == 584)
    assert_check("Источник IMDb не SQL", built["sources"]["imdb_score"] == "kp_api")


def test_add_defaults_accepts_sql_when_imdb_id_matches() -> None:
    """Проверяет, что совпадающий imdb_id разрешает SQL-кандидата."""
    print("\n11.7) Проверяем identity gate по imdb_id")
    sql_data = {
        "tconst": "tt1234567",
        "title": "SQL Title",
        "original_title": "SQL Original",
        "year": 2020,
        "genres": ["драма"],
        "imdb_rating": 7.8,
        "imdb_votes": 1000,
    }
    api_data = {
        "externalId": {"imdb": "tt1234567"},
        "name": "API Title",
        "year": 2021,
        "rating": {"kp": 8.0, "imdb": 6.0},
        "votes": {"kp": 2000, "imdb": 300},
        "genres": [{"name": "драма"}],
    }

    built = title_resolve.build_add_defaults_by_priority("Input Title", sql_data, api_data, None)

    assert_check("SQL принят по imdb_id", built["sql_identity"]["accepted"] is True)
    assert_check("Причина принятия imdb_id_match", built["sql_identity"]["reason"] == "imdb_id_match")
    assert_check("IMDb rating может быть из SQL", built["defaults"]["raw_scores"]["imdb_score"] == 7.8)
    assert_check("Источник IMDb SQL", built["sources"]["imdb_score"] == "imdb_sql")


def test_add_defaults_accepts_sql_when_title_year_match_without_imdb_id() -> None:
    """Проверяет fallback identity по похожему названию и году."""
    print("\n11.8) Проверяем identity gate по title/year без imdb_id")
    sql_data = {
        "title": "Псих",
        "original_title": "Псих",
        "year": 2020,
        "genres": ["драма"],
        "imdb_rating": 7.1,
        "imdb_votes": 600,
    }
    api_data = {
        "name": "Псих",
        "year": 2020,
        "rating": {"kp": 7.3, "imdb": 7.0},
        "votes": {"kp": 1000, "imdb": 500},
        "genres": [{"name": "драма"}],
    }

    built = title_resolve.build_add_defaults_by_priority("Псих", sql_data, api_data, None)

    assert_check("SQL принят по title/year", built["sql_identity"]["accepted"] is True)
    assert_check("Причина принятия title_year_match", built["sql_identity"]["reason"] == "title_year_match")
    assert_check("IMDb rating остаётся из SQL", built["defaults"]["raw_scores"]["imdb_score"] == 7.1)


def test_add_defaults_rejects_sql_when_year_differs_more_than_one() -> None:
    """Проверяет reject при несовпадении года больше чем на один."""
    print("\n11.9) Проверяем identity gate по year_mismatch")
    sql_data = {
        "title": "Псих",
        "original_title": "Псих",
        "year": 2010,
        "genres": ["драма"],
        "imdb_rating": 6.0,
        "imdb_votes": 3480,
    }
    api_data = {
        "name": "Псих",
        "year": 2020,
        "rating": {"kp": 7.3, "imdb": 7.0},
        "votes": {"kp": 1000, "imdb": 500},
        "genres": [{"name": "драма"}],
    }

    built = title_resolve.build_add_defaults_by_priority("Псих", sql_data, api_data, None)

    assert_check("SQL отклонён по году", built["sql_identity"]["accepted"] is False)
    assert_check("Причина reject year_mismatch", built["sql_identity"]["reason"] == "year_mismatch")
    assert_check("IMDb rating берётся из API после reject", built["defaults"]["raw_scores"]["imdb_score"] == 7.0)


def test_add_defaults_keeps_sql_only_flow() -> None:
    """Проверяет, что SQL-only сценарий не сломан при отсутствии API."""
    print("\n11.10) Проверяем SQL-only add-flow")
    sql_data = {
        "title": "Mad",
        "original_title": "Mad",
        "year": 2010,
        "genres": ["Animation", "Comedy"],
        "imdb_rating": 6.0,
        "imdb_votes": 3480,
    }

    built = title_resolve.build_add_defaults_by_priority("Mad", sql_data, None, None)
    defaults = built["defaults"]

    assert_check("SQL-only кандидат принят", built["sql_identity"]["accepted"] is True)
    assert_check("Причина SQL-only зафиксирована", built["sql_identity"]["reason"] == "sql_only")
    assert_check("IMDb rating приходит из SQL-only", defaults["raw_scores"]["imdb_score"] == 6.0)
    assert_check("Источник IMDb SQL-only", built["sources"]["imdb_score"] == "imdb_sql")


def test_add_record_country_selection() -> None:
    """Проверяет выбор страны по номеру при добавлении записи."""
    print("\n11.10.1) Проверяем выбор страны при добавлении записи")

    assert_check(
        "Пустой ввод страны = Россия",
        tmdb_country_options.parse_single_country_index("", 26) == 1,
    )
    assert_check(
        "Номер 2 = США",
        tmdb_country_options.choose_single_country_label(input_func=lambda _prompt: "2") == "США",
    )
    assert_check(
        "Enter по умолчанию = Россия",
        tmdb_country_options.choose_single_country_label(input_func=lambda _prompt: "") == "Россия",
    )
    assert_check(
        "0 = KP без фильтра страны",
        tmdb_country_options.choose_single_country_label(input_func=lambda _prompt: "0") == "",
    )

    with patch("ui.console.request.loop_input", return_value="Euphoria"):
        with patch("ui.console.request.tmdb_country_options.choose_single_country_label", return_value="США"):
            with patch("ui.console.request.resolve_title_for_training", return_value=({"title": "Euphoria"}, {}, {})) as resolve_training:
                result = request_ui.request_api_defaults()

    resolve_training.assert_called_once_with("Euphoria", "США", False)
    assert_check("request_api_defaults передаёт выбранную страну", result == ({"title": "Euphoria"}, {}, {}))


def test_add_resolver_second_pass_sql_after_identity_mismatch() -> None:
    """Проверяет second-pass SQL после first-pass mismatch, если API не дал IMDb."""
    print("\n11.11) Проверяем second-pass SQL после identity mismatch")
    first_sql = {
        "title": "Mad",
        "original_title": "Mad",
        "year": 2010,
        "genres": ["Animation", "Comedy"],
        "imdb_rating": 6.0,
        "imdb_votes": 3480,
    }
    second_sql = {
        "title": "Псих",
        "original_title": "Псих",
        "year": 2020,
        "genres": ["драма"],
        "imdb_rating": 7.0,
        "imdb_votes": 584,
    }
    kp_data = {
        "name": "Псих",
        "year": 2020,
        "rating": {"kp": 7.372},
        "votes": {"kp": 173943},
        "genres": [{"name": "драма"}],
    }

    with patch(
        "dataset.title_resolve.sql_search.search_title_in_sql",
        side_effect=[{"ok": True, "data": first_sql}, {"ok": True, "data": second_sql}],
    ) as mocked_sql:
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": True, "data": kp_data}):
            resolved = title_resolve.resolve_title_data_for_add("Псих")

    defaults = resolved["defaults"]
    assert_check("First-pass SQL отклонён", resolved["sql_identity"]["accepted"] is False)
    assert_check("Second-pass SQL принят", resolved["sql_second_pass_identity"]["accepted"] is True)
    assert_check("Second-pass вызвал SQL второй раз", mocked_sql.call_count == 2)
    assert_check("Final IMDb rating из second-pass SQL", defaults["raw_scores"]["imdb_score"] == 7.0)
    assert_check("Final IMDb votes из second-pass SQL", defaults["raw_scores"]["imdb_votes"] == 584)
    assert_check("Источник IMDb second-pass", resolved["sources"]["imdb_score"] == "imdb_sql_second_pass")
    assert_check("Статус second-pass принят", resolved["statuses"]["sql_second_pass"] == "найдено и принято")


def test_add_resolver_skips_second_pass_when_api_has_imdb() -> None:
    """Проверяет, что second-pass не нужен, если KP API уже дал IMDb."""
    print("\n11.12) Проверяем skip second-pass при IMDb из API")
    first_sql = {
        "title": "Mad",
        "original_title": "Mad",
        "year": 2010,
        "genres": ["Animation", "Comedy"],
        "imdb_rating": 6.0,
        "imdb_votes": 3480,
    }
    kp_data = {
        "name": "Псих",
        "year": 2020,
        "rating": {"kp": 7.372, "imdb": 7.0},
        "votes": {"kp": 173943, "imdb": 584},
        "genres": [{"name": "драма"}],
    }

    with patch("dataset.title_resolve.sql_search.search_title_in_sql", return_value={"ok": True, "data": first_sql}) as mocked_sql:
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": True, "data": kp_data}):
            resolved = title_resolve.resolve_title_data_for_add("Псих")

    defaults = resolved["defaults"]
    assert_check("First-pass SQL отклонён", resolved["sql_identity"]["accepted"] is False)
    assert_check("Second-pass SQL не вызывается", mocked_sql.call_count == 1)
    assert_check("IMDb rating берётся из API", defaults["raw_scores"]["imdb_score"] == 7.0)
    assert_check("Источник IMDb KP API", resolved["sources"]["imdb_score"] == "kp_api")
    assert_check(
        "Статус second-pass сообщает, что он не нужен",
        resolved["statuses"]["sql_second_pass"] == "не требуется, IMDb взят из KP API",
    )


def test_add_resolver_rejects_second_pass_sql_mismatch() -> None:
    """Проверяет, что second-pass SQL тоже проходит identity gate."""
    print("\n11.13) Проверяем reject second-pass SQL mismatch")
    first_sql = {
        "title": "Mad",
        "original_title": "Mad",
        "year": 2010,
        "genres": ["Animation", "Comedy"],
        "imdb_rating": 6.0,
        "imdb_votes": 3480,
    }
    second_sql = {
        "title": "Bad",
        "original_title": "Bad",
        "year": 2011,
        "genres": ["Comedy"],
        "imdb_rating": 5.5,
        "imdb_votes": 100,
    }
    kp_data = {
        "name": "Псих",
        "year": 2020,
        "rating": {"kp": 7.372},
        "votes": {"kp": 173943},
        "genres": [{"name": "драма"}],
    }

    with patch(
        "dataset.title_resolve.sql_search.search_title_in_sql",
        side_effect=[{"ok": True, "data": first_sql}, {"ok": True, "data": second_sql}],
    ):
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": True, "data": kp_data}):
            resolved = title_resolve.resolve_title_data_for_add("Псих")

    defaults = resolved["defaults"]
    assert_check("First-pass SQL отклонён", resolved["sql_identity"]["accepted"] is False)
    assert_check("Second-pass SQL тоже отклонён", resolved["sql_second_pass_identity"]["accepted"] is False)
    assert_check("Чужой second-pass IMDb не используется", defaults["raw_scores"]["imdb_score"] is None)
    assert_check("Источник IMDb остаётся пустым", resolved["sources"]["imdb_score"] is None)
    assert_check("Статус second-pass показывает reject", "отклонено" in resolved["statuses"]["sql_second_pass"])


def test_add_resolver_accepted_first_pass_skips_second_pass() -> None:
    """Проверяет, что accepted first-pass не запускает second-pass."""
    print("\n11.14) Проверяем accepted first-pass без second-pass")
    first_sql = {
        "tconst": "tt1234567",
        "title": "Псих",
        "original_title": "Псих",
        "year": 2020,
        "genres": ["драма"],
        "imdb_rating": 7.0,
        "imdb_votes": 584,
    }
    kp_data = {
        "externalId": {"imdb": "tt1234567"},
        "name": "Псих",
        "year": 2020,
        "rating": {"kp": 7.372},
        "votes": {"kp": 173943},
        "genres": [{"name": "драма"}],
    }

    with patch("dataset.title_resolve.sql_search.search_title_in_sql", return_value={"ok": True, "data": first_sql}) as mocked_sql:
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": True, "data": kp_data}):
            resolved = title_resolve.resolve_title_data_for_add("Псих")

    assert_check("First-pass SQL принят", resolved["sql_identity"]["accepted"] is True)
    assert_check("Second-pass не вызывается", mocked_sql.call_count == 1)
    assert_check("Статуса second-pass нет", "sql_second_pass" not in resolved["statuses"])
    assert_check("Источник IMDb first-pass", resolved["sources"]["imdb_score"] == "imdb_sql")


def test_add_resolver_sql_only_skips_second_pass() -> None:
    """Проверяет, что SQL-only при отсутствии API не запускает second-pass."""
    print("\n11.15) Проверяем SQL-only resolver без second-pass")
    sql_data = {
        "title": "Mad",
        "original_title": "Mad",
        "year": 2010,
        "genres": ["Animation", "Comedy"],
        "imdb_rating": 6.0,
        "imdb_votes": 3480,
    }

    with patch("dataset.title_resolve.sql_search.search_title_in_sql", return_value={"ok": True, "data": sql_data}) as mocked_sql:
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": False, "error": "not_found", "details": "not found"}):
            with patch("dataset.title_resolve.api_tmdb.search_tv_by_name", return_value=[]):
                resolved = title_resolve.resolve_title_data_for_add("Mad")

    assert_check("SQL-only принят", resolved["sql_identity"]["accepted"] is True)
    assert_check("Second-pass не вызывается без API candidate", mocked_sql.call_count == 1)
    assert_check("Статуса second-pass нет в SQL-only", "sql_second_pass" not in resolved["statuses"])
    assert_check("IMDb остаётся из SQL-only", resolved["sources"]["imdb_score"] == "imdb_sql")


def test_add_resolver_prioritizes_sql_and_kp() -> None:
    """Проверяет приоритет SQL для IMDb и KP API для KP/жанров/описания."""
    print("\n11.16) Проверяем приоритеты источников SQL + KP API")
    sql_data = {
        "tconst": "tt7654321",
        "title": "SQL Title",
        "original_title": "SQL Original",
        "year": 2022,
        "genres": ["триллер"],
        "imdb_rating": 7.7,
        "imdb_votes": 12345,
    }
    kp_data = {
        "externalId": {"imdb": "tt7654321"},
        "name": "KP Title",
        "year": 2024,
        "rating": {"kp": 8.1, "imdb": 6.1},
        "votes": {"kp": 5000, "imdb": 100},
        "genres": [{"name": "драма"}],
        "description": "KP description",
    }

    with patch("dataset.title_resolve.sql_search.search_title_in_sql", return_value={"ok": True, "data": sql_data}):
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": True, "data": kp_data}):
            resolved = title_resolve.resolve_title_data_for_add("Input Title")

    defaults = resolved["defaults"]
    assert_check("Title берётся из KP API", defaults["main_info"]["title"] == "KP Title")
    assert_check("Год берётся из KP API", defaults["main_info"]["year"] == 2024)
    assert_check("IMDb rating берётся из SQL", defaults["raw_scores"]["imdb_score"] == 7.7)
    assert_check("KP rating берётся из KP API", defaults["raw_scores"]["kp_score"] == 8.1)
    assert_check("Источник IMDb зафиксирован как SQL", resolved["sources"]["imdb_score"] == "imdb_sql")
    assert_check("Источник жанров зафиксирован как KP API", resolved["sources"]["genres"] == "kp_api")
    assert_check("TMDb не вызывается при успешном KP", resolved["statuses"]["tmdb_api"] == "не найдено")


def test_add_resolver_uses_tmdb_when_kp_fails() -> None:
    """Проверяет fallback на TMDb для жанров/описания, без подстановки KP/IMDb оценок."""
    print("\n11.17) Проверяем fallback на TMDb при падении KP API")
    sql_data = {
        "tconst": "tt2468135",
        "title": "SQL Title",
        "original_title": "SQL Original",
        "year": 2021,
        "genres": ["триллер"],
        "imdb_rating": 7.5,
        "imdb_votes": 2222,
    }
    tmdb_details = {
        "id": 10,
        "name": "TMDb Title",
        "original_name": "TMDb Original",
        "first_air_date": "2023-01-01",
        "genres": [{"name": "драма"}],
        "overview": "TMDb overview",
        "external_ids": {"imdb_id": "tt2468135"},
        "credits": {},
    }

    with patch("dataset.title_resolve.sql_search.search_title_in_sql", return_value={"ok": True, "data": sql_data}):
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": False, "error": "network_error", "details": "timeout"}):
            with patch("dataset.title_resolve.api_tmdb.search_tv_by_name", return_value=[{"id": 10, "name": "TMDb Title", "vote_count": 10}]):
                with patch("dataset.title_resolve.api_tmdb.get_tv_details", return_value=tmdb_details):
                    resolved = title_resolve.resolve_title_data_for_add("Input Title")

    defaults = resolved["defaults"]
    assert_check("KP API отмечен как ошибка", resolved["statuses"]["kp_api"] == "ошибка")
    assert_check("TMDb найден", resolved["statuses"]["tmdb_api"] == "найдено")
    assert_check("IMDb остаётся из SQL", defaults["raw_scores"]["imdb_score"] == 7.5)
    assert_check("KP score не берётся из TMDb", defaults["raw_scores"]["kp_score"] is None)
    assert_check("Жанры берутся из TMDb, если KP не сработал", resolved["sources"]["genres"] == "tmdb_api")
    assert_check("Описание берётся из TMDb", resolved["sources"]["description"] == "tmdb_api")


def test_add_resolver_offline_without_sql_is_manual() -> None:
    """Проверяет, что полный offline/no-match сценарий остаётся ручным."""
    print("\n11.18) Проверяем полный offline/manual сценарий resolver-а")

    with patch("dataset.title_resolve.sql_search.search_title_in_sql", return_value={"ok": False, "details": "not_found"}):
        with patch("dataset.title_resolve.api.find_series_raw", return_value={"ok": False, "error": "network_error", "details": "timeout"}):
            with patch("dataset.title_resolve.api_tmdb.search_tv_by_name", side_effect=RuntimeError("offline")):
                resolved = title_resolve.resolve_title_data_for_add("Manual Only")

    assert_check("Ничего не найдено", resolved["found"] is False)
    assert_check("Defaults не создаются до согласия пользователя", resolved["defaults"] is None)
    assert_check("SQL не найден", resolved["statuses"]["sql"] == "не найдено")
    assert_check("KP API ошибка", resolved["statuses"]["kp_api"] == "ошибка")
    assert_check("TMDb API ошибка", resolved["statuses"]["tmdb_api"] == "ошибка")


def test_candidate_pool_cross_criteria_keys_survive_save_load() -> None:
    """Проверяет, что одинаковый title/year в разных criteria_name не схлопывается."""
    print("\n12) Проверяем cross-criteria ключи candidate_pool")

    candidate_pool.save_candidate_pool({
        "legacy-one": {
            "title": "Метод",
            "alternative_title": "",
            "year": 2015,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.3,
            "kp_votes": 1000,
            "imdb_score": 7.4,
            "imdb_votes": 5000,
            "genres": ["драма"],
        },
        "legacy-two": {
            "title": "Метод",
            "alternative_title": "",
            "year": 2015,
            "criteria_name": "tmdb_RU_hidden",
            "kp_score": 7.8,
            "kp_votes": 1500,
            "imdb_score": 7.5,
            "imdb_votes": 4500,
            "genres": ["драма"],
        },
    })

    pool = candidate_pool.load_candidate_pool()
    criteria_names = {candidate.get("criteria_name") for candidate in pool.values()}

    assert_check("В пуле остаются обе criteria-версии", len(pool) == 2)
    assert_check("Сохраняются оба criteria_name", criteria_names == {"tmdb_RU_quality", "tmdb_RU_hidden"})
    assert_check(
        "Ключи пула становятся criteria-aware",
        set(pool.keys()) == {"tmdb_ru_quality|метод|2015", "tmdb_ru_hidden|метод|2015"}
    )


def test_candidate_pool_same_criteria_duplicates_keep_best() -> None:
    """Проверяет дедуп внутри одного criteria-aware ключа."""
    print("\n13) Проверяем дедуп внутри одного pool_entry_key")

    candidate_pool.save_candidate_pool({
        "first": {
            "title": "Метод",
            "alternative_title": "",
            "year": 2015,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.1,
            "kp_votes": 800,
            "imdb_score": 7.0,
            "imdb_votes": 2500,
            "genres": ["драма"],
        },
        "second": {
            "title": "Метод",
            "alternative_title": "",
            "year": 2015,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 8.4,
            "kp_votes": 2200,
            "imdb_score": 8.0,
            "imdb_votes": 9000,
            "genres": ["драма"],
        },
    })

    pool = candidate_pool.load_candidate_pool()
    saved_candidate = next(iter(pool.values()))

    assert_check("Дубликаты в одном criteria-aware ключе схлопываются", len(pool) == 1)
    assert_check("Остаётся лучший кандидат по текущему score-правилу", saved_candidate["kp_score"] == 8.4)


def test_candidate_pool_duplicate_scripts_helpers() -> None:
    """Checks offline helpers used by scripts/candidate_pool_* duplicate tools."""
    print("\n13b) Проверяем helper для scripts candidate_pool duplicate tools")

    candidate_pool.save_candidate_pool({
        "method_quality": {
            "title": "Method",
            "alternative_title": "",
            "year": 2015,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.1,
            "kp_votes": 800,
            "imdb_score": 7.0,
            "imdb_votes": 2500,
            "genres": ["drama"],
        },
        "method_hidden": {
            "title": "Method",
            "alternative_title": "",
            "year": 2015,
            "criteria_name": "tmdb_RU_hidden",
            "kp_score": 8.4,
            "kp_votes": 2200,
            "imdb_score": 8.0,
            "imdb_votes": 9000,
            "genres": ["drama"],
        },
        "trigger_one": {
            "title": "Trigger",
            "alternative_title": "",
            "year": 2018,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.5,
            "kp_votes": 1000,
            "imdb_score": 7.3,
            "imdb_votes": 3000,
            "genres": ["drama"],
        },
        "trigger_two": {
            "title": "Triggers",
            "alternative_title": "",
            "year": 2018,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.0,
            "kp_votes": 500,
            "imdb_score": 7.1,
            "imdb_votes": 1500,
            "genres": ["drama"],
        },
    })

    groups = pool_duplicate_tools.find_exact_duplicate_groups()
    pairs = pool_duplicate_tools.find_similar_title_pairs()

    assert_check("duplicate script helper находит exact title+year повтор", len(groups) == 1)
    assert_check("duplicate script helper ставит лучший exact candidate первым", groups[0][0].candidate["kp_score"] == 8.4)
    assert_check(
        "similar script helper находит похожие названия",
        any(
            {pair["left"].candidate["title"], pair["right"].candidate["title"]} == {"Trigger", "Triggers"}
            for pair in pairs
        )
    )

    trigger_entry = next(
        entry
        for entry in pool_duplicate_tools.load_pool_entries()
        if entry.candidate["title"] == "Triggers"
    )
    removed = pool_duplicate_tools.delete_entries_by_keys({trigger_entry.key})
    pool = candidate_pool.load_candidate_pool()

    assert_check("duplicate script helper удаляет выбранный key", removed == 1)
    assert_check("duplicate script helper не удаляет соседний похожий title", any(item.get("title") == "Trigger" for item in pool.values()))


def test_load_candidate_pool_is_read_only() -> None:
    """Проверяет, что load_candidate_pool больше не пишет файл на чтении."""
    print("\n14) Проверяем read-only поведение load_candidate_pool")

    candidate_pool.save_candidate_pool({
        "one": {
            "title": "Метод",
            "alternative_title": "",
            "year": 2015,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.5,
            "kp_votes": 1000,
            "imdb_score": 7.3,
            "imdb_votes": 3000,
            "genres": ["драма"],
        },
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    with patch("candidates.candidate_pool.save_candidate_pool", side_effect=RuntimeError("load must not save")):
        pool = candidate_pool.load_candidate_pool()

    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    assert_check("load_candidate_pool возвращает данные", len(pool) == 1)
    assert_check("load_candidate_pool не меняет mtime файла", before_mtime == after_mtime)


def test_read_path_keeps_watched_in_json() -> None:
    """Проверяет, что read-path не удаляет watched из JSON и не переписывает файл."""
    print("\n15) Проверяем read-path: watched остаётся в JSON")

    storage_movie.add_movie(make_movie(title="Watched In Pool", user_score=8.0))
    candidate_pool.init_candidate_pool()
    pool_data = {
        "watched-entry": {
            "title": "Watched In Pool",
            "alternative_title": "",
            "year": 2024,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.5,
            "kp_votes": 1000,
            "imdb_score": 7.3,
            "imdb_votes": 3000,
            "genres": ["драма"],
        },
    }
    with open(constant.CANDIDATE_POOL_JSON, "w", encoding="utf-8") as file:
        json.dump(pool_data, file, ensure_ascii=False, indent=4)

    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    with patch("candidates.candidate_pool.save_candidate_pool", side_effect=RuntimeError("read must not save")):
        candidates = candidate_pool.get_all_candidates()
        by_criteria = candidate_pool.get_candidates_by_criteria("tmdb_RU_quality")
    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    pool = candidate_pool.load_candidate_pool()

    assert_check("get_all_candidates не переписывает JSON", before_mtime == after_mtime)
    assert_check("Watched candidate остаётся в JSON", any(
        item.get("title") == "Watched In Pool" for item in pool.values()
    ))
    assert_check("get_all_candidates возвращает watched из storage", any(
        item.get("title") == "Watched In Pool" for item in candidates
    ))
    assert_check("get_candidates_by_criteria возвращает watched из storage", any(
        item.get("title") == "Watched In Pool" for item in by_criteria
    ))


def test_write_path_purges_watched_from_json() -> None:
    """Проверяет, что save_candidate_pool удаляет watched из сохранённого JSON."""
    print("\n16) Проверяем write-path: watched purge при save")

    storage_movie.add_movie(make_movie(title="Watched On Save", user_score=8.0))
    candidate_pool.init_candidate_pool()
    pool_data = {
        "watched-entry": {
            "title": "Watched On Save",
            "alternative_title": "",
            "year": 2024,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 7.5,
            "kp_votes": 1000,
            "imdb_score": 7.3,
            "imdb_votes": 3000,
            "genres": ["драма"],
        },
        "other-entry": {
            "title": "Still In Pool",
            "alternative_title": "",
            "year": 2023,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 8.0,
            "kp_votes": 2000,
            "imdb_score": 7.8,
            "imdb_votes": 5000,
            "genres": ["драма"],
        },
    }
    with open(constant.CANDIDATE_POOL_JSON, "w", encoding="utf-8") as file:
        json.dump(pool_data, file, ensure_ascii=False, indent=4)

    candidate_pool.save_candidate_pool(candidate_pool.load_candidate_pool())
    pool = candidate_pool.load_candidate_pool()

    assert_check("Watched candidate удалён из JSON", all(
        item.get("title") != "Watched On Save" for item in pool.values()
    ))
    assert_check("Непросмотренный кандидат остаётся в JSON", any(
        item.get("title") == "Still In Pool" for item in pool.values()
    ))


def test_get_pool_stats_reports_raw_storage_watched_ready() -> None:
    """Проверяет согласованные счётчики pool stats для UI."""
    print("\n17) Проверяем get_pool_stats: raw/storage/watched/ready/incomplete")

    storage_movie.add_movie(make_movie(title="Watched Stats", user_score=8.0))
    candidate_pool.init_candidate_pool()
    pool_data = {
        "legacy-watched": {
            "title": "Watched Stats",
            "alternative_title": "",
            "year": 2024,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": None,
            "kp_votes": None,
            "imdb_score": 7.3,
            "imdb_votes": 3000,
            "genres": ["драма"],
        },
        "legacy-ready": {
            "title": "Ready Stats",
            "alternative_title": "",
            "year": 2023,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 8.0,
            "kp_votes": 2000,
            "imdb_score": 7.8,
            "imdb_votes": 5000,
            "genres": ["драма"],
        },
        "legacy-incomplete": {
            "title": "Incomplete Stats",
            "alternative_title": "",
            "year": 2022,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": None,
            "kp_votes": None,
            "imdb_score": 7.1,
            "imdb_votes": 4000,
            "genres": ["драма"],
        },
    }
    with open(constant.CANDIDATE_POOL_JSON, "w", encoding="utf-8") as file:
        json.dump(pool_data, file, ensure_ascii=False, indent=4)

    stats = candidate_pool.get_pool_stats()
    scoped_stats = candidate_pool.get_pool_stats(criteria_name="tmdb_RU_quality")

    assert_check("raw_total считает записи JSON", stats["raw_total"] == 3)
    assert_check("storage_total считает normalized pool", stats["storage_total"] == 3)
    assert_check("watched_total видит просмотренного в pool", stats["watched_total"] == 1)
    assert_check("active_total = storage - watched", stats["active_total"] == 2)
    assert_check("ready_total считает complete", stats["ready_total"] == 1)
    assert_check("incomplete_total считает неполных", stats["incomplete_total"] == 2)
    assert_check("Scoped stats совпадают с общими для одного criteria", scoped_stats["storage_total"] == 3)


def test_build_prediction_filter_defaults_from_saved_criteria() -> None:
    """Проверяет, что top prediction может взять defaults из candidate_criteria.json."""
    print("\n18) Проверяем prediction filter defaults из criteria")

    candidate_pool.save_named_criteria("tmdb_RU_quality", {
        "country": "Россия",
        "count": 20,
        "min_kp": 7.0,
        "min_kp_votes": 100,
        "min_imdb": 6.5,
        "min_imdb_votes": 500,
        "min_year": 2020,
        "max_year": 2024,
        "genres": ["драма"],
        "excluded_genres": ["комедия"],
    })

    defaults = candidate_pool.build_prediction_filter_defaults("tmdb_RU_quality")

    assert_check("criteria_name сохраняется в defaults", defaults["criteria_name"] == "tmdb_RU_quality")
    assert_check("country подтягивается из criteria", defaults["country"] == "Россия")
    assert_check("min_kp_score подтягивается из criteria", defaults["min_kp_score"] == 7.0)
    assert_check("include_genres подтягиваются из criteria", defaults["include_genres"] == ["драма"])
    assert_check("exclude_genres подтягиваются из criteria", defaults["exclude_genres"] == ["комедия"])
    assert_check("year_min подтягивается из criteria", defaults["year_min"] == 2020)
    assert_check("year_max подтягивается из criteria", defaults["year_max"] == 2024)


def test_prediction_filters_apply_saved_criteria_defaults() -> None:
    """Проверяет, что runtime filters из criteria реально отбирают pool без его изменения."""
    print("\n19) Проверяем применение criteria defaults к runtime filters")

    candidate_pool.save_named_criteria("tmdb_RU_quality", {
        "min_kp": 7.5,
        "genres": ["драма"],
        "excluded_genres": ["комедия"],
    })
    candidate_pool.save_candidate_pool({
        "one": {
            "title": "Drama Match",
            "year": 2022,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 8.0,
            "kp_votes": 1000,
            "imdb_score": 7.5,
            "imdb_votes": 5000,
            "genres": ["драма"],
        },
        "two": {
            "title": "Comedy Skip",
            "year": 2022,
            "criteria_name": "tmdb_RU_quality",
            "kp_score": 8.5,
            "kp_votes": 1200,
            "imdb_score": 7.8,
            "imdb_votes": 6000,
            "genres": ["комедия"],
        },
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    candidates = candidate_pool.get_all_candidates()
    filters = candidate_pool.build_prediction_filter_defaults("tmdb_RU_quality")
    filtered = candidate_pool.filter_saved_candidates_for_prediction(candidates, filters)
    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    assert_check("Runtime filters не переписывают pool JSON", before_mtime == after_mtime)
    assert_check("Criteria defaults отбирают только подходящий жанр", len(filtered) == 1)
    assert_check("Остаётся drama-кандидат", filtered[0]["title"] == "Drama Match")


def _runtime_genre_filter(include_genres=None, exclude_genres=None, candidates=None):
    if candidates is None:
        candidates = []
    return candidate_pool.filter_saved_candidates_for_prediction(
        candidates,
        {
            "criteria_name": None,
            "source": None,
            "country": None,
            "year_min": None,
            "year_max": None,
            "include_genres": include_genres or [],
            "exclude_genres": exclude_genres or [],
            "min_kp_score": None,
            "min_kp_votes": None,
            "min_imdb_score": None,
            "min_imdb_votes": None,
            "min_tmdb_score": None,
            "min_tmdb_votes": None,
            "only_complete": False,
        },
    )


def _genre_ready_candidate(title: str, genres: list) -> dict:
    return {
        "title": title,
        "year": 2022,
        "criteria_name": "tmdb_RU_quality",
        "kp_score": 8.0,
        "kp_votes": 1000,
        "imdb_score": 7.5,
        "imdb_votes": 5000,
        "genres": genres,
    }


def test_genre_normalization_runtime_filters() -> None:
    """Проверяет RU/EN genre aliases только в runtime-фильтрации pool."""
    print("\n20) Проверяем runtime genre normalization")

    drama_candidate = _genre_ready_candidate("Drama EN", ["Drama"])
    mystery_candidate = _genre_ready_candidate("Mystery EN", ["Mystery"])
    crime_candidate = _genre_ready_candidate("Crime EN", ["Crime"])
    reality_candidate = _genre_ready_candidate("Reality EN", ["Reality"])
    drama_only_candidate = _genre_ready_candidate("Only Drama", ["Drama"])

    assert_check(
        "RU drama матчится с EN Drama",
        len(_runtime_genre_filter(["драма"], candidates=[drama_candidate])) == 1,
    )
    assert_check(
        "RU детектив матчится с EN Mystery",
        len(_runtime_genre_filter(["детектив"], candidates=[mystery_candidate])) == 1,
    )
    assert_check(
        "RU криминал матчится с EN Crime",
        len(_runtime_genre_filter(["криминал"], candidates=[crime_candidate])) == 1,
    )
    assert_check(
        "Exclude реалити отсекает EN Reality",
        len(_runtime_genre_filter(exclude_genres=["реалити"], candidates=[reality_candidate])) == 0,
    )
    assert_check(
        "Unknown genre не ломает фильтр и не матчится",
        len(_runtime_genre_filter(["странный жанр"], candidates=[drama_only_candidate])) == 0,
    )
    assert_check(
        "normalize_genre_list убирает дубли и регистр",
        pool_genres.normalize_genre_list([" Drama ", "драма", "DRAMA"]) == ["drama"],
    )

    candidate_pool.save_candidate_pool({
        "one": drama_candidate,
        "two": reality_candidate,
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    candidates = candidate_pool.get_all_candidates()
    filtered = _runtime_genre_filter(["драма"], candidates=candidates)
    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    assert_check("Top/runtime filter не переписывает candidate_pool.json", before_mtime == after_mtime)
    assert_check("Runtime filter по RU drama находит EN Drama из pool", len(filtered) == 1)
    assert_check("Runtime filter оставляет EN Drama title", filtered[0]["title"] == "Drama EN")


def test_canonical_runtime_filters_use_genre_keys_and_country_codes() -> None:
    """Проверяет, что top prediction filters сравнивают canonical keys, а не raw strings."""
    print("\n20b) Проверяем canonical runtime filters для genre_keys/country_codes")

    def _filter_candidates(candidates: list, **filters) -> list:
        payload = {
            "criteria_name": None,
            "source": None,
            "country": None,
            "year_min": None,
            "year_max": None,
            "include_genres": [],
            "exclude_genres": [],
            "min_kp_score": None,
            "min_kp_votes": None,
            "min_imdb_score": None,
            "min_imdb_votes": None,
            "only_complete": False,
        }
        payload.update(filters)
        return candidate_pool.filter_saved_candidates_for_prediction(candidates, payload)

    def _ready(title: str, **fields) -> dict:
        return candidate_schema.normalize_candidate_record({
            "title": title,
            "year": 2020,
            "kp_score": 8.0,
            "kp_votes": 1000,
            "imdb_score": 7.5,
            "imdb_votes": 5000,
            **fields,
        })

    kr_drama_crime = _ready(
        "KR Drama Crime",
        countries=["KR", "South Korea"],
        genres=["Drama", "Crime"],
    )
    kr_comedy = _ready(
        "KR Comedy",
        countries=["KR", "South Korea"],
        genres=["Comedy"],
    )
    ru_drama = _ready(
        "RU Drama",
        countries=["RU", "Россия"],
        genres=["Drama"],
    )
    jp_drama = _ready(
        "JP Drama",
        countries=["JP", "Япония"],
        genres=["Drama"],
    )
    unknown_country = _ready(
        "Unknown Country",
        countries=["Atlantis"],
        genres=["Drama"],
    )
    unknown_genre = _ready(
        "Unknown Genre",
        countries=["KR"],
        genres=["TotallyUnknownGenreXYZ"],
    )

    assert_check(
        "RU drama filter матчится с EN Drama через genre_keys",
        len(_filter_candidates([kr_drama_crime], include_genres=["драма"])) == 1,
    )
    assert_check(
        "Include драма, криминал находит mixed EN candidate",
        len(_filter_candidates([kr_drama_crime], include_genres=["драма", "криминал"])) == 1,
    )
    assert_check(
        "Exclude комедия отсекает Comedy candidate",
        len(_filter_candidates([kr_comedy], exclude_genres=["комедия"])) == 0,
    )
    assert_check(
        "Exclude комедия не отсекает Drama/Crime candidate",
        len(_filter_candidates([kr_drama_crime], exclude_genres=["комедия"])) == 1,
    )
    assert_check(
        "Country filter KR матчится с raw KR/South Korea через country_codes",
        len(_filter_candidates([kr_drama_crime], country="KR")) == 1,
    )
    assert_check(
        "Country filter Южная Корея матчится с KR candidate",
        len(_filter_candidates([kr_drama_crime], country="Южная Корея")) == 1,
    )
    assert_check(
        "Country filter KR не пропускает RU candidate",
        len(_filter_candidates([ru_drama], country="KR")) == 0,
    )
    assert_check(
        "Country filter KR,JP матчит обе страны (OR)",
        len(_filter_candidates([kr_drama_crime, ru_drama, jp_drama], country=["KR", "JP"])) == 2,
    )
    assert_check(
        "Country filter KR,JP через запятую в строке",
        len(_filter_candidates([kr_drama_crime, ru_drama, jp_drama], country="KR, JP")) == 2,
    )
    assert_check(
        "Unknown country filter не матчит candidate и не падает",
        len(_filter_candidates([kr_drama_crime], country="Atlantis")) == 0,
    )
    assert_check(
        "Unknown include genre не матчит candidate и не падает",
        len(_filter_candidates([kr_drama_crime], include_genres=["странный жанр"])) == 0,
    )
    assert_check(
        "Candidate без mappable genre_keys не проходит include filter",
        len(_filter_candidates([unknown_genre], include_genres=["драма"])) == 0,
    )
    assert_check(
        "Candidate без country_codes не проходит country filter",
        len(_filter_candidates([unknown_country], country="KR")) == 0,
    )

    candidate_pool.save_candidate_pool({
        "kr": kr_drama_crime,
        "ru": ru_drama,
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    filtered = _filter_candidates(
        candidate_pool.get_all_candidates(),
        country="KR",
        include_genres=["драма"],
        exclude_genres=["комедия"],
    )
    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    assert_check("Canonical runtime filter read-only для pool JSON", before_mtime == after_mtime)
    assert_check("Combined KR+drama filter оставляет KR Drama Crime", len(filtered) == 1)
    assert_check("Combined filter сохраняет title", filtered[0]["title"] == "KR Drama Crime")

    genre_options = candidate_pool.collect_prediction_genre_options([kr_drama_crime, kr_comedy])
    assert_check("Genre options показывают display labels", set(genre_options) == {"Драма", "Криминал", "Комедия"})


def test_prediction_numeric_string_filters_are_safe() -> None:
    """Проверяет, что runtime numeric filters не падают на числах-строках из JSON."""
    print("\n20c) Проверяем runtime numeric filters на строковых числах")

    def _filter_candidates(candidates: list, **filters) -> list:
        payload = {
            "criteria_name": None,
            "source": None,
            "country": None,
            "year_min": None,
            "year_max": None,
            "include_genres": [],
            "exclude_genres": [],
            "min_kp_score": None,
            "min_kp_votes": None,
            "min_imdb_score": None,
            "min_imdb_votes": None,
            "only_complete": False,
        }
        payload.update(filters)
        return candidate_pool.filter_saved_candidates_for_prediction(candidates, payload)

    string_number_candidate = {
        "title": "String Number Candidate",
        "year": "2021",
        "kp_score": "7.5",
        "kp_votes": "50000",
        "imdb_score": "7.1",
        "imdb_votes": "10000",
        "genres": [],
        "countries": [],
    }
    lower_score_candidate = dict(string_number_candidate, title="Lower KP", kp_score="6.9")
    garbage_score_candidate = dict(string_number_candidate, title="Garbage KP", kp_score="unknown")
    none_score_candidate = dict(string_number_candidate, title="None KP", kp_score=None)
    empty_score_candidate = dict(string_number_candidate, title="Empty KP", kp_score="")
    garbage_votes_candidate = dict(string_number_candidate, title="Garbage Votes", kp_votes="abc")

    passed = _filter_candidates(
        [string_number_candidate],
        min_kp_score=7.0,
        min_kp_votes=10000,
        min_imdb_score=7.0,
        year_min=2020,
    )
    assert_check("Строковые score/votes/year проходят numeric filters", len(passed) == 1)
    assert_check("Строковое число ниже минимума отсекается", len(_filter_candidates([lower_score_candidate], min_kp_score=7.0)) == 0)
    assert_check("Нечисловой score не падает и отсекается", len(_filter_candidates([garbage_score_candidate], min_kp_score=7.0)) == 0)
    assert_check("None score при активном min отсекается", len(_filter_candidates([none_score_candidate], min_kp_score=7.0)) == 0)
    assert_check("Пустой score при активном min отсекается", len(_filter_candidates([empty_score_candidate], min_kp_score=7.0)) == 0)
    assert_check("None score без min filter не отсекается сам по себе", len(_filter_candidates([none_score_candidate])) == 1)
    assert_check("Пустой score без min filter не отсекается сам по себе", len(_filter_candidates([empty_score_candidate])) == 1)
    assert_check("Строковые votes проходят numeric min", len(_filter_candidates([string_number_candidate], min_kp_votes=10000)) == 1)
    assert_check("Мусорные votes не падают и отсекаются", len(_filter_candidates([garbage_votes_candidate], min_kp_votes=10000)) == 0)

    candidate_pool.save_candidate_pool({"string-number": string_number_candidate})
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    read_only_filtered = _filter_candidates(
        candidate_pool.get_all_candidates(),
        min_kp_score=7.0,
        min_kp_votes=10000,
        min_imdb_score=7.0,
        year_min=2020,
    )
    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    assert_check("Numeric runtime filter не переписывает candidate_pool.json", before_mtime == after_mtime)
    assert_check("Numeric runtime filter из JSON оставляет строкового кандидата", len(read_only_filtered) == 1)


def test_prediction_filter_accepts_numeric_strings() -> None:
    """Regression: numeric runtime filters accept numeric values saved as strings."""
    print("\n20c.1) Проверяем numeric strings в runtime filters")

    candidate = {
        "title": "Numeric Strings",
        "year": "2021",
        "kp_score": "7.5",
        "kp_votes": "100",
        "imdb_score": "7.1",
        "imdb_votes": "1000",
        "genres": [],
        "countries": [],
    }
    filtered = candidate_pool.filter_saved_candidates_for_prediction(
        [candidate],
        {
            "criteria_name": None,
            "source": None,
            "country": None,
            "year_min": 2020,
            "year_max": None,
            "include_genres": [],
            "exclude_genres": [],
            "min_kp_score": 7.0,
            "min_kp_votes": 50,
            "min_imdb_score": 7.0,
            "min_imdb_votes": None,
            "only_complete": False,
        },
    )
    assert_check("Numeric strings проходят фильтр без TypeError", len(filtered) == 1)


def test_prediction_filter_rejects_invalid_numeric_strings() -> None:
    """Regression: invalid numeric strings fail active min filters without TypeError."""
    print("\n20c.2) Проверяем invalid numeric strings в runtime filters")

    def _candidate(kp_score) -> dict:
        return {
            "title": f"Invalid Numeric {kp_score}",
            "year": "2021",
            "kp_score": kp_score,
            "kp_votes": "100",
            "imdb_score": "7.1",
            "imdb_votes": "1000",
            "genres": [],
            "countries": [],
        }

    filters = {
        "criteria_name": None,
        "source": None,
        "country": None,
        "year_min": None,
        "year_max": None,
        "include_genres": [],
        "exclude_genres": [],
        "min_kp_score": 7.0,
        "min_kp_votes": None,
        "min_imdb_score": None,
        "min_imdb_votes": None,
        "only_complete": False,
    }

    for value in ("unknown", "N/A", "", "seven", "7.5/10"):
        filtered = candidate_pool.filter_saved_candidates_for_prediction([_candidate(value)], filters)
        assert_check(f"Invalid numeric '{value}' не проходит и не падает", len(filtered) == 0)


def test_coerce_candidate_number_handles_runtime_values() -> None:
    """Regression: shared numeric coercion for candidate runtime filters."""
    print("\n20c.2b) Проверяем coerce_candidate_number")

    assert_check("7.5 -> 7.5", candidate_schema.coerce_candidate_number(7.5) == 7.5)
    assert_check("50000 -> 50000", candidate_schema.coerce_candidate_number(50000) == 50000)
    assert_check('"7.5" -> 7.5', candidate_schema.coerce_candidate_number("7.5") == 7.5)
    assert_check('"50000" -> 50000', candidate_schema.coerce_candidate_number("50000") == 50000)
    assert_check('" 2021 " -> 2021', candidate_schema.coerce_candidate_number(" 2021 ") == 2021)
    assert_check('"7,5" -> 7.5', candidate_schema.coerce_candidate_number("7,5") == 7.5)
    assert_check("None -> None", candidate_schema.coerce_candidate_number(None) is None)
    assert_check('"" -> None', candidate_schema.coerce_candidate_number("") is None)
    assert_check('"unknown" -> None', candidate_schema.coerce_candidate_number("unknown") is None)
    assert_check('"N/A" -> None', candidate_schema.coerce_candidate_number("N/A") is None)
    assert_check('"7.5/10" -> None', candidate_schema.coerce_candidate_number("7.5/10") is None)
    assert_check("True -> None", candidate_schema.coerce_candidate_number(True) is None)
    assert_check("False -> None", candidate_schema.coerce_candidate_number(False) is None)


def test_prediction_filter_does_not_mutate_input_candidates() -> None:
    """Regression: runtime filters normalize an in-memory copy, not the input dict."""
    print("\n20c.3) Проверяем, что runtime filters не мутируют input candidates")

    candidate = {
        "title": "Immutable Runtime Candidate",
        "year": 2021,
        "kp_score": 7.5,
        "kp_votes": 100,
        "imdb_score": 7.1,
        "imdb_votes": 1000,
        "genres": ["Drama"],
        "countries": ["KR", "South Korea"],
    }
    before = copy.deepcopy(candidate)
    filtered = candidate_pool.filter_saved_candidates_for_prediction(
        [candidate],
        {
            "criteria_name": None,
            "source": None,
            "country": "KR",
            "year_min": None,
            "year_max": None,
            "include_genres": ["драма"],
            "exclude_genres": [],
            "min_kp_score": None,
            "min_kp_votes": None,
            "min_imdb_score": None,
            "min_imdb_votes": None,
            "only_complete": False,
        },
    )

    assert_check("Runtime filter работает через normalized copy", len(filtered) == 1)
    assert_check("Input candidate dict не изменился", candidate == before)
    assert_check("genre_keys не добавлен в input candidate", "genre_keys" not in candidate)
    assert_check("country_codes не добавлен в input candidate", "country_codes" not in candidate)
    assert_check("genres_display не добавлен в input candidate", "genres_display" not in candidate)
    assert_check("country_display не добавлен в input candidate", "country_display" not in candidate)


def test_top_prediction_title_identity_dedupe() -> None:
    """Проверяет read-only dedupe top prediction по title_identity_key без изменения pool JSON."""
    print("\n20d) Проверяем top prediction dedupe по title_identity_key")

    def _candidate(**fields) -> dict:
        base = {
            "title": "Same Show",
            "year": 2020,
            "kp_score": 7.0,
            "kp_votes": 1000,
            "imdb_score": 7.0,
            "imdb_votes": 1000,
            "kp_status": "done",
            "is_complete": True,
            "genres": ["Drama"],
            "countries": ["RU"],
        }
        base.update(fields)
        return candidate_schema.normalize_candidate_record(base)

    crit_a = _candidate(criteria_name="crit_a", predict=7.5)
    crit_b = _candidate(criteria_name="crit_b", predict=8.2)
    deduped = candidate_pool.dedupe_ranked_predictions_by_title_identity([crit_a, crit_b])
    assert_check(
        "Same title/year из разных criteria -> один кандидат",
        len(deduped) == 1 and deduped[0]["criteria_name"] == "crit_b",
    )

    low_predict = _candidate(criteria_name="low", predict=7.0)
    high_predict = _candidate(criteria_name="high", predict=9.0)
    deduped_score = candidate_pool.dedupe_ranked_predictions_by_title_identity([low_predict, high_predict])
    assert_check(
        "При разных predicted score остаётся кандидат с большим score",
        len(deduped_score) == 1 and deduped_score[0]["criteria_name"] == "high",
    )

    incomplete = _candidate(
        criteria_name="incomplete",
        kp_score=None,
        kp_votes=None,
        imdb_score=None,
        imdb_votes=None,
        kp_status="missing",
        is_complete=False,
    )
    ready = _candidate(criteria_name="ready", kp_score=8.5, kp_votes=5000)
    deduped_ready = candidate_pool.dedupe_ranked_predictions_by_title_identity([incomplete, ready])
    assert_check(
        "Без predict остаётся ready/complete кандидат",
        len(deduped_ready) == 1 and deduped_ready[0]["criteria_name"] == "ready",
    )

    other_year = _candidate(title="Same Show", year=2021, criteria_name="other_year", predict=8.0)
    deduped_years = candidate_pool.dedupe_ranked_predictions_by_title_identity([crit_a, other_year])
    assert_check("Same title, different year не склеиваются", len(deduped_years) == 2)

    other_title = _candidate(title="Other Show", year=2020, criteria_name="other_title", predict=8.0)
    deduped_titles = candidate_pool.dedupe_ranked_predictions_by_title_identity([crit_a, other_title])
    assert_check("Same year, different title не склеиваются", len(deduped_titles) == 2)

    equal_first = _candidate(criteria_name="first", predict=8.0, kp_score=7.0)
    equal_second = _candidate(criteria_name="second", predict=8.0, kp_score=7.0)
    deduped_order = candidate_pool.dedupe_ranked_predictions_by_title_identity([equal_first, equal_second])
    assert_check(
        "При полном равенстве остаётся первый по исходному порядку",
        len(deduped_order) == 1 and deduped_order[0]["criteria_name"] == "first",
    )

    candidate_pool.save_candidate_pool({
        "dup_a": crit_a,
        "dup_b": crit_b,
        "unique": other_title,
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    weights = storage_data.load_weights()
    ranked = candidate_pool.rank_candidates_by_predict(candidate_pool.get_all_candidates(), weights)
    deduped_flow = candidate_pool.dedupe_ranked_predictions_by_title_identity(ranked)
    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    assert_check("Top prediction dedupe не переписывает candidate_pool.json", before_mtime == after_mtime)
    assert_check(
        "Rank+dedupe flow склеивает title/year дубли",
        len(deduped_flow) == 2 and all(
            item["title"] in {"Same Show", "Other Show"} for item in deduped_flow
        ),
    )


def test_contributions_readiness_gate() -> None:
    """Проверяет readiness gate для feature contributions без изменения pool/model."""
    print("\n21) Проверяем readiness gate для contributions")

    ready_candidate = _genre_ready_candidate("Ready Title", ["Drama"])
    incomplete_candidate = {
        "title": "Incomplete Title",
        "year": 2021,
        "imdb_score": 7.0,
        "imdb_votes": 1000,
        "kp_score": None,
        "kp_votes": None,
    }

    ready_only = candidate_pool.select_ready_candidates_for_contributions(
        [ready_candidate, incomplete_candidate]
    )
    assert_check("Ready candidate проходит contributions gate", len(ready_only) == 1)
    assert_check(
        "Incomplete candidate отфильтровывается",
        candidate_pool.select_ready_candidates_for_contributions([incomplete_candidate]) == [],
    )

    not_ready_message = candidate_pool.candidate_not_ready_for_contributions_message(incomplete_candidate)
    assert_check(
        "Incomplete candidate получает понятное сообщение",
        not_ready_message is not None and "kp_score" in not_ready_message,
    )
    assert_check(
        "Ready candidate не получает not-ready сообщение",
        candidate_pool.candidate_not_ready_for_contributions_message(ready_candidate) is None,
    )

    weights = storage_data.load_weights()
    contribution_calls = {"count": 0}

    def _count_contributions(candidate, loaded_weights):
        contribution_calls["count"] += 1
        return {
            "title": candidate.get("title"),
            "year": candidate.get("year"),
            "predict": 0.0,
            "positive": [],
            "negative": [],
            "criteria_name": "",
        }

    with patch("candidates.candidate_pool.candidate_feature_contributions", side_effect=_count_contributions):
        incomplete_reports = candidate_pool.build_contribution_reports_for_ready_candidates(
            [incomplete_candidate],
            weights,
        )

    assert_check("Incomplete candidate не вызывает contributions", contribution_calls["count"] == 0)
    assert_check("Reports для incomplete-only пустой", incomplete_reports == [])

    contribution_calls["count"] = 0
    with patch("candidates.candidate_pool.candidate_feature_contributions", side_effect=_count_contributions):
        mixed_reports = candidate_pool.build_contribution_reports_for_ready_candidates(
            [ready_candidate, incomplete_candidate],
            weights,
        )

    assert_check("Contributions считаются только для ready", contribution_calls["count"] == 1)
    assert_check(
        "Report строится только для ready candidate",
        len(mixed_reports) == 1 and mixed_reports[0]["title"] == "Ready Title",
    )

    candidate_pool.save_candidate_pool({
        "one": ready_candidate,
        "two": incomplete_candidate,
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    candidate_pool.build_contribution_reports_for_ready_candidates(
        candidate_pool.get_all_candidates(),
        weights,
    )
    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    assert_check("Contributions gate не переписывает candidate_pool.json", before_mtime == after_mtime)


def test_candidate_service_read_only_facade() -> None:
    """Проверяет read-only facade candidates.service без записи в JSON."""
    print("\n22) Проверяем candidates.service read-only facade")

    ready_candidate = _genre_ready_candidate("Service Ready", ["Drama"])
    incomplete_candidate = {
        "title": "Service Incomplete",
        "year": 2021,
        "criteria_name": "tmdb_RU_quality",
        "imdb_score": 7.0,
        "imdb_votes": 1000,
        "kp_score": None,
        "kp_votes": None,
    }

    candidate_pool.save_candidate_pool({
        "one": ready_candidate,
        "two": incomplete_candidate,
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    all_view = candidate_service.get_pool_view()
    scoped_view = candidate_service.get_pool_view(criteria_name="tmdb_RU_quality")
    stats_view = candidate_service.get_pool_stats_view()
    contribution_view = candidate_service.get_contribution_ready_view(all_view)

    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    assert_check(
        "get_pool_view(None) совпадает с get_all_candidates()",
        all_view == candidate_pool.get_all_candidates(),
    )
    assert_check(
        "get_pool_view(criteria) совпадает с get_candidates_by_criteria()",
        scoped_view == candidate_pool.get_candidates_by_criteria("tmdb_RU_quality"),
    )
    assert_check(
        "get_pool_stats_view stats совместимы с get_pool_stats()",
        stats_view["stats"] == candidate_pool.get_pool_stats(),
    )
    assert_check(
        "get_pool_stats_view lines не пустые при непустом pool",
        len(stats_view["lines"]) > 0,
    )
    assert_check(
        "get_contribution_ready_view пропускает ready candidate",
        len(contribution_view["ready_candidates"]) == 1,
    )
    assert_check(
        "get_contribution_ready_view относит incomplete к skipped",
        len(contribution_view["skipped_incomplete"]) == 1,
    )
    assert_check(
        "get_contribution_ready_view даёт not-ready message",
        contribution_view["not_ready_messages"][0]["message"] is not None,
    )
    assert_check(
        "service read-only функции не переписывают candidate_pool.json",
        before_mtime == after_mtime,
    )


def test_candidate_service_top_prediction_view() -> None:
    """Проверяет read-only подготовку top prediction через candidates.service."""
    print("\n23) Проверяем candidates.service top prediction view")

    ready_candidate = _genre_ready_candidate("Top Ready", ["Drama"])
    incomplete_candidate = {
        "title": "Top Incomplete",
        "year": 2021,
        "criteria_name": "tmdb_RU_quality",
        "imdb_score": 7.0,
        "imdb_votes": 1000,
        "kp_score": None,
        "kp_votes": None,
        "genres": ["Drama"],
    }

    candidate_pool.save_candidate_pool({
        "one": ready_candidate,
        "two": incomplete_candidate,
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    top_view = candidate_service.get_global_top_prediction_view()
    filters = {
        "criteria_name": "tmdb_RU_quality",
        "source": None,
        "country": None,
        "year_min": None,
        "year_max": None,
        "include_genres": [],
        "exclude_genres": [],
        "min_kp_score": None,
        "min_kp_votes": None,
        "min_imdb_score": None,
        "min_imdb_votes": None,
        "min_tmdb_score": None,
        "min_tmdb_votes": None,
        "only_complete": False,
    }
    filter_view = candidate_service.get_prediction_filter_view(top_view["candidates"], filters)

    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    assert_check(
        "get_global_top_prediction_view candidates совпадают с get_pool_view()",
        top_view["candidates"] == candidate_service.get_pool_view(),
    )
    assert_check(
        "get_global_top_prediction_view stats совпадают с get_pool_stats_view()",
        top_view["stats"] == candidate_service.get_pool_stats_view()["stats"],
    )
    assert_check(
        "get_global_top_prediction_view is_empty=False для непустого pool",
        top_view["is_empty"] is False,
    )
    assert_check(
        "get_prediction_filter_view оставляет ready candidate",
        filter_view["ready_count"] == 1,
    )
    assert_check(
        "get_prediction_filter_view относит incomplete к skipped",
        filter_view["skipped_incomplete_count"] == 1,
    )
    assert_check(
        "get_prediction_filter_view совпадает с filter_saved_candidates_for_prediction",
        filter_view["filtered_count"]
        == len(candidate_pool.filter_saved_candidates_for_prediction(top_view["candidates"], filters)),
    )
    assert_check(
        "top prediction service view не переписывает candidate_pool.json",
        before_mtime == after_mtime,
    )

    ranking_view = candidate_service.rank_top_prediction_candidates(
        filter_view["ready_candidates"],
        storage_data.load_weights(),
    )
    assert_check(
        "rank_top_prediction_candidates возвращает scored candidates",
        len(ranking_view["candidates"]) == 1,
    )
    assert_check(
        "rank_top_prediction_candidates содержит predict",
        ranking_view["candidates"][0].get("predict") is not None,
    )


def test_top_prediction_ui_read_only_helpers() -> None:
    """Проверяет read-only helpers для top prediction UI."""
    print("\n23.1) Проверяем top prediction UI read-only helpers")

    long_description = "Описание " + ("очень длинное " * 30)
    candidate_pool.save_candidate_pool({
        "dup_a": {
            "title": "Эпидемия",
            "year": 2019,
            "criteria_name": "first",
            "kp_score": 7.8,
            "kp_votes": 1000,
            "imdb_score": 7.2,
            "imdb_votes": 2000,
            "countries": ["Россия"],
            "genres": ["Драма", "Триллер"],
            "description": long_description,
            "kp_status": "done",
            "is_complete": True,
        },
        "dup_b": {
            "title": "Эпидемия",
            "year": 2019,
            "criteria_name": "second",
            "kp_score": 8.0,
            "kp_votes": 2000,
            "imdb_score": 7.4,
            "imdb_votes": 3000,
            "countries": ["Россия"],
            "genres": ["Драма"],
            "description": "short",
            "kp_status": "done",
            "is_complete": True,
        },
        "other": {
            "title": "Other",
            "year": 2020,
            "criteria_name": "first",
            "kp_score": 7.0,
            "kp_votes": 1000,
            "imdb_score": 7.0,
            "imdb_votes": 1000,
            "countries": ["США"],
            "genres": ["Комедия"],
            "kp_status": "done",
            "is_complete": True,
        },
    })
    before_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns

    genre_view = candidate_service.get_prediction_genre_options_view()
    filtered = candidate_pool.filter_saved_candidates_for_prediction(
        candidate_pool.get_all_candidates(),
        {
            "criteria_name": None,
            "source": None,
            "country": None,
            "year_min": None,
            "year_max": None,
            "include_genres": [],
            "exclude_genres": [],
            "min_kp_score": None,
            "min_kp_votes": None,
            "min_imdb_score": None,
            "min_imdb_votes": None,
            "min_tmdb_score": 99,
            "min_tmdb_votes": 99_999_999,
            "only_complete": False,
        },
    )
    deduped = candidate_pool.dedupe_ranked_predictions_by_title_identity([
        {"title": "Эпидемия", "year": 2019, "predict": 7.10},
        {"title": "Эпидемия", "year": 2019, "predict": 8.20},
        {"title": "Other", "year": 2020, "predict": 7.50},
    ])
    short_description = candidate_pool.format_candidate_description({"description": long_description}, limit=200)
    missing_description = candidate_pool.format_candidate_description({})
    card_output = io.StringIO()
    with contextlib.redirect_stdout(card_output):
        interface_funcs._print_prediction_candidate_card(
            1,
            {
                "title": "Эпидемия",
                "year": 2019,
                "kp_score": 7.8,
                "imdb_score": 7.2,
                "countries": ["Россия"],
                "genres": ["Драма", "Триллер"],
                "description": long_description,
                "predict_score": 7.75,
            },
        )

    after_mtime = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    filter_source = inspect.getsource(interface_funcs._request_prediction_candidate_filters)

    assert_check("Genre options view видит Триллер из saved pool", "Триллер" in genre_view["genres"])
    assert_check("Genre options view подписан как saved pool", "saved" in genre_view["label"] or "сохранённым" in genre_view["label"])
    assert_check("Top prediction не применяет min_tmdb_score/min_tmdb_votes", len(filtered) == 3)
    assert_check("Top prediction UI больше не спрашивает TMDb filters", "Минимальный TMDb" not in filter_source and "Минимум голосов TMDb" not in filter_source)
    assert_check("Top prediction UI жанры через numbered list", "_choose_prediction_genre_list" in filter_source)
    assert_check("Top prediction UI страна через numbered list", "_choose_prediction_country" in filter_source)
    assert_check("Top prediction UI не спрашивает source", "_choose_prediction_source" not in filter_source)
    assert_check(
        "Top prediction country index parser принимает multi-select",
        interface_funcs._parse_prediction_country_indexes("1, 2", 26) == [1, 2],
    )
    assert_check(
        "Top prediction country index parser принимает пробелы",
        interface_funcs._parse_prediction_country_indexes("1 2", 26) == [1, 2],
    )
    assert_check(
        "Top prediction genre index parser принимает multi-select",
        interface_funcs._parse_prediction_genre_indexes("1, 3", 5) == [1, 3],
    )
    assert_check(
        "Top prediction genre index parser пустой ввод = не важно",
        interface_funcs._parse_prediction_genre_indexes("", 5) == [],
    )
    assert_check(
        "TMDb genre index parser пустой ввод = не важно",
        interface_funcs._parse_tmdb_genre_indexes("") == [],
    )
    assert_check("Dedupe оставляет один title/year с лучшим predict", len(deduped) == 2 and deduped[0]["predict"] == 8.20)
    assert_check("Описание режется до 200 символов", len(short_description) <= 200 and short_description.endswith("..."))
    assert_check("Пустое описание даёт нет данных", missing_description == "нет данных")
    card_text = card_output.getvalue()
    assert_check("Карточка выводит title/year", "Эпидемия (2019)" in card_text)
    assert_check("Карточка выводит KP / IMDb", "KP: 7.8 / IMDb: 7.2" in card_text)
    assert_check("Карточка выводит страну и жанры", "Россия" in card_text and "Драма, Триллер" in card_text)
    assert_check("Карточка выводит прогноз", "Прогноз: 7.75" in card_text)
    assert_check("Top prediction helpers не переписывают candidate_pool.json", before_mtime == after_mtime)


def test_candidate_service_prediction_filter_defaults_view() -> None:
    """Проверяет read-only defaults view для top prediction через candidates.service."""
    print("\n24) Проверяем candidates.service prediction filter defaults view")

    candidate_pool.save_named_criteria("tmdb_RU_quality", {
        "country": "Россия",
        "min_kp": 7.0,
        "genres": ["драма"],
        "excluded_genres": ["комедия"],
        "min_year": 2020,
        "max_year": 2024,
    })

    pool_mtime_before = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    criteria_mtime_before = Path(constant.CRITERIA_POOL_JSON).stat().st_mtime_ns

    defaults_view = candidate_service.get_prediction_filter_defaults_view("tmdb_RU_quality")
    missing_view = candidate_service.get_prediction_filter_defaults_view(None)
    unknown_view = candidate_service.get_prediction_filter_defaults_view("missing_criteria")

    pool_mtime_after = Path(constant.CANDIDATE_POOL_JSON).stat().st_mtime_ns
    criteria_mtime_after = Path(constant.CRITERIA_POOL_JSON).stat().st_mtime_ns

    expected_defaults = candidate_pool.build_prediction_filter_defaults("tmdb_RU_quality")
    expected_lines = candidate_pool.format_prediction_filter_default_lines(expected_defaults)

    assert_check(
        "get_prediction_filter_defaults_view возвращает те же defaults",
        defaults_view["defaults"] == expected_defaults,
    )
    assert_check(
        "get_prediction_filter_defaults_view возвращает те же lines",
        defaults_view["lines"] == expected_lines,
    )
    assert_check(
        "has_defaults=True для выбранного criteria_name",
        defaults_view["has_defaults"] is True,
    )
    assert_check(
        "criteria_name=None даёт базовые defaults без ошибки",
        missing_view["defaults"] == candidate_pool.build_prediction_filter_defaults(None),
    )
    assert_check(
        "has_defaults=False без criteria_name",
        missing_view["has_defaults"] is False,
    )
    assert_check(
        "unknown criteria_name не ломает defaults view",
        unknown_view["defaults"] == candidate_pool.build_prediction_filter_defaults("missing_criteria"),
    )
    assert_check(
        "defaults view не переписывает candidate_pool.json",
        pool_mtime_before == pool_mtime_after,
    )
    assert_check(
        "defaults view не переписывает candidate_criteria.json",
        criteria_mtime_before == criteria_mtime_after,
    )


def test_candidate_service_mark_watched_in_pool() -> None:
    """Проверяет write-flow удаления просмотренного кандидата через candidates.service."""
    print("\n25) Проверяем candidates.service mark watched in pool")

    candidate_pool.save_candidate_pool({
        "one": {
            "title": "Ева, рожай!",
            "alternative_title": "",
            "year": 2022,
            "criteria_name": "A",
            "kp_score": 7.0,
            "kp_votes": 1000,
            "imdb_score": 0,
            "imdb_votes": 0,
            "genres": ["комедия"],
        },
        "two": {
            "title": "Ева рожай",
            "alternative_title": "",
            "year": 2022,
            "criteria_name": "B",
            "kp_score": 8.0,
            "kp_votes": 2000,
            "imdb_score": 0,
            "imdb_votes": 0,
            "genres": ["комедия"],
        },
        "three": {
            "title": "Другой сериал",
            "alternative_title": "",
            "year": 2022,
            "criteria_name": "A",
            "kp_score": 8.0,
            "kp_votes": 500,
            "imdb_score": 0,
            "imdb_votes": 0,
            "genres": ["драма"],
        },
    })

    target_candidate = {
        "title": "Ева рожай",
        "alternative_title": "",
        "year": 2022,
        "criteria_name": "X",
    }

    with patch("candidates.candidate_pool.remove_candidate_from_pool", wraps=candidate_pool.remove_candidate_from_pool) as remove_mock:
        result = candidate_service.mark_candidate_watched_in_pool(target_candidate)

    pool = candidate_pool.load_candidate_pool()

    assert_check("Service делегирует в remove_candidate_from_pool", remove_mock.call_count == 1)
    assert_check("Service removed=True при cross-criteria match", result["removed"] is True)
    assert_check("Service удаляет обе fuzzy-записи", result["removed_count"] == 2)
    assert_check("После service mark watched в pool остался 1 кандидат", len(pool) == 1)
    assert_check(
        "В pool остался только другой сериал",
        next(iter(pool.values())).get("title") == "Другой сериал",
    )

    with patch("candidates.service.mark_candidate_watched_in_pool", wraps=candidate_service.mark_candidate_watched_in_pool) as service_mock:
        from dataset import dataset_records

        dataset_records._cleanup_candidate_pool(target_candidate)

    assert_check(
        "Mark watched cleanup path использует candidates.service",
        service_mock.call_count == 1,
    )


def test_candidate_service_retry_kp_enrichment() -> None:
    """Проверяет retry KP view/write-flow через candidates.service."""
    print("\n26) Проверяем candidates.service retry KP enrichment")

    candidate_pool.save_named_criteria("legacy", {"country": "Россия"})
    candidate_pool.save_candidate_pool({
        "one": {
            "title": "Retry Service Candidate",
            "alternative_title": "",
            "year": 2020,
            "criteria_name": "legacy",
            "kp_score": None,
            "kp_votes": None,
            "imdb_score": 7.3,
            "imdb_votes": 4200,
            "kp_status": "missing",
        },
    })

    retry_view = candidate_service.get_retry_kp_view()
    assert_check("get_retry_kp_view видит incomplete candidate", retry_view["incomplete_count"] == 1)
    assert_check(
        "get_retry_kp_view совпадает с get_incomplete_candidates",
        retry_view["incomplete_candidates"]
        == candidate_pool.get_incomplete_candidates(candidate_pool.load_candidate_pool()),
    )

    kp_movie = {
        "id": 777,
        "name": "Retry Service Candidate",
        "year": 2020,
        "rating": {"kp": 8.1},
        "votes": {"kp": 12000},
        "description": "Filled from KP",
    }

    with patch("candidates.kp_enrichment.kp_match_is_safe", return_value=(True, None)):
        with patch("candidates.candidate_pool.api.find_series_raw", return_value={"ok": True, "data": kp_movie}):
            with patch(
                "candidates.candidate_pool.retry_kp_enrichment_for_pool",
                wraps=candidate_pool.retry_kp_enrichment_for_pool,
            ) as retry_mock:
                result = candidate_service.retry_kp_enrichment_in_pool(limit=1)

    assert_check("Service делегирует в retry_kp_enrichment_for_pool", retry_mock.call_count == 1)
    assert_check("Service retry saved_pool=True при попытке", result["saved_pool"] is True)
    assert_check("Service retry kp_found=1", result["stats"]["kp_found"] == 1)

    pool = candidate_pool.load_candidate_pool()
    candidate = next(item for item in pool.values() if item.get("title") == "Retry Service Candidate")
    assert_check("После service retry кандидат complete", candidate["is_complete"] is True)
    assert_check("После service retry kp_status=done", candidate["kp_status"] == "done")


def test_candidate_service_tmdb_import_result() -> None:
    """Проверяет TMDb import view/preview/write-flow через candidates.service."""
    print("\n27) Проверяем candidates.service TMDb import result")

    country_output = []
    country_codes = interface_funcs.request_tmdb_country_codes(
        input_func=lambda _prompt: "1",
        output_func=country_output.append,
    )
    country_output_text = "\n".join(country_output)
    assert_check(
        "TMDb country options содержат русские названия для кодов",
        tmdb_country_options.COUNTRY_NAMES_RU_BY_CODE["RU"] == "Россия"
        and tmdb_country_options.COUNTRY_NAMES_RU_BY_CODE["US"] == "США"
        and tmdb_country_options.COUNTRY_NAMES_RU_BY_CODE["GB"] == "Великобритания",
    )
    assert_check(
        "TMDb country parser возвращает ISO-коды по номерам",
        country_codes == ["RU"] and tmdb_country_options.parse_country_indexes("1,2,3") == ["RU", "US", "GB"],
    )
    assert_check(
        "TMDb country UI просит номера стран",
        "Введите номера стран, по которым будет производиться поиск:" in country_output_text
        and "Список:" in country_output_text,
    )
    assert_check(
        "TMDb country UI показывает названия без ISO-кодов",
        "1. Россия" in country_output_text
        and "2. США" in country_output_text
        and "3. Великобритания" in country_output_text
        and "Россия RU" not in country_output_text,
    )
    country_list_lines = [
        line
        for line in country_output
        if "." in line and "Список:" not in line
    ]
    assert_check(
        "TMDb country UI разбивает список стран по 5 пунктов в строке",
        len(country_list_lines) > 1
        and all(line.count(".") <= 5 for line in country_list_lines),
    )

    import inspect

    flow_source = inspect.getsource(interface_funcs.import_tmdb_result_to_common_pool_flow)
    assert_check(
        "UI import flow вызывает candidate_service.import_tmdb_result_to_pool",
        "candidate_service.import_tmdb_result_to_pool" in flow_source,
    )
    assert_check(
        "UI import flow не вызывает import_tmdb напрямую",
        "import_tmdb_result_to_common_pool(" not in flow_source,
    )

    tmdb_dir = Path(constant.DATA_DIR) / "candidate_pool"
    tmdb_dir.mkdir(parents=True, exist_ok=True)
    result_path = tmdb_dir / "tmdb_candidate_pool_service_test.json"
    result = {
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "candidates": [
            {
                "title": "Service Import Candidate",
                "year": 2023,
                "tmdb_id": 501,
                "tmdb_rating": 7.5,
                "tmdb_votes": 100,
                "imdb_rating": 7.0,
                "imdb_votes": 5000,
            }
        ],
    }
    result_path.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

    with patch("candidates.import_tmdb.OUTPUT_DIR", tmdb_dir):
        files_view = candidate_service.get_tmdb_import_files_view()
        assert_check("get_tmdb_import_files_view не пустой", files_view["is_empty"] is False)
        assert_check(
            "get_tmdb_import_files_view находит result file",
            result_path.name in files_view["file_names"],
        )

        missing_preview = candidate_service.load_tmdb_result_import_preview(tmdb_dir / "missing.json")
        assert_check("preview missing file ok=False", missing_preview["ok"] is False)

        bad_path = tmdb_dir / "tmdb_candidate_pool_bad.json"
        bad_path.write_text(json.dumps({"candidates": "not a list"}), encoding="utf-8")
        bad_preview = candidate_service.load_tmdb_result_import_preview(bad_path)
        assert_check("preview invalid candidates ok=False", bad_preview["ok"] is False)

        preview = candidate_service.load_tmdb_result_import_preview(result_path)
        assert_check("preview ok=True", preview["ok"] is True)
        assert_check("preview candidate_count=1", preview["candidate_count"] == 1)
        assert_check(
            "preview default_criteria_name",
            preview["default_criteria_name"] == "tmdb_RU_quality",
        )

        with patch(
            "candidates.import_tmdb.import_tmdb_result_to_common_pool",
            wraps=tmdb_import.import_tmdb_result_to_common_pool,
        ) as import_mock:
            import_result = candidate_service.import_tmdb_result_to_pool(
                result_path,
                criteria_name="tmdb_RU_quality",
            )

    assert_check("Service делегирует в import_tmdb_result_to_common_pool", import_mock.call_count == 1)
    assert_check("Service import ok=True", import_result["ok"] is True)
    assert_check("Service import stats read=1", import_result["stats"]["read"] == 1)
    assert_check("Service import возвращает result_file", import_result["result_file"] == str(result_path))
    assert_check(
        "Service import возвращает criteria_name",
        import_result["criteria_name"] == "tmdb_RU_quality",
    )

    pool = candidate_pool.load_candidate_pool()
    assert_check(
        "После service import кандидат в pool",
        any(item.get("title") == "Service Import Candidate" for item in pool.values()),
    )

    criteria = candidate_pool.load_candidate_criteria()["tmdb_RU_quality"]
    assert_check("Service import metadata source", criteria["source"] == "tmdb_imdb_kp_v1")
    assert_check("Service import metadata country", criteria["country"] == "RU")
    assert_check("Service import metadata mode", criteria["mode"] == "quality")
    assert_check(
        "Service import metadata result_file",
        criteria["result_file"].endswith("tmdb_candidate_pool_service_test.json"),
    )
    assert_check("Service import metadata candidate_count", criteria["candidate_count"] == 1)

    candidate_pool.save_named_criteria("tmdb_RU_quality", {
        "criteria_name": "tmdb_RU_quality",
        "genres": ["драма"],
        "excluded_genres": ["мелодрама"],
        "min_kp": 7.0,
        "max_kp": 9.0,
        "min_imdb": 6.5,
        "max_imdb": 8.5,
        "min_year": 2010,
        "max_year": 2025,
        "custom_note": "service import filters",
    })
    second_result_path = Path(constant.DATA_DIR) / "tmdb_service_filters.json"
    second_result_path.write_text(json.dumps({
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "candidates": [
            {
                "title": "Service Filters Candidate",
                "year": 2024,
                "tmdb_id": 502,
                "tmdb_rating": 7.2,
                "tmdb_votes": 60,
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    candidate_service.import_tmdb_result_to_pool(second_result_path, criteria_name="tmdb_RU_quality")
    criteria = candidate_pool.load_candidate_criteria()["tmdb_RU_quality"]
    assert_check("Service import сохраняет genres", criteria["genres"] == ["драма"])
    assert_check("Service import сохраняет excluded_genres", criteria["excluded_genres"] == ["мелодрама"])
    assert_check("Service import сохраняет min_kp", criteria["min_kp"] == 7.0)
    assert_check("Service import сохраняет max_kp", criteria["max_kp"] == 9.0)
    assert_check("Service import сохраняет min_imdb", criteria["min_imdb"] == 6.5)
    assert_check("Service import сохраняет max_imdb", criteria["max_imdb"] == 8.5)
    assert_check("Service import сохраняет min_year", criteria["min_year"] == 2010)
    assert_check("Service import сохраняет max_year", criteria["max_year"] == 2025)
    assert_check("Service import сохраняет custom unknown field", criteria["custom_note"] == "service import filters")

    other_result_path = Path(constant.DATA_DIR) / "tmdb_service_cross_criteria.json"
    other_result_path.write_text(json.dumps({
        "criteria_name": "tmdb_US_quality",
        "country": "US",
        "mode": "quality",
        "candidates": [
            {
                "title": "Service Import Candidate",
                "year": 2023,
                "tmdb_id": 503,
                "tmdb_rating": 7.1,
                "tmdb_votes": 40,
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")
    candidate_service.import_tmdb_result_to_pool(other_result_path, criteria_name="tmdb_US_quality")
    pool = candidate_pool.load_candidate_pool()
    cross_criteria_keys = [
        key for key, item in pool.items()
        if item.get("title") == "Service Import Candidate" and item.get("year") == 2023
    ]
    assert_check("Service import cross-criteria хранит 2 записи", len(cross_criteria_keys) == 2)

    storage_movie.add_movie(make_movie(title="Service Watched Import", user_score=8.0))
    watched_result_path = Path(constant.DATA_DIR) / "tmdb_service_watched.json"
    watched_result_path.write_text(json.dumps({
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "candidates": [
            {
                "title": "Service Watched Import",
                "year": 2024,
                "tmdb_id": 504,
                "tmdb_rating": 7.0,
                "tmdb_votes": 30,
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")
    watched_result = candidate_service.import_tmdb_result_to_pool(
        watched_result_path,
        criteria_name="tmdb_RU_quality",
    )
    pool = candidate_pool.load_candidate_pool()
    assert_check("Service import watched_skipped=1", watched_result["stats"]["watched_skipped"] == 1)
    assert_check(
        "Service import не сохраняет watched candidate",
        all(item.get("title") != "Service Watched Import" for item in pool.values()),
    )


def test_candidate_service_tmdb_build_and_auto_import() -> None:
    """Проверяет TMDb build/save/auto-import facade через candidates.service."""
    print("\n28) Проверяем candidates.service TMDb build and auto-import")

    import inspect

    build_flow_source = inspect.getsource(interface_funcs.run_tmdb_candidate_pool_flow)
    auto_import_source = inspect.getsource(interface_funcs.maybe_auto_import_tmdb_result)
    assert_check(
        "UI build flow вызывает candidate_service.build_tmdb_candidate_pool",
        "candidate_service.build_tmdb_candidate_pool" in build_flow_source,
    )
    assert_check(
        "UI build flow вызывает candidate_service.save_tmdb_build_result",
        "candidate_service.save_tmdb_build_result" in build_flow_source,
    )
    assert_check(
        "UI build flow не вызывает build_candidate_pool напрямую",
        "build_candidate_pool(" not in build_flow_source,
    )
    assert_check(
        "UI build flow переименовывает режимы поиска",
        "Поиск по популярным" in build_flow_source
        and "Поиск по недооценённым" in build_flow_source
        and "Лучшие по качеству" not in build_flow_source
        and "Скрытые находки" not in build_flow_source,
    )
    assert_check(
        "Auto-import helper по умолчанию использует candidate_service.import_tmdb_result_to_pool",
        "candidate_service.import_tmdb_result_to_pool" in auto_import_source,
    )

    build_result = {
        "country": "RU",
        "mode": "quality",
        "criteria_name": "tmdb_RU_quality",
        "stats": {"discover_total": 2, "final_candidates": 1},
        "candidates": [{"title": "Build Service Candidate", "year": 2022}],
    }
    fake_json_path = Path(constant.DATA_DIR) / "candidate_pool_RU_quality.json"
    fake_csv_path = fake_json_path.with_suffix(".csv")

    with patch("candidates.tmdb_candidate_pool.build_candidate_pool", return_value=build_result) as build_mock:
        result = candidate_service.build_tmdb_candidate_pool(
            country="RU",
            pages=1,
            details_limit=5,
            mode="quality",
            criteria_name="tmdb_RU_quality",
        )
    assert_check("Service build делегирует в build_candidate_pool", build_mock.call_count == 1)
    assert_check("Service build возвращает snapshot", result["criteria_name"] == "tmdb_RU_quality")

    with patch(
        "candidates.tmdb_candidate_pool.save_candidate_pool_result",
        return_value=(fake_json_path, fake_csv_path),
    ) as save_mock:
        save_result = candidate_service.save_tmdb_build_result(build_result, is_test_run=False)
    assert_check("Service save делегирует в save_candidate_pool_result", save_mock.call_count == 1)
    assert_check("Service save возвращает json_path", save_result["json_path"] == fake_json_path)

    with patch("candidates.tmdb_candidate_pool.build_candidate_pool", return_value=build_result):
        with patch(
            "candidates.tmdb_candidate_pool.save_candidate_pool_result",
            return_value=(fake_json_path, fake_csv_path),
        ):
            combined = candidate_service.build_and_save_tmdb_candidate_pool(
                country="RU",
                pages=1,
                details_limit=5,
                mode="quality",
                criteria_name="tmdb_RU_quality",
                is_test_run=False,
            )
    assert_check("build_and_save ok=True", combined["ok"] is True)
    assert_check("build_and_save возвращает paths", combined["json_path"] == fake_json_path)
    assert_check("build_and_save возвращает candidates", len(combined["candidates"]) == 1)

    criteria_name = candidate_service.build_tmdb_criteria_name("RU", "quality", year_min=2020, min_tmdb_score=7.5)
    assert_check(
        "Service build_tmdb_criteria_name совпадает с tmdb_candidate_pool",
        criteria_name == tmdb_candidate_pool.build_tmdb_criteria_name("RU", "quality", year_min=2020, min_tmdb_score=7.5),
    )

    import_calls = []

    def capture_import(result_path, criteria_name=None):
        import_calls.append((result_path, criteria_name))
        return {
            "ok": True,
            "stats": {
                "ok": True,
                "read": 1,
                "added": 1,
                "updated": 0,
                "watched_skipped": 0,
                "skipped_watched": 0,
                "duplicates": 0,
                "skipped_duplicates": 0,
                "errors": 0,
                "criteria_name": criteria_name,
                "pool_size_before": 0,
                "pool_size_after": 1,
                "pool_size": 1,
            },
            "result_file": str(result_path),
            "criteria_name": criteria_name,
            "error": None,
        }

    with patch("candidates.service.import_tmdb_result_to_pool", side_effect=capture_import) as service_import_mock:
        stats = interface_funcs.maybe_auto_import_tmdb_result(
            fake_json_path,
            "tmdb_RU_quality",
            input_func=lambda _prompt: "",
            output_func=lambda _line: None,
        )
    assert_check("Auto-import default вызывает service import", service_import_mock.call_count == 1)
    assert_check(
        "Auto-import default передаёт path и criteria_name",
        import_calls == [(fake_json_path, "tmdb_RU_quality")],
    )
    assert_check("Auto-import default возвращает stats", stats["ok"] is True)


def test_candidate_service_polish_flows() -> None:
    """Проверяет J5 polish facade: menu stats, mark watched view, delete, duplicates."""
    print("\n29) Проверяем candidates.service polish flows")

    import inspect

    global_menu_source = inspect.getsource(__import__("ui.console.global_menu", fromlist=["global_menu"]))
    mark_watched_source = inspect.getsource(interface_funcs.mark_candidate_as_watched)
    delete_source = inspect.getsource(interface_funcs.delete_candidate_pool)
    duplicates_source = inspect.getsource(interface_funcs.show_suspicious_candidate_duplicates)

    assert_check(
        "global_menu использует candidate_service.get_pool_stats_view",
        "candidate_service.get_pool_stats_view" in global_menu_source,
    )
    assert_check(
        "mark_candidate_as_watched использует get_mark_watched_view",
        "candidate_service.get_mark_watched_view" in mark_watched_source,
    )
    assert_check(
        "delete_candidate_pool использует delete_candidate_pool_criteria",
        "candidate_service.delete_candidate_pool_criteria" in delete_source,
    )
    assert_check(
        "show_suspicious_candidate_duplicates использует get_suspicious_duplicates_view",
        "candidate_service.get_suspicious_duplicates_view" in duplicates_source,
    )

    top_prediction_source = inspect.getsource(interface_funcs.show_global_candidate_top)
    contributions_source = inspect.getsource(interface_funcs.show_candidate_contributions)
    collect_legacy_source = inspect.getsource(interface_funcs.collect_candidate_pool)

    assert_check(
        "show_global_candidate_top использует rank_top_prediction_candidates",
        "candidate_service.rank_top_prediction_candidates" in top_prediction_source,
    )
    assert_check(
        "show_candidate_contributions использует build_contribution_reports",
        "candidate_service.build_contribution_reports" in contributions_source,
    )
    assert_check(
        "collect_candidate_pool использует collect_candidates_legacy",
        "candidate_service.collect_candidates_legacy" in collect_legacy_source,
    )
    assert_check(
        "interface_funcs не импортирует candidate_pool напрямую",
        "from candidates import candidate_pool" not in inspect.getsource(interface_funcs),
    )

    candidate_pool.save_named_criteria("legacy", {"country": "Россия"})
    candidate_pool.save_candidate_pool({
        "legacy|Polish Service One|2020": {
            "title": "Polish Service One",
            "year": 2020,
            "criteria_name": "legacy",
            "kp_score": 7.0,
            "kp_votes": 1000,
            "imdb_score": 7.1,
            "imdb_votes": 2000,
        },
        "legacy|Polish Service Two|2021": {
            "title": "Polish Service Two",
            "year": 2021,
            "criteria_name": "legacy",
            "kp_score": None,
            "kp_votes": None,
            "imdb_score": 7.0,
            "imdb_votes": 1500,
            "kp_status": "missing",
        },
        "other|Polish Service Other|2022": {
            "title": "Polish Service Other",
            "year": 2022,
            "criteria_name": "other",
            "kp_score": 7.5,
            "kp_votes": 900,
            "imdb_score": 7.4,
            "imdb_votes": 1800,
        },
    })

    watched_view = candidate_service.get_mark_watched_view("legacy")
    assert_check("get_mark_watched_view видит 2 candidates", len(watched_view["candidates"]) == 2)
    assert_check(
        "get_mark_watched_view lines совпадают с get_pool_stats_view",
        watched_view["lines"] == candidate_service.get_pool_stats_view("legacy")["lines"],
    )

    incomplete_candidate = next(
        item for item in watched_view["candidates"] if item.get("title") == "Polish Service Two"
    )
    assert_check(
        "is_pool_candidate_incomplete совпадает с candidate_pool",
        candidate_service.is_pool_candidate_incomplete(incomplete_candidate)
        == candidate_pool.is_candidate_incomplete(incomplete_candidate),
    )

    with patch(
        "candidates.candidate_pool.delete_criteria_and_candidates",
        wraps=candidate_pool.delete_criteria_and_candidates,
    ) as delete_mock:
        delete_result = candidate_service.delete_candidate_pool_criteria("legacy")

    assert_check("delete_candidate_pool_criteria делегирует в delete_criteria_and_candidates", delete_mock.call_count == 1)
    assert_check("delete_candidate_pool_criteria deleted=True", delete_result["deleted"] is True)
    assert_check("delete_candidate_pool_criteria удаляет 2 candidates", delete_result["deleted_candidates"] == 2)

    remaining_pool = candidate_pool.load_candidate_pool()
    assert_check("После delete остаётся только other criteria", len(remaining_pool) == 1)
    assert_check(
        "После delete остаётся Polish Service Other",
        next(iter(remaining_pool.values())).get("title") == "Polish Service Other",
    )

    candidate_pool.save_candidate_pool({
        "dup_a|Polish Dup Alpha|2020": {
            "title": "Polish Dup Alpha",
            "year": 2020,
            "criteria_name": "dup_a",
        },
        "dup_b|Polish Dup Alphа|2020": {
            "title": "Polish Dup Alphа",
            "year": 2020,
            "criteria_name": "dup_b",
        },
    })

    with patch(
        "candidates.candidate_pool.find_suspicious_duplicates",
        wraps=candidate_pool.find_suspicious_duplicates,
    ) as duplicates_mock:
        duplicates_view = candidate_service.get_suspicious_duplicates_view()

    assert_check("get_suspicious_duplicates_view делегирует в find_suspicious_duplicates", duplicates_mock.call_count == 1)
    assert_check("get_suspicious_duplicates_view count совпадает", duplicates_view["count"] == len(duplicates_view["pairs"]))
    assert_check(
        "get_suspicious_duplicates_view is_empty корректен",
        duplicates_view["is_empty"] == (duplicates_view["count"] == 0),
    )


def test_candidate_schema_normalizes_legacy_complete_record() -> None:
    """Проверяет legacy backfill для complete-кандидата без is_complete/kp_status."""
    print("\n15) Проверяем schema backfill для complete legacy candidate")

    normalized = candidate_schema.normalize_candidate_record({
        "title": "Legacy Complete",
        "year": 2020,
        "kp_score": 7.2,
        "kp_votes": 1500,
        "imdb_score": 7.4,
        "imdb_votes": 7000,
        "extra_field": "keep_me",
    })

    assert_check("Legacy candidate получает criteria_name=legacy", normalized["criteria_name"] == "legacy")
    assert_check("Legacy candidate получает source=legacy", normalized["source"] == "legacy")
    assert_check("Legacy candidate становится complete", normalized["is_complete"] is True)
    assert_check("Legacy candidate получает kp_status=done", normalized["kp_status"] == "done")
    assert_check("Unknown fields сохраняются", normalized["extra_field"] == "keep_me")


def test_candidate_schema_marks_missing_kp() -> None:
    """Проверяет incomplete-кандидата без KP score/votes."""
    print("\n16) Проверяем schema completeness без KP")

    candidate = candidate_schema.normalize_candidate_record({
        "title": "No KP Yet",
        "year": 2021,
        "imdb_score": 7.1,
        "imdb_votes": 5000,
        "kp_score": None,
        "kp_votes": None,
    })
    completeness = candidate_schema.compute_completeness(candidate)

    assert_check("Candidate без KP не complete", candidate["is_complete"] is False)
    assert_check("missing_fields содержит kp_score", "kp_score" in completeness["missing_fields"])
    assert_check("missing_fields содержит kp_votes", "kp_votes" in completeness["missing_fields"])
    assert_check("kp_status не done без KP данных", completeness["kp_status"] != "done")


def test_candidate_schema_preserves_specific_kp_status() -> None:
    """Проверяет сохранение конкретного KP-статуса при отсутствии KP данных."""
    print("\n17) Проверяем сохранение конкретного kp_status")

    candidate = candidate_schema.normalize_candidate_record({
        "title": "KP Not Found",
        "year": 2021,
        "kp_status": "not_found",
        "kp_score": None,
        "kp_votes": None,
        "imdb_score": 7.0,
        "imdb_votes": 3000,
    })

    assert_check("kp_status=not_found сохраняется", candidate["kp_status"] == "not_found")
    assert_check("Candidate со статусом not_found остаётся incomplete", candidate["is_complete"] is False)


def test_candidate_schema_ready_for_predict_requires_kp_and_imdb() -> None:
    """Проверяет readiness gate по единой schema-функции."""
    print("\n18) Проверяем readiness gate schema")

    ready_candidate = {
        "title": "Ready",
        "year": 2020,
        "kp_score": 7.5,
        "kp_votes": 1000,
        "imdb_score": 7.4,
        "imdb_votes": 5000,
    }
    incomplete_candidate = {
        "title": "Incomplete",
        "year": 2020,
        "kp_score": 7.5,
        "kp_votes": 1000,
        "imdb_score": 7.4,
        "imdb_votes": None,
    }

    assert_check("is_ready_for_predict=True только при полном KP+IMDb", candidate_schema.is_ready_for_predict(ready_candidate) is True)
    assert_check("is_ready_for_predict=False если не хватает поля", candidate_schema.is_ready_for_predict(incomplete_candidate) is False)


def test_kp_country_alias_matching() -> None:
    """Проверяет alias matching для Южной Кореи без ослабления других стран."""
    print("\n19a) Проверяем KP country alias matching")

    assert_check(
        "Южная Корея matches South Korea",
        kp_enrichment.countries_match("Южная Корея", ["South Korea"]) is True,
    )
    assert_check(
        "Южная Корея matches Republic of Korea",
        kp_enrichment.countries_match("Южная Корея", ["Republic of Korea"]) is True,
    )
    assert_check(
        "Южная Корея matches Korea, Republic of",
        kp_enrichment.countries_match("Южная Корея", ["Korea, Republic of"]) is True,
    )
    assert_check(
        "KR matches Южная Корея",
        kp_enrichment.countries_match("KR", ["Южная Корея"]) is True,
    )
    assert_check(
        "Япония не matches South Korea",
        kp_enrichment.countries_match("Япония", ["South Korea"]) is False,
    )
    assert_check(
        "Россия не matches South Korea",
        kp_enrichment.countries_match("Россия", ["South Korea"]) is False,
    )
    assert_check(
        "Россия matches KP label Россия",
        kp_enrichment.countries_match(
            "Россия",
            kp_enrichment.extract_kp_country_values({"countries": [{"name": "Россия"}]}),
        ) is True,
    )
    assert_check(
        "Япония matches KP label Япония",
        kp_enrichment.countries_match(
            "Япония",
            kp_enrichment.extract_kp_country_values({"countries": [{"name": "Япония"}]}),
        ) is True,
    )


def test_kp_lookup_kr_country_aliases() -> None:
    """Проверяет, что KR lookup с South Korea в KP не падает в country_not_found."""
    print("\n19b) Проверяем KR KP lookup country aliases")

    candidate = {
        "title": "Гоблин",
        "original_title": "Guardian: The Lonely and Great God",
        "year": 2016,
    }
    kp_movie = {
        "id": 1005016,
        "name": "Гоблин",
        "year": 2016,
        "type": "tv-series",
        "countries": [{"name": "South Korea"}],
        "rating": {"kp": 8.5},
        "votes": {"kp": 12000},
    }

    def fake_fetch_json(url, **kwargs):
        return {"ok": True, "data": {"docs": [kp_movie]}}

    with patch("apis.kp_api.fetch_json", side_effect=fake_fetch_json):
        lookup = kp_enrichment.lookup_kp_via_api(
            candidate,
            ["Гоблин"],
            "Южная Корея",
        )

    assert_check("KR lookup status=found", lookup["status"] == "found")
    assert_check("KR lookup не country_not_found", lookup.get("error") != "country_not_found")
    assert_check("KR lookup возвращает movie", lookup.get("movie") is not None)


def test_kp_country_from_iso2_mapping() -> None:
    """Проверяет перевод ISO-2 TMDb country в русское название для KP API."""
    print("\n19) Проверяем kp_country_from_iso2 mapping")

    assert_check("JP -> Япония", kp_enrichment.kp_country_from_iso2("JP") == "Япония")
    assert_check("FR -> Франция", kp_enrichment.kp_country_from_iso2("FR") == "Франция")
    assert_check("UA -> Украина", kp_enrichment.kp_country_from_iso2("UA") == "Украина")
    assert_check("US -> США", kp_enrichment.kp_country_from_iso2("US") == "США")
    assert_check("Неизвестный ISO -> пустая строка", kp_enrichment.kp_country_from_iso2("XX") == "")
    assert_check(
        "Неизвестный ISO не превращается в Россия",
        kp_enrichment.kp_country_from_iso2("XX") != "Россия",
    )

    for code in tmdb_country_options.COUNTRY_CODE_ORDER:
        expected_label = tmdb_country_options.COUNTRY_NAMES_RU_BY_CODE[code]
        assert_check(
            f"KP mapping покрывает TMDb UI country {code}",
            kp_enrichment.kp_country_from_iso2(code) == expected_label,
        )


def test_kp_series_type_filter() -> None:
    """Проверяет расширенный KP type-фильтр для anime/animated-series без ослабления match-check."""
    print("\n20) Проверяем KP series type filter")

    tv_series = {"type": "tv-series", "name": "Regular Series"}
    anime_series = {"type": "anime", "name": "Anime Series", "isSeries": True}
    animated_series = {"type": "animated-series", "name": "Animated Series"}
    cartoon_serial = {
        "type": "cartoon",
        "name": "Cartoon Serial",
        "seasonsInfo": [{"number": 1}],
    }
    cartoon_film = {"type": "cartoon", "name": "Cartoon Film"}
    movie = {"type": "movie", "name": "Feature Film"}
    film = {"type": "film", "name": "Another Film"}

    assert_check("tv-series проходит type-фильтр", api.is_series(tv_series) is True)
    assert_check("anime проходит type-фильтр", api.is_series(anime_series) is True)
    assert_check("animated-series проходит type-фильтр", api.is_series(animated_series) is True)
    assert_check("cartoon с serial markers проходит", api.is_series(cartoon_serial) is True)
    assert_check("cartoon без serial markers блокируется", api.is_series(cartoon_film) is False)
    assert_check("movie блокируется", api.is_series(movie) is False)
    assert_check("film блокируется", api.is_series(film) is False)

    candidate = {
        "title": "Наруто",
        "original_title": "ナルト",
        "year": 2002,
    }
    wrong_anime = {
        "name": "Совсем другой anime",
        "year": 2002,
        "type": "anime",
    }
    assert_check("anime type не bypass-ит match-check", api.is_series(wrong_anime) is True)
    is_safe, reason = kp_enrichment.kp_match_is_safe(candidate, wrong_anime)
    assert_check("anime type всё ещё требует title match", is_safe is False)
    assert_check("anime type всё ещё требует title mismatch reason", reason == "title_mismatch")


def test_kp_tmdb_build_debug_traces_rejection_details() -> None:
    """Проверяет TEMP KP debug traces: query, counts, rejection summary."""
    print("\n20) Проверяем KP TMDb build debug traces")

    candidate = {
        "title": "Offline Alpha Series",
        "original_title": "Offline Alpha Original",
        "year": 2013,
    }
    kp_movie = {
        "id": 9001,
        "name": "Completely Different Beta",
        "year": 2013,
        "type": "tv-series",
        "countries": [{"name": "Япония"}],
    }
    traces: list[dict] = []

    def fake_fetch_json(url, **kwargs):
        return {"ok": True, "data": {"docs": [kp_movie]}}

    with patch("apis.kp_api.fetch_json", side_effect=fake_fetch_json):
        lookup = kp_enrichment.lookup_kp_via_api(
            candidate,
            ["Offline Alpha Series"],
            "Япония",
            attempt_traces=traces,
        )

    assert_check("Lookup остаётся rejected", lookup["status"] == "rejected")
    assert_check("Debug trace записан", len(traces) == 1)
    trace = traces[0]
    assert_check("Trace содержит candidate title", trace["candidate_title"] == "Offline Alpha Series")
    assert_check("Trace содержит KP query", trace["kp_query"] == "Offline Alpha Series")
    assert_check("Trace содержит KP country", trace["kp_country"] == "Япония")
    assert_check("Trace содержит series count", trace["kp_series_count"] == 1)
    assert_check("Trace содержит KP selected title", trace["kp_selected_title"] == "Completely Different Beta")
    assert_check("Trace содержит reject_reason", trace["reject_reason"] == "title_mismatch")
    assert_check("Trace содержит similarity", trace["title_similarity"] is not None)
    assert_check(
        "Trace summary объясняет rejection",
        "title_mismatch" in str(trace.get("rejection_summary") or ""),
    )


def test_kp_tmdb_build_debug_session_save() -> None:
    """Проверяет сохранение отдельного *_kp_debug.json рядом с build result."""
    print("\n21) Проверяем сохранение KP debug report")

    session = kp_tmdb_build_debug.KpBuildDebugSession(country="JP", criteria_name="test_jp")
    session.start_candidate({"title": "Debug Save", "original_title": "Debug", "year": 2020}, "Япония")
    session.finish_candidate(
        {"title": "Debug Save", "kp_status": "not_found"},
        {"status": "not_found", "error": "not_found"},
    )
    report = session.to_report()

    with tempfile.TemporaryDirectory() as temp_dir:
        json_path = Path(temp_dir) / "candidate_pool_JP_quality.json"
        json_path.write_text("{}", encoding="utf-8")
        debug_path = kp_tmdb_build_debug.save_kp_debug_report(report, json_path)
        assert_check("Debug report сохранён", debug_path.is_file())
        saved = json.loads(debug_path.read_text(encoding="utf-8"))
        assert_check("Debug report содержит entries", len(saved.get("entries") or []) == 1)
        assert_check("Debug report содержит attempts", len(saved["entries"][0].get("attempts") or []) == 0)


def test_retry_kp_enrichment_makes_candidate_complete() -> None:
    """Проверяет, что retry KP переводит кандидата в complete после успешного fill."""
    print("\n20) Проверяем retry KP -> complete")

    candidate_pool.save_candidate_pool({
        "one": {
            "title": "Retry Candidate",
            "alternative_title": "",
            "year": 2020,
            "criteria_name": "legacy",
            "kp_score": None,
            "kp_votes": None,
            "imdb_score": 7.3,
            "imdb_votes": 4200,
            "kp_status": "missing",
        },
    })

    kp_movie = {
        "id": 555,
        "name": "Retry Candidate",
        "year": 2020,
        "rating": {"kp": 8.1},
        "votes": {"kp": 12000},
        "description": "Filled from KP",
    }

    with patch("candidates.kp_enrichment.kp_match_is_safe", return_value=(True, None)):
        with patch("candidates.candidate_pool.api.find_series_raw", return_value={"ok": True, "data": kp_movie}):
            stats = candidate_pool.retry_kp_enrichment_for_pool(limit=1)

    pool = candidate_pool.load_candidate_pool()
    candidate = next(
        item for item in pool.values()
        if item.get("title") == "Retry Candidate"
    )

    assert_check("Retry действительно нашёл KP", stats["kp_found"] == 1)
    assert_check("После retry кандидат complete", candidate["is_complete"] is True)
    assert_check("После retry kp_status=done", candidate["kp_status"] == "done")


def test_retry_kp_uses_candidate_country_when_criteria_missing() -> None:
    """Regression: retry KP prefers candidate country over missing criteria fallback."""
    print("\n20) Проверяем retry KP country fallback по candidate country")

    candidate_pool.save_candidate_criteria({})
    candidate_pool.save_candidate_pool({
        "one": {
            "title": "Гоблин",
            "alternative_title": "",
            "year": 2016,
            "criteria_name": "missing criteria",
            "country_codes": ["KR"],
            "country_display": "Южная Корея",
            "countries": ["KR", "South Korea"],
            "kp_score": None,
            "kp_votes": None,
            "imdb_score": 7.3,
            "imdb_votes": 4200,
            "kp_status": "missing",
        },
    })
    captured_countries = []

    def fake_lookup(candidate, queries, country, **_kwargs):
        captured_countries.append(country)
        return {
            "status": "not_found",
            "movie": None,
            "error": "not_found",
            "reject_reason": None,
            "query": None,
            "attempts": 1,
        }

    with patch("candidates.kp_enrichment.lookup_kp_via_api", side_effect=fake_lookup):
        stats = candidate_pool.retry_kp_enrichment_for_pool(limit=1)

    assert_check("Retry попытался добрать KP", stats["attempted"] == 1)
    assert_check("Retry вызвал KP lookup", captured_countries == ["Южная Корея"])
    assert_check("Retry не fallback-ит в Россию при KR candidate", captured_countries[0] != "Россия")


def test_candidate_schema_keeps_unknown_fields() -> None:
    """Проверяет, что schema-нормализация не удаляет неизвестные поля."""
    print("\n20) Проверяем сохранение unknown fields")

    normalized = candidate_schema.normalize_candidate_record({
        "title": "Unknown Fields",
        "year": 2022,
        "custom_blob": {"hello": "world"},
        "signals": ["keep"],
    })

    assert_check("Unknown dict field сохраняется", normalized["custom_blob"] == {"hello": "world"})
    assert_check("Signals сохраняются", normalized["signals"] == ["keep"])


def test_candidate_genre_schema_normalization() -> None:
    """Проверяет alias -> genre_key -> display label без мутации raw genres."""
    print("\n20a) Проверяем candidate genre_schema normalization")

    case_one = candidate_schema.normalize_candidate_record({
        "title": "Genre Case 1",
        "year": 2020,
        "genres": ["Drama", "драма"],
    })
    assert_check("Drama+драма -> один key drama", case_one["genre_keys"] == ["drama"])
    assert_check("Drama+драма -> один display Драма", case_one["genres_display"] == ["Драма"])

    case_two = candidate_schema.normalize_candidate_record({
        "title": "Genre Case 2",
        "year": 2020,
        "genres": ["Biography", "Drama", "History", "драма", "Боевик и Приключения"],
    })
    assert_check(
        "Mixed raw genres -> canonical keys без дублей",
        case_two["genre_keys"] == ["biography", "drama", "history", "action_adventure"],
    )
    assert_check(
        "Mixed raw genres -> display labels",
        case_two["genres_display"] == ["Биография", "Драма", "История", "Боевик/приключения"],
    )

    case_three = candidate_schema.normalize_candidate_record({
        "title": "Genre Case 3",
        "year": 2020,
        "genres": ["Comedy", "Romance", "драма", "комедия"],
    })
    assert_check(
        "Comedy/Romance/драма/комедия -> comedy, romance, drama",
        case_three["genre_keys"] == ["comedy", "romance", "drama"],
    )

    case_four = candidate_schema.normalize_candidate_record({
        "title": "Genre Case 4",
        "year": 2020,
        "genres": ["Action", "Adventure", "Боевик и Приключения"],
    })
    assert_check(
        "Action+Adventure -> один action_adventure",
        case_four["genre_keys"] == ["action_adventure"],
    )
    assert_check(
        "Action+Adventure -> один display label",
        case_four["genres_display"] == ["Боевик/приключения"],
    )

    unknown_case = candidate_schema.normalize_candidate_record({
        "title": "Genre Unknown",
        "year": 2020,
        "genres": ["Drama", "TotallyUnknownGenreXYZ"],
    })
    assert_check("Unknown genre не ломает normalize", unknown_case["genre_keys"] == ["drama"])
    assert_check("Unknown genre не попадает в display", unknown_case["genres_display"] == ["Драма"])

    legacy_raw = {
        "title": "Legacy Raw",
        "year": 2020,
        "genres": ["Drama", "Comedy"],
    }
    legacy_snapshot = copy.deepcopy(legacy_raw)
    normalized_legacy = candidate_schema.normalize_candidate_record(legacy_raw)
    assert_check("legacy candidate genres не мутируется", legacy_raw == legacy_snapshot)
    assert_check("legacy genres сохраняются в normalized copy", normalized_legacy["genres"] == ["Drama", "Comedy"])

    priority_case = candidate_schema.normalize_candidate_record({
        "title": "Priority Case",
        "year": 2020,
        "imdb_genres": ["Crime"],
        "genres_tmdb": ["Drama"],
        "genres": ["Comedy"],
    })
    assert_check(
        "imdb_genres имеет приоритет над tmdb/legacy",
        priority_case["genre_keys"] == ["crime", "drama", "comedy"],
    )


def test_build_candidate_features_uses_genre_keys() -> None:
    """build_candidate_features использует genre_keys/imdb_genres, а не пустой genres."""
    print("\n20a.0) Проверяем build_candidate_features через genre_keys")

    candidate = {
        "title": "Genre Keys Only",
        "year": 2020,
        "kp_score": 8.0,
        "kp_votes": 10000,
        "imdb_score": 8.0,
        "imdb_votes": 10000,
        "imdb_genres": ["Crime", "Drama"],
        "genres": [],
    }
    features = candidate_pool.build_candidate_features(candidate)

    assert_check("has_crime from imdb_genres/genre_keys", features.get("has_crime", 0) > 0)
    assert_check("has_drama from imdb_genres/genre_keys", features.get("has_drama", 0) > 0)
    assert_check("has_action stays 0", features.get("has_action", 0) == 0)
    assert_check("bias присутствует", features.get(constant.BIAS_FEATURE, 0) == 1.0)


def test_candidate_genre_keys_to_dataset_genres_maps_supported_keys() -> None:
    """Maps supported pool genre keys into the current dataset genre vector."""
    print("\n20a.1) Проверяем candidate genre_keys -> dataset genres")

    result = candidate_to_dataset.candidate_genre_keys_to_dataset_genres(
        ["drama", "crime", "comedy"]
    )
    dataset_genre = result["dataset_genre"]

    assert_check("status ok для supported keys", result["status"] == "ok")
    assert_check("has_drama = 1", dataset_genre["has_drama"] == 1)
    assert_check("has_crime = 1", dataset_genre["has_crime"] == 1)
    assert_check("has_comedy = 1", dataset_genre["has_comedy"] == 1)
    assert_check(
        "полный набор constant.GENRE",
        set(dataset_genre.keys()) == set(constant.GENRE),
    )
    assert_check(
        "остальные genre features остаются 0",
        all(
            dataset_genre[feature] == 0
            for feature in constant.GENRE
            if feature not in {"has_drama", "has_crime", "has_comedy"}
        ),
    )
    assert_check(
        "mapped_genre_keys сохраняет вход",
        result["mapped_genre_keys"] == ["drama", "crime", "comedy"],
    )
    assert_check("unmapped_genre_keys пуст", result["unmapped_genre_keys"] == [])


def test_candidate_genre_keys_to_dataset_genres_maps_mystery_to_detective() -> None:
    """Maps pool mystery key to has_detective, not has_mystery."""
    print("\n20a.2) Проверяем mystery -> has_detective")

    result = candidate_to_dataset.candidate_genre_keys_to_dataset_genres(["mystery"])
    dataset_genre = result["dataset_genre"]

    assert_check("status ok для mystery", result["status"] == "ok")
    assert_check("mystery -> has_detective", dataset_genre["has_detective"] == 1)
    assert_check(
        "has_mystery не создаётся",
        "has_mystery" not in dataset_genre,
    )
    assert_check("mapped_genre_keys == ['mystery']", result["mapped_genre_keys"] == ["mystery"])


def test_candidate_genre_keys_to_dataset_genres_reports_unmapped() -> None:
    """Reports unmapped pool genre keys without failing."""
    print("\n20a.3) Проверяем unmapped genre keys")

    result = candidate_to_dataset.candidate_genre_keys_to_dataset_genres(["drama", "history"])
    dataset_genre = result["dataset_genre"]

    assert_check("status partial", result["status"] == "partial")
    assert_check("drama mapped", dataset_genre["has_drama"] == 1)
    assert_check("history unmapped", result["unmapped_genre_keys"] == ["history"])
    assert_check("mapped_genre_keys == ['drama']", result["mapped_genre_keys"] == ["drama"])


def test_candidate_genre_keys_to_dataset_genres_missing() -> None:
    """Returns missing status for empty or fully unmapped input."""
    print("\n20a.4) Проверяем missing status")

    for genre_keys in ([], None, ["history"], ["totally_unknown_key"]):
        result = candidate_to_dataset.candidate_genre_keys_to_dataset_genres(genre_keys)
        assert_check(
            f"missing status for {genre_keys!r}",
            result["status"] == "missing",
        )
        assert_check(
            f"all genre values are 0 for {genre_keys!r}",
            all(value == 0 for value in result["dataset_genre"].values()),
        )
        assert_check(
            f"полный набор constant.GENRE for {genre_keys!r}",
            set(result["dataset_genre"].keys()) == set(constant.GENRE),
        )


def test_raw_genres_to_dataset_genres_maps_mystery_to_detective() -> None:
    """Raw EN/RU mystery labels map to has_detective, not has_mystery."""
    print("\n20a.4b) Проверяем raw genres -> has_detective")

    for raw_genres in (["Mystery"], ["детектив"], ["Mystery", "детектив"]):
        result = candidate_to_dataset.raw_genres_to_dataset_genres(raw_genres)
        assert_check(
            f"has_detective for {raw_genres!r}",
            result["dataset_genre"]["has_detective"] == 1,
        )
        assert_check(
            f"has_mystery не создаётся for {raw_genres!r}",
            "has_mystery" not in result["dataset_genre"],
        )
        assert_check(
            f"genre_keys содержит mystery for {raw_genres!r}",
            "mystery" in result["genre_keys"],
        )


def test_raw_genres_to_dataset_genres_maps_en_ru_supported() -> None:
    """Raw EN/RU supported genres map to current dataset has_* features."""
    print("\n20a.4c) Проверяем raw EN/RU genres -> has_*")

    result = candidate_to_dataset.raw_genres_to_dataset_genres(
        ["Drama", "Crime", "драма", "криминал"]
    )

    assert_check("status ok", result["status"] == "ok")
    assert_check("has_drama = 1", result["dataset_genre"]["has_drama"] == 1)
    assert_check("has_crime = 1", result["dataset_genre"]["has_crime"] == 1)
    assert_check(
        "полный набор constant.GENRE",
        set(result["dataset_genre"].keys()) == set(constant.GENRE),
    )
    assert_check(
        "дубли raw не дублируют genre_keys",
        result["genre_keys"] == ["drama", "crime"],
    )


def test_raw_genres_to_dataset_genres_reports_unmapped_raw() -> None:
    """Unknown raw genres are reported without raising."""
    print("\n20a.4d) Проверяем unmapped raw genres")

    result = candidate_to_dataset.raw_genres_to_dataset_genres(
        ["Drama", "TotallyUnknownGenreXYZ", "history"]
    )

    assert_check("status partial", result["status"] == "partial")
    assert_check("has_drama = 1", result["dataset_genre"]["has_drama"] == 1)
    assert_check(
        "unknown raw in unmapped_raw_genres",
        result["unmapped_raw_genres"] == ["TotallyUnknownGenreXYZ"],
    )
    assert_check(
        "history key unmapped for current model",
        result["unmapped_genre_keys"] == ["history"],
    )


def test_raw_genres_to_dataset_genres_missing() -> None:
    """Empty or fully unresolvable raw genres return missing status."""
    print("\n20a.4e) Проверяем missing status для raw genres")

    for raw_genres in ([], None, ["TotallyUnknownGenreXYZ"]):
        result = candidate_to_dataset.raw_genres_to_dataset_genres(raw_genres)
        assert_check(
            f"missing status for {raw_genres!r}",
            result["status"] == "missing",
        )
        assert_check(
            f"all genre values are 0 for {raw_genres!r}",
            all(value == 0 for value in result["dataset_genre"].values()),
        )


def test_candidate_transfer_payload_uses_genre_keys_mapper() -> None:
    """Candidate transfer defaults use pool genre_keys mapper for has_* features."""
    print("\n20a.5) Проверяем build_candidate_transfer_payload + genre_keys mapper")

    candidate = {
        "title": "Mapped Transfer",
        "year": 2021,
        "kp_score": 7.5,
        "kp_votes": 100,
        "imdb_score": 7.1,
        "imdb_votes": 1000,
        "genres": ["Mystery"],
        "genre_keys": ["mystery", "drama"],
    }
    payload = title_resolve.build_candidate_transfer_payload(candidate)
    genre_defaults = payload["defaults"][scheme.GENRE]

    assert_check("mystery -> has_detective", genre_defaults["has_detective"] == 1)
    assert_check("drama -> has_drama", genre_defaults["has_drama"] == 1)
    assert_check("has_mystery не создаётся", "has_mystery" not in genre_defaults)
    assert_check(
        "полный набор constant.GENRE",
        set(genre_defaults.keys()) == set(constant.GENRE),
    )


def test_candidate_transfer_payload_falls_back_to_raw_genres_without_genre_keys() -> None:
    """Candidate transfer keeps raw genres fallback when genre_keys are absent."""
    print("\n20a.6) Проверяем fallback на raw genres без genre_keys")

    candidate = {
        "title": "Legacy Transfer",
        "year": 2021,
        "kp_score": 7.5,
        "kp_votes": 100,
        "imdb_score": 7.1,
        "imdb_votes": 1000,
        "genres": ["драма", "криминал"],
    }
    payload = title_resolve.build_candidate_transfer_payload(candidate)
    genre_defaults = payload["defaults"][scheme.GENRE]

    assert_check("драма -> has_drama", genre_defaults["has_drama"] == 1)
    assert_check("криминал -> has_crime", genre_defaults["has_crime"] == 1)
    assert_check("has_mystery не создаётся", "has_mystery" not in genre_defaults)


def test_candidate_transfer_payload_does_not_mutate_candidate() -> None:
    """Candidate transfer payload builder does not mutate the input candidate."""
    print("\n20a.7) Проверяем, что transfer payload не мутирует candidate")

    candidate = {
        "title": "Immutable Transfer",
        "year": 2021,
        "kp_score": 7.5,
        "kp_votes": 100,
        "imdb_score": 7.1,
        "imdb_votes": 1000,
        "genres": ["Mystery"],
        "genre_keys": ["mystery", "drama"],
    }
    before = copy.deepcopy(candidate)
    title_resolve.build_candidate_transfer_payload(candidate)

    assert_check("candidate dict не изменился", candidate == before)


def test_candidate_genre_transfer_preview_maps_genre_keys() -> None:
    """Preview maps pool genre_keys to active has_* features."""
    print("\n20a.8) Проверяем genre transfer preview для mapped genre_keys")

    preview = title_resolve.build_candidate_genre_transfer_preview({
        "genre_keys": ["mystery", "drama"],
        "genres": ["Mystery"],
    })

    assert_check("mapper_status ok", preview["mapper_status"] == "ok")
    assert_check("used_fallback False", preview["used_fallback"] is False)
    assert_check("has_detective active", "has_detective" in preview["active_has_features"])
    assert_check("has_drama active", "has_drama" in preview["active_has_features"])
    assert_check("has_mystery не создаётся", "has_mystery" not in preview["dataset_genre"])
    assert_check(
        "preview совпадает с transfer defaults",
        preview["dataset_genre"]
        == title_resolve.build_candidate_transfer_genre_defaults({
            "genre_keys": ["mystery", "drama"],
            "genres": ["Mystery"],
        }),
    )


def test_candidate_genre_transfer_preview_reports_partial_unmapped() -> None:
    """Preview reports unmapped pool genre keys for partial mapper status."""
    print("\n20a.9) Проверяем partial/unmapped в genre transfer preview")

    preview = title_resolve.build_candidate_genre_transfer_preview({
        "genre_keys": ["drama", "history"],
    })

    assert_check("mapper_status partial", preview["mapper_status"] == "partial")
    assert_check("drama mapped", preview["mapped_genre_keys"] == ["drama"])
    assert_check("history unmapped", preview["unmapped_genre_keys"] == ["history"])
    assert_check("has_drama active", preview["dataset_genre"]["has_drama"] == 1)


def test_candidate_genre_transfer_preview_falls_back_to_raw_genres() -> None:
    """Preview uses raw genres fallback when genre_keys are absent."""
    print("\n20a.10) Проверяем fallback preview без genre_keys")

    preview = title_resolve.build_candidate_genre_transfer_preview({
        "genres": ["драма", "криминал"],
    })

    assert_check("mapper_status missing", preview["mapper_status"] == "missing")
    assert_check("used_fallback True", preview["used_fallback"] is True)
    assert_check("has_drama active", preview["dataset_genre"]["has_drama"] == 1)
    assert_check("has_crime active", preview["dataset_genre"]["has_crime"] == 1)
    assert_check("raw_genres сохранены", preview["raw_genres"] == ["драма", "криминал"])


def test_candidate_transfer_payload_falls_back_to_imdb_genres() -> None:
    """Fallback transfer defaults use imdb_genres when common genres field is empty."""
    print("\n20a.10b) Проверяем fallback transfer defaults через imdb_genres")

    candidate = {
        "title": "IMDb Fallback Transfer",
        "year": 2021,
        "genres": [],
        "imdb_genres": ["Crime", "Drama"],
    }
    payload = title_resolve.build_candidate_transfer_payload(candidate)
    genre_defaults = payload["defaults"][scheme.GENRE]

    assert_check("Crime -> has_crime", genre_defaults["has_crime"] == 1)
    assert_check("Drama -> has_drama", genre_defaults["has_drama"] == 1)


def test_candidate_genre_transfer_preview_falls_back_to_imdb_genres() -> None:
    """Preview fallback uses imdb_genres when genre_keys and genres are empty."""
    print("\n20a.10c) Проверяем preview fallback через imdb_genres")

    preview = title_resolve.build_candidate_genre_transfer_preview({
        "genres": [],
        "imdb_genres": ["Crime"],
        "genres_tmdb": ["Drama"],
    })

    assert_check("used_fallback True", preview["used_fallback"] is True)
    assert_check("has_crime active", preview["dataset_genre"]["has_crime"] == 1)
    assert_check("has_drama active", preview["dataset_genre"]["has_drama"] == 1)
    assert_check(
        "raw_genres включает imdb и tmdb",
        preview["raw_genres"] == ["Crime", "Drama"],
    )


def test_candidate_genre_transfer_preview_warns_when_all_zero_with_raw_signals() -> None:
    """Preview warns when raw genres exist but no has_* defaults were resolved."""
    print("\n20a.11) Проверяем warn_all_genres_zero в genre transfer preview")

    preview = title_resolve.build_candidate_genre_transfer_preview({
        "genres": ["TotallyUnknownGenreXYZ"],
    })

    assert_check("has_raw_genre_signals True", preview["has_raw_genre_signals"] is True)
    assert_check("warn_all_genres_zero True", preview["warn_all_genres_zero"] is True)
    assert_check("active_has_features пуст", preview["active_has_features"] == [])


def test_candidate_genre_transfer_preview_does_not_mutate_candidate() -> None:
    """Genre transfer preview is read-only for the input candidate."""
    print("\n20a.12) Проверяем, что genre transfer preview не мутирует candidate")

    candidate = {
        "title": "Preview Immutable",
        "genres": ["Mystery"],
        "genre_keys": ["mystery", "drama"],
        "imdb_genres": ["Crime"],
    }
    before = copy.deepcopy(candidate)
    title_resolve.build_candidate_genre_transfer_preview(candidate)

    assert_check("candidate dict не изменился", candidate == before)


def test_mark_candidate_as_watched_prints_genre_transfer_preview() -> None:
    """mark_candidate_as_watched shows genre preview before the add form."""
    print("\n20a.13) Проверяем genre preview в mark_candidate_as_watched")

    mark_watched_source = inspect.getsource(interface_funcs.mark_candidate_as_watched)
    helper_source = inspect.getsource(interface_funcs.print_candidate_genre_transfer_preview)

    assert_check(
        "mark_candidate_as_watched вызывает build_candidate_genre_transfer_preview",
        "build_candidate_genre_transfer_preview" in mark_watched_source,
    )
    assert_check(
        "mark_candidate_as_watched вызывает print_candidate_genre_transfer_preview",
        "print_candidate_genre_transfer_preview" in mark_watched_source,
    )
    assert_check(
        "preview печатается до request_all_scores",
        mark_watched_source.index("print_candidate_genre_transfer_preview")
        < mark_watched_source.index("request_all_scores"),
    )
    assert_check(
        "helper показывает pool genre_keys",
        "Pool genre_keys" in helper_source,
    )
    assert_check(
        "helper показывает fallback",
        "fallback по raw genres" in helper_source,
    )
    assert_check(
        "helper показывает partial unmapped",
        "Не удалось сопоставить" in helper_source,
    )
    assert_check(
        "helper показывает all-zero warning",
        "warn_all_genres_zero" in helper_source or "raw-жанры" in helper_source,
    )


def test_candidate_country_schema_normalization() -> None:
    """Проверяет alias -> country_codes -> country_display без мутации raw countries."""
    print("\n20b) Проверяем candidate country_schema normalization")

    kr_case = candidate_schema.normalize_candidate_record({
        "title": "KR Case",
        "year": 2020,
        "countries": ["KR", "South Korea"],
    })
    assert_check("KR+South Korea -> country_codes [KR]", kr_case["country_codes"] == ["KR"])
    assert_check("KR+South Korea -> country_display Южная Корея", kr_case["country_display"] == "Южная Корея")

    kr_alias_case = candidate_schema.normalize_candidate_record({
        "title": "KR Alias Case",
        "year": 2020,
        "countries": ["Корея Южная"],
    })
    assert_check("Корея Южная -> country_codes [KR]", kr_alias_case["country_codes"] == ["KR"])
    assert_check("Корея Южная -> country_display Южная Корея", kr_alias_case["country_display"] == "Южная Корея")

    jp_case = candidate_schema.normalize_candidate_record({
        "title": "JP Case",
        "year": 2020,
        "countries": ["JP", "Япония"],
    })
    assert_check("JP+Япония -> country_codes [JP]", jp_case["country_codes"] == ["JP"])
    assert_check("JP+Япония -> country_display Япония", jp_case["country_display"] == "Япония")

    ru_case = candidate_schema.normalize_candidate_record({
        "title": "RU Case",
        "year": 2020,
        "countries": ["RU", "Россия"],
    })
    assert_check("RU+Россия -> country_codes [RU]", ru_case["country_codes"] == ["RU"])
    assert_check("RU+Россия -> country_display Россия", ru_case["country_display"] == "Россия")

    priority_case = candidate_schema.normalize_candidate_record({
        "title": "Country Priority Case",
        "year": 2020,
        "tmdb_origin_countries": ["KR"],
        "countries": ["South Korea"],
    })
    assert_check(
        "tmdb_origin_countries имеет приоритет над countries",
        priority_case["country_codes"] == ["KR"],
    )
    assert_check(
        "priority case -> country_display Южная Корея",
        priority_case["country_display"] == "Южная Корея",
    )

    unknown_case = candidate_schema.normalize_candidate_record({
        "title": "Unknown Country Case",
        "year": 2020,
        "countries": ["Atlantis"],
    })
    assert_check("Unknown country -> country_codes пустой", unknown_case["country_codes"] == [])
    assert_check("Unknown country -> country_display None", unknown_case["country_display"] is None)
    assert_check(
        "Unknown country UI fallback на raw countries",
        country_schema.candidate_country_for_display(unknown_case) == "Atlantis",
    )

    legacy_raw = {
        "title": "Legacy Country Raw",
        "year": 2020,
        "countries": ["KR", "South Korea"],
    }
    legacy_snapshot = copy.deepcopy(legacy_raw)
    normalized_legacy = candidate_schema.normalize_candidate_record(legacy_raw)
    assert_check("legacy candidate countries не мутируется", legacy_raw == legacy_snapshot)
    assert_check(
        "legacy countries сохраняются в normalized copy",
        normalized_legacy["countries"] == ["KR", "South Korea"],
    )

    coproduction_case = candidate_schema.normalize_candidate_record({
        "title": "Coproduction Case",
        "year": 2020,
        "countries": ["US", "GB"],
    })
    assert_check("Coproduction -> оба ISO-2 code", coproduction_case["country_codes"] == ["US", "GB"])
    assert_check("Coproduction display -> primary US label", coproduction_case["country_display"] == "США")

    candidate_pool.save_candidate_pool({
        "legacy-one": {
            "title": "Country Save Case",
            "alternative_title": "",
            "year": 2021,
            "criteria_name": "tmdb_KR_quality",
            "countries": ["KR", "South Korea"],
            "kp_score": 7.5,
            "kp_votes": 1000,
            "imdb_score": 7.4,
            "imdb_votes": 5000,
            "genres": ["драма"],
        },
    })
    saved_pool = candidate_pool.load_candidate_pool()
    saved_candidate = next(iter(saved_pool.values()))
    assert_check("Write path сохраняет country_codes", saved_candidate.get("country_codes") == ["KR"])
    assert_check("Write path сохраняет country_display", saved_candidate.get("country_display") == "Южная Корея")


def test_remove_candidate_from_pool() -> None:
    """Проверяет удаление просмотренного кандидата из общего пула по совпадающему названию и году."""
    print("\n12) Проверяем удаление просмотренного кандидата из пула")

    candidate_pool.save_candidate_pool({
        "one": {
            "title": "Ева, рожай!",
            "alternative_title": "",
            "year": 2022,
            "criteria_name": "A",
            "kp_score": 7.0,
            "kp_votes": 1000,
            "imdb_score": 0,
            "imdb_votes": 0,
            "genres": ["комедия"],
        },
        "two": {
            "title": "Ева рожай",
            "alternative_title": "",
            "year": 2022,
            "criteria_name": "B",
            "kp_score": 8.0,
            "kp_votes": 2000,
            "imdb_score": 0,
            "imdb_votes": 0,
            "genres": ["комедия"],
        },
        "three": {
            "title": "Другой сериал",
            "alternative_title": "",
            "year": 2022,
            "criteria_name": "A",
            "kp_score": 8.0,
            "kp_votes": 500,
            "imdb_score": 0,
            "imdb_votes": 0,
            "genres": ["драма"],
        },
    })

    removed = candidate_pool.remove_candidate_from_pool({
        "title": "Ева рожай",
        "alternative_title": "",
        "year": 2022,
        "criteria_name": "X",
    })
    pool = candidate_pool.load_candidate_pool()

    assert_check("Удалён совпавший кандидат из единого общего пула", removed == 2)
    assert_check("В пуле остался только другой сериал", len(pool) == 1)


def test_candidate_pool_genre_filters() -> None:
    """Проверяет фильтрацию кандидатов по жанрам и жанрам-исключениям."""
    print("\n13) Проверяем фильтрацию пула кандидатов по жанрам")

    movie = {
        "genres": [
            {"name": "драма"},
            {"name": "триллер"},
        ]
    }

    assert_check(
        "Кандидат проходит по обязательному жанру",
        candidate_pool.movie_matches_genres(movie, ["драма"], [])
    )
    assert_check(
        "Кандидат отсекается по исключенному жанру",
        candidate_pool.movie_matches_genres(movie, ["драма"], ["триллер"]) is False
    )
    assert_check(
        "При пустом обязательном списке жанров кандидат проходит",
        candidate_pool.movie_matches_genres(movie, [], []) is True
    )


def test_tmdb_candidate_pool_criteria_name() -> None:
    """Проверяет ручной/auto criteria_name и совместимость импорта старых result."""
    print("\n14) Проверяем criteria_name для TMDb candidate_pool")

    def fake_details(_tmdb_id, **_kwargs):
        return {
            "id": 101,
            "name": "TMDb Candidate",
            "original_name": "TMDb Candidate",
            "first_air_date": "2021-01-01",
            "origin_country": ["RU"],
            "production_countries": [{"iso_3166_1": "RU", "name": "Russia"}],
            "genres": [{"name": "Drama"}],
            "vote_average": 8.0,
            "vote_count": 100,
            "popularity": 10,
            "external_ids": {},
            "credits": {},
        }

    with patch("candidates.tmdb_candidate_pool.api_tmdb.load_tmdb_token", return_value="token"):
        with patch("candidates.tmdb_candidate_pool.api_tmdb.discover_tv_candidates", return_value=[{"id": 101, "vote_count": 100, "popularity": 10}]):
            with patch("candidates.tmdb_candidate_pool.api_tmdb.get_tv_details", side_effect=fake_details):
                with patch("candidates.tmdb_candidate_pool.connect_imdb", return_value=None):
                    with patch(
                        "candidates.tmdb_candidate_pool.enrich_from_kp_api_if_needed",
                        side_effect=lambda candidate, _country, _stats, **_kwargs: candidate,
                    ):
                        manual_result = tmdb_candidate_pool.build_candidate_pool(
                            country="RU",
                            pages=1,
                            details_limit=1,
                            mode="quality",
                            criteria_name="Русские драмы 2020+",
                        )
                        auto_result = tmdb_candidate_pool.build_candidate_pool(
                            country="RU",
                            pages=1,
                            details_limit=1,
                            mode="quality",
                            year_min=2020,
                            min_tmdb_score=7.0,
                        )

    assert_check("Ручное criteria_name сохраняется в result", manual_result["criteria_name"] == "Русские драмы 2020+")
    assert_check("Ручное criteria_name сохраняется в candidate", manual_result["candidates"][0]["criteria_name"] == "Русские драмы 2020+")
    assert_check("Auto criteria_name создаётся по параметрам", auto_result["criteria_name"] == "tmdb_RU_quality_2020plus_min7")
    assert_check("Auto criteria_name сохраняется в settings", auto_result["settings"]["criteria_name"] == "tmdb_RU_quality_2020plus_min7")

    old_result_path = Path(constant.DATA_DIR) / "old_tmdb_result.json"
    old_result = {
        "country": "RU",
        "mode": "quality",
        "candidates": [
            {
                "title": "Old TMDb Candidate",
                "year": 2022,
                "tmdb_id": 200,
                "genres_tmdb": ["Drama"],
                "tmdb_rating": 7.5,
                "tmdb_votes": 50,
            }
        ],
    }
    old_result_path.write_text(json.dumps(old_result, ensure_ascii=False), encoding="utf-8")
    stats = tmdb_candidate_pool.import_tmdb_result_to_common_pool(old_result_path)
    pool = candidate_pool.load_candidate_pool()

    assert_check("Старый result без criteria_name импортируется", stats["ok"] is True)
    assert_check("Для старого result используется fallback criteria_name", stats["criteria_name"] == "tmdb_RU_quality")
    assert_check(
        "Common candidate получает fallback criteria_name",
        any(candidate.get("criteria_name") == "tmdb_RU_quality" for candidate in pool.values())
    )


def test_tmdb_candidate_pool_network_error_skip_gate() -> None:
    """Проверяет, что TMDb Details и KP API пропускаются после 3 подряд сетевых ошибок."""
    print("\n14.0b) Проверяем skip-gate после 3 сетевых ошибок TMDb/KP")

    discover_items = [
        {"id": index, "vote_count": 100, "popularity": 10}
        for index in range(1, 6)
    ]
    progress_events = []

    def fake_details(tmdb_id, **_kwargs):
        return {
            "id": tmdb_id,
            "name": f"Network Gate {tmdb_id}",
            "original_name": f"Network Gate {tmdb_id}",
            "first_air_date": "2024-01-01",
            "origin_country": ["RU"],
            "production_countries": [{"iso_3166_1": "RU", "name": "Russia"}],
            "genres": [{"name": "Drama"}],
            "vote_average": 7.5,
            "vote_count": 100,
            "popularity": 10,
            "external_ids": {},
            "credits": {},
        }

    tmdb_details_calls = []

    def failing_details(tmdb_id, **_kwargs):
        tmdb_details_calls.append(tmdb_id)
        raise RuntimeError("tmdb offline")

    try:
        tmdb_candidate_pool.set_progress_reporter(lambda source, status: progress_events.append((source, status)))
        with patch("candidates.tmdb_candidate_pool.api_tmdb.load_tmdb_token", return_value="token"):
            with patch("candidates.tmdb_candidate_pool.api_tmdb.discover_tv_candidates", return_value=discover_items):
                with patch("candidates.tmdb_candidate_pool.api_tmdb.get_tv_details", side_effect=failing_details):
                    with patch("candidates.tmdb_candidate_pool.connect_imdb", return_value=None):
                        tmdb_result = tmdb_candidate_pool.build_candidate_pool(
                            country="RU",
                            pages=1,
                            details_limit=5,
                            mode="quality",
                            kp_api_limit=0,
                        )
    finally:
        tmdb_candidate_pool.set_progress_reporter(None)

    assert_check("TMDb Details останавливает сетевые попытки после 3 ошибок", len(tmdb_details_calls) == 3)
    assert_check("TMDb Details stats считает ошибки", tmdb_result["stats"]["tmdb_details_errors"] == 3)
    assert_check("TMDb Details stats считает пропуски", tmdb_result["stats"]["tmdb_details_skipped_after_errors"] == 2)
    assert_check(
        "TMDb Details выводит Пропущено",
        any(source == "TMDb Details" and "Пропущено" in status for source, status in progress_events),
    )

    progress_events = []
    kp_lookup_calls = []

    def failing_kp_lookup(*_args, **_kwargs):
        kp_lookup_calls.append(1)
        return {
            "status": "error",
            "movie": None,
            "error": "network_error",
            "reject_reason": None,
            "query": "offline",
            "attempts": 1,
        }

    try:
        tmdb_candidate_pool.set_progress_reporter(lambda source, status: progress_events.append((source, status)))
        with patch("candidates.tmdb_candidate_pool.api_tmdb.load_tmdb_token", return_value="token"):
            with patch("candidates.tmdb_candidate_pool.api_tmdb.discover_tv_candidates", return_value=discover_items):
                with patch("candidates.tmdb_candidate_pool.api_tmdb.get_tv_details", side_effect=fake_details):
                    with patch("candidates.tmdb_candidate_pool.connect_imdb", return_value=None):
                        with patch("candidates.kp_enrichment.lookup_kp_via_api", side_effect=failing_kp_lookup):
                            kp_result = tmdb_candidate_pool.build_candidate_pool(
                                country="RU",
                                pages=1,
                                details_limit=5,
                                mode="quality",
                            )
    finally:
        tmdb_candidate_pool.set_progress_reporter(None)

    assert_check("KP API останавливает сетевые попытки после 3 ошибок", len(kp_lookup_calls) == 3)
    assert_check("KP API stats считает ошибки", kp_result["stats"]["kp_api_errors"] == 3)
    assert_check("KP API stats считает пропуски", kp_result["stats"]["kp_api_skipped_after_errors"] == 2)
    assert_check(
        "KP API выводит Пропущено",
        any(source == "KP API" and status == "Пропущено" for source, status in progress_events),
    )
    assert_check(
        "KP API сохраняет статус skipped_network_errors у пропущенных кандидатов",
        any(candidate.get("kp_status") == "skipped_network_errors" for candidate in kp_result["candidates"]),
    )


def test_tmdb_candidate_pool_discover_genre_filters() -> None:
    """Checks optional TMDb Discover genre filters without real network calls."""
    print("\n14.1) Проверяем TMDb Discover genre filters")

    captured_params = []

    def fake_cached_tmdb_get(_endpoint, params, _cache_path, **_kwargs):
        captured_params.append(dict(params))
        return {"results": [], "total_pages": 1}

    with patch("apis.tmdb_api.cached_tmdb_get", side_effect=fake_cached_tmdb_get):
        tmdb_api.discover_tv_candidates(
            country="RU",
            vote_average_gte=6.3,
            vote_count_gte=10,
            max_pages=1,
            with_genres="18|9648|80",
            without_genres="10766|10764",
            token="token",
        )

    assert_check(
        "TMDb API params сохраняют with_genres как строку",
        captured_params[0]["with_genres"] == "18|9648|80",
    )
    assert_check(
        "TMDb API params сохраняют without_genres как строку",
        captured_params[0]["without_genres"] == "10766|10764",
    )
    assert_check(
        "TMDb genre helper label содержит TMDb",
        "TMDb" in tmdb_genre_options.TMDB_DISCOVER_GENRE_TITLE
        and "TMDb OR" in tmdb_genre_options.TMDB_INCLUDE_OR_LABEL
        and "TMDb AND" in tmdb_genre_options.TMDB_INCLUDE_AND_LABEL,
    )

    def run_tmdb_genre_helper(answers: list[str]):
        helper_output = []
        helper_prompts = []
        helper_answers = iter(answers)

        def input_func(prompt):
            helper_prompts.append(prompt)
            return next(helper_answers)

        with_genres, without_genres = interface_funcs.request_tmdb_discover_genre_filters(
            input_func=input_func,
            output_func=helper_output.append,
        )
        return with_genres, without_genres, "\n".join(helper_output), "\n".join(helper_prompts)

    enter_with_genres, enter_without_genres, enter_output, _enter_prompts = run_tmdb_genre_helper(["", ""])
    single_with_genres, single_without_genres, single_output, _single_prompts = run_tmdb_genre_helper(["1", ""])
    or_with_genres, or_without_genres, helper_output_text, _or_prompts = run_tmdb_genre_helper(["1,2,3", "1", ""])
    and_with_genres, _and_without_genres, and_output, _and_prompts = run_tmdb_genre_helper(["1,2,3", "2", ""])
    _exclude_with_genres, exclude_without_genres, exclude_output, _exclude_prompts = run_tmdb_genre_helper(["", "1,2,3,4"])
    _all_exclude_with_genres, all_exclude_without_genres, all_exclude_output, _all_exclude_prompts = run_tmdb_genre_helper(["", "все"])

    assert_check(
        "TMDb genre helper Enter include/exclude возвращает no filter",
        enter_with_genres is None and enter_without_genres is None,
    )
    assert_check(
        "TMDb genre helper не спрашивает OR/AND при пустом include",
        "Как применять выбранные жанры (TMDb)?" not in enter_output,
    )
    assert_check(
        "TMDb genre helper один include-жанр не спрашивает OR/AND",
        single_with_genres == "18"
        and single_without_genres is None
        and "Как применять выбранные жанры (TMDb)?" not in single_output,
    )
    assert_check(
        "TMDb genre helper собирает OR строку без сетевых запросов",
        or_with_genres == "18|9648|80" and or_without_genres is None,
    )
    assert_check(
        "TMDb genre helper собирает AND строку без сетевых запросов",
        and_with_genres == "18,9648,80" and "Режим: все выбранные одновременно" in and_output,
    )
    assert_check(
        "TMDb genre helper собирает exclude строку без сетевых запросов",
        exclude_without_genres == "10766|10764|10767|10763"
        and "Как применять выбранные жанры (TMDb)?" not in exclude_output,
    )
    assert_check(
        "TMDb genre helper поддерживает пункт все для exclude",
        all_exclude_without_genres == "10766|10764|10767|10763|10762|99"
        and "все >>" in all_exclude_output,
    )
    assert_check(
        "TMDb genre helper показывает русские labels без ID",
        "Драма" in helper_output_text
        and "Детектив / мистика" in helper_output_text
        and "18" not in helper_output_text
        and "9648" not in helper_output_text
        and "10766" not in exclude_output,
    )
    assert_check(
        "TMDb genre helper labels содержат (TMDb)",
        "Жанры для поиска (TMDb)" in helper_output_text
        and "Include жанры (TMDb)" in helper_output_text
        and "Exclude жанры (TMDb)" in exclude_output,
    )
    tmdb_labels = [option["label"] for option in tmdb_genre_options.TV_GENRE_OPTIONS]
    assert_check(
        "TMDb genre picker содержит только TMDb TV genres и не содержит Триллер",
        "Анимация" in tmdb_labels and "Триллер" not in tmdb_labels,
    )
    assert_check(
        "Runtime prediction genres могут нормализовать Триллер",
        pool_genres.normalize_genre_list(["Триллер", "thriller"]) == ["thriller"],
    )

    discover_calls = []

    def fake_discover(**kwargs):
        discover_calls.append(dict(kwargs))
        return []

    with patch("candidates.tmdb_candidate_pool.api_tmdb.load_tmdb_token", return_value="token"):
        with patch("candidates.tmdb_candidate_pool.api_tmdb.discover_tv_candidates", side_effect=fake_discover):
            tmdb_candidate_pool.build_candidate_pool(
                country="RU",
                pages=1,
                details_limit=0,
                mode="quality",
                with_genres="18|9648|80",
            )
            tmdb_candidate_pool.build_candidate_pool(
                country="RU",
                pages=1,
                details_limit=0,
                mode="quality",
                without_genres="10766|10764",
            )
            empty_result = tmdb_candidate_pool.build_candidate_pool(
                country="RU",
                pages=1,
                details_limit=0,
                mode="quality",
                with_genres="",
                without_genres="   ",
            )
            none_result = tmdb_candidate_pool.build_candidate_pool(
                country="RU",
                pages=1,
                details_limit=0,
                mode="quality",
                with_genres=None,
                without_genres=None,
            )

    assert_check(
        "build_candidate_pool прокидывает with_genres в Discover",
        discover_calls[0].get("with_genres") == "18|9648|80",
    )
    assert_check(
        "build_candidate_pool прокидывает without_genres в Discover",
        discover_calls[1].get("without_genres") == "10766|10764",
    )
    assert_check(
        "Пустой with_genres не передаётся в Discover",
        "with_genres" not in discover_calls[2],
    )
    assert_check(
        "Пустой without_genres не передаётся в Discover",
        "without_genres" not in discover_calls[2],
    )
    assert_check(
        "None with_genres не передаётся в Discover",
        "with_genres" not in discover_calls[3],
    )
    assert_check(
        "None without_genres не передаётся в Discover",
        "without_genres" not in discover_calls[3],
    )
    assert_check(
        "Result settings сохраняют with_genres=None для пустого ввода",
        empty_result["settings"]["with_genres"] is None,
    )
    assert_check(
        "Result settings сохраняют without_genres=None для None ввода",
        none_result["settings"]["without_genres"] is None,
    )

    with patch("candidates.tmdb_candidate_pool.build_candidate_pool", return_value={"ok": True}) as build_mock:
        candidate_service.build_tmdb_candidate_pool(
            country="RU",
            pages=1,
            details_limit=5,
            mode="quality",
            with_genres="18,80",
            without_genres="10766|10764",
        )
    assert_check(
        "Service facade прокидывает with_genres",
        build_mock.call_args.kwargs["with_genres"] == "18,80",
    )
    assert_check(
        "Service facade прокидывает without_genres",
        build_mock.call_args.kwargs["without_genres"] == "10766|10764",
    )

    import inspect

    build_flow_source = inspect.getsource(interface_funcs.run_tmdb_candidate_pool_flow)
    assert_check(
        "UI flow спрашивает жанры для TMDb Discover",
        "request_tmdb_discover_genre_filters" in build_flow_source
        and "Include жанры (TMDb)" in build_flow_source,
    )
    assert_check(
        "UI flow подписывает TMDb exclude genres",
        "Exclude жанры (TMDb)" in build_flow_source,
    )
    assert_check(
        "UI flow передаёт genre filters через candidate_service",
        "candidate_service.build_tmdb_candidate_pool" in build_flow_source
        and "with_genres=with_genres" in build_flow_source
        and "without_genres=without_genres" in build_flow_source,
    )
    prediction_filter_source = inspect.getsource(interface_funcs._request_prediction_candidate_filters)
    default_lines = candidate_pool.format_prediction_filter_default_lines({
        "country": None,
        "year_min": None,
        "year_max": None,
        "include_genres": ["драма"],
        "exclude_genres": ["комедия"],
    })
    assert_check(
        "Prediction filter UI label указывает saved pool",
        "saved pool" in prediction_filter_source
        and "по сохранённым данным pool" in prediction_filter_source,
    )
    assert_check(
        "Prediction defaults label указывает saved pool",
        any("saved pool" in line for line in default_lines),
    )


def test_tmdb_import_keeps_cross_criteria_entries() -> None:
    """Проверяет, что TMDb import не теряет одинаковый title/year из разных criteria."""
    print("\n16) Проверяем cross-criteria TMDb import")

    first_result_path = Path(constant.DATA_DIR) / "tmdb_result_quality.json"
    second_result_path = Path(constant.DATA_DIR) / "tmdb_result_hidden.json"

    first_result = {
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "candidates": [
            {
                "title": "Метод",
                "year": 2015,
                "tmdb_id": 301,
                "genres_tmdb": ["Drama"],
                "tmdb_rating": 7.5,
                "tmdb_votes": 50,
            }
        ],
    }
    second_result = {
        "criteria_name": "tmdb_RU_hidden",
        "country": "RU",
        "mode": "hidden",
        "candidates": [
            {
                "title": "Метод",
                "year": 2015,
                "tmdb_id": 302,
                "genres_tmdb": ["Drama"],
                "tmdb_rating": 7.7,
                "tmdb_votes": 40,
            }
        ],
    }

    first_result_path.write_text(json.dumps(first_result, ensure_ascii=False), encoding="utf-8")
    second_result_path.write_text(json.dumps(second_result, ensure_ascii=False), encoding="utf-8")

    first_stats = tmdb_candidate_pool.import_tmdb_result_to_common_pool(first_result_path)
    second_stats = tmdb_candidate_pool.import_tmdb_result_to_common_pool(second_result_path)
    pool = candidate_pool.load_candidate_pool()
    method_candidates = [
        candidate
        for candidate in pool.values()
        if candidate.get("title") == "Метод" and candidate.get("year") == 2015
    ]
    criteria_names = {candidate.get("criteria_name") for candidate in method_candidates}

    assert_check("Первый TMDb import успешен", first_stats["ok"] is True)
    assert_check("Второй TMDb import успешен", second_stats["ok"] is True)
    assert_check("После двух import остаются обе criteria-версии", criteria_names == {"tmdb_RU_quality", "tmdb_RU_hidden"})
    assert_check("TMDb import не схлопывает одинаковый title/year из разных criteria", len(method_candidates) == 2)

def test_import_tmdb_module_normalizes_schema_and_keeps_unknown_fields() -> None:
    """Проверяет новый import_tmdb модуль: criteria, schema и unknown fields."""
    print("\n17) Проверяем import_tmdb module schema normalization")

    result_path = Path(constant.DATA_DIR) / "tmdb_module_result.json"
    result = {
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "candidates": [
            {
                "title": "Schema Candidate",
                "year": 2022,
                "tmdb_id": 401,
                "tmdb_rating": 7.9,
                "tmdb_votes": 77,
                "imdb_rating": 7.2,
                "imdb_votes": 8800,
                "custom_payload": {"hello": "world"},
            }
        ],
    }
    result_path.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

    stats = tmdb_import.import_tmdb_result_to_common_pool(result_path)
    pool = candidate_pool.load_candidate_pool()
    candidate = next(
        item for item in pool.values()
        if item.get("title") == "Schema Candidate"
    )

    assert_check("Новый import_tmdb import успешен", stats["ok"] is True)
    assert_check("criteria_name перенесён в common pool", candidate["criteria_name"] == "tmdb_RU_quality")
    assert_check("source нормализован", candidate["source"] == "tmdb_imdb_kp_v1")
    assert_check("tmdb_score нормализован", candidate["tmdb_score"] == 7.9)
    assert_check("Unknown field не теряется", candidate["custom_payload"] == {"hello": "world"})


def test_import_tmdb_module_skips_watched_by_title_identity() -> None:
    """Проверяет watched-skip через title/year identity."""
    print("\n18) Проверяем watched skip в import_tmdb module")

    storage_movie.add_movie(make_movie(title="Watched Import", user_score=8.0))
    result_path = Path(constant.DATA_DIR) / "tmdb_watched_result.json"
    result = {
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "candidates": [
            {
                "title": "Watched Import",
                "year": 2024,
                "tmdb_id": 402,
                "tmdb_rating": 7.1,
                "tmdb_votes": 42,
            }
        ],
    }
    result_path.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

    stats = tmdb_import.import_tmdb_result_to_common_pool(result_path)
    pool = candidate_pool.load_candidate_pool()

    assert_check("Watched candidate пропускается", stats["watched_skipped"] == 1)
    assert_check("Совместимый alias skipped_watched тоже заполнен", stats["skipped_watched"] == 1)
    assert_check(
        "Watched candidate не попадает в common pool",
        all(candidate.get("title") != "Watched Import" for candidate in pool.values())
    )


def test_import_tmdb_preserves_existing_criteria_filters() -> None:
    """Проверяет, что TMDb import не затирает существующие filters criteria."""
    print("\n19) Проверяем сохранение existing criteria filters при TMDb import")

    candidate_pool.save_candidate_pool({})
    candidate_pool.save_candidate_criteria({})
    storage_data.clean_dataset()

    candidate_pool.save_named_criteria("tmdb_RU_quality", {
        "criteria_name": "tmdb_RU_quality",
        "genres": ["драма"],
        "excluded_genres": ["мелодрама"],
        "min_kp": 7.0,
        "custom_note": "ручная настройка",
    })

    result_path = Path(constant.DATA_DIR) / "tmdb_preserve_filters.json"
    result = {
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "settings": {"criteria_name": "tmdb_RU_quality", "pages": 1},
        "candidates": [
            {
                "title": "Preserve Filters",
                "year": 2024,
                "tmdb_id": 501,
                "tmdb_rating": 7.8,
                "tmdb_votes": 80,
            }
        ],
    }
    result_path.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

    stats = tmdb_import.import_tmdb_result_to_common_pool(result_path)
    criteria = candidate_pool.load_candidate_criteria()["tmdb_RU_quality"]

    assert_check("TMDb import успешен", stats["ok"] is True)
    assert_check("genres сохраняются после import", criteria["genres"] == ["драма"])
    assert_check("excluded_genres сохраняются после import", criteria["excluded_genres"] == ["мелодрама"])
    assert_check("min_kp сохраняется после import", criteria["min_kp"] == 7.0)
    assert_check("Unknown custom field сохраняется", criteria["custom_note"] == "ручная настройка")
    assert_check("Metadata source обновляется", criteria["source"] == "tmdb_imdb_kp_v1")
    assert_check("Metadata result_file обновляется", criteria["result_file"].endswith("tmdb_preserve_filters.json"))


def test_import_tmdb_creates_new_criteria_with_safe_defaults() -> None:
    """Проверяет, что новый criteria создаётся с безопасными default filters и metadata."""
    print("\n20) Проверяем создание нового criteria при TMDb import")

    candidate_pool.save_candidate_pool({})
    candidate_pool.save_candidate_criteria({})
    storage_data.clean_dataset()

    result_path = Path(constant.DATA_DIR) / "tmdb_new_criteria.json"
    result = {
        "criteria_name": "tmdb_US_quality",
        "country": "US",
        "mode": "quality",
        "settings": {"criteria_name": "tmdb_US_quality", "pages": 1},
        "candidates": [
            {
                "title": "New Criteria Candidate",
                "year": 2023,
                "tmdb_id": 502,
                "tmdb_rating": 7.0,
                "tmdb_votes": 55,
            }
        ],
    }
    result_path.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

    stats = tmdb_import.import_tmdb_result_to_common_pool(result_path)
    criteria = candidate_pool.load_candidate_criteria()["tmdb_US_quality"]

    assert_check("Новый criteria import успешен", stats["ok"] is True)
    assert_check("Новый criteria получает default genres", criteria["genres"] == [])
    assert_check("Новый criteria получает default excluded_genres", criteria["excluded_genres"] == [])
    assert_check("Новый criteria получает default min_kp=None", criteria["min_kp"] is None)
    assert_check("Новый criteria получает metadata country", criteria["country"] == "US")
    assert_check("Новый criteria получает metadata mode", criteria["mode"] == "quality")
    assert_check("Новый criteria получает candidate_count", criteria["candidate_count"] == 1)


def test_import_tmdb_repeated_import_keeps_custom_unknown_fields() -> None:
    """Проверяет, что repeated import не стирает custom unknown fields в criteria."""
    print("\n21) Проверяем repeated import без потери custom fields")

    candidate_pool.save_candidate_pool({})
    candidate_pool.save_candidate_criteria({})
    storage_data.clean_dataset()

    candidate_pool.save_named_criteria("tmdb_RU_quality", {
        "criteria_name": "tmdb_RU_quality",
        "genres": ["триллер"],
        "custom_note": "keep_me",
        "custom_blob": {"owner": "user"},
    })

    first_result_path = Path(constant.DATA_DIR) / "tmdb_repeat_first.json"
    second_result_path = Path(constant.DATA_DIR) / "tmdb_repeat_second.json"
    base_result = {
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "settings": {"criteria_name": "tmdb_RU_quality"},
        "candidates": [
            {
                "title": "Repeat Candidate",
                "year": 2022,
                "tmdb_id": 503,
                "tmdb_rating": 7.4,
                "tmdb_votes": 44,
            }
        ],
    }
    first_result_path.write_text(json.dumps(base_result, ensure_ascii=False), encoding="utf-8")
    second_result_path.write_text(json.dumps(base_result, ensure_ascii=False), encoding="utf-8")

    first_stats = tmdb_import.import_tmdb_result_to_common_pool(first_result_path)
    second_stats = tmdb_import.import_tmdb_result_to_common_pool(second_result_path)
    criteria = candidate_pool.load_candidate_criteria()["tmdb_RU_quality"]

    assert_check("Первый repeated import успешен", first_stats["ok"] is True)
    assert_check("Второй repeated import успешен", second_stats["ok"] is True)
    assert_check("custom_note переживает repeated import", criteria["custom_note"] == "keep_me")
    assert_check("custom_blob переживает repeated import", criteria["custom_blob"] == {"owner": "user"})
    assert_check("genres тоже переживают repeated import", criteria["genres"] == ["триллер"])


def test_import_tmdb_stats_stay_stable_with_criteria_merge() -> None:
    """Проверяет, что merge criteria metadata не ломает import stats."""
    print("\n22) Проверяем стабильность import stats при merge criteria")

    candidate_pool.save_candidate_pool({})
    candidate_pool.save_candidate_criteria({})
    storage_data.clean_dataset()

    result_path = Path(constant.DATA_DIR) / "tmdb_stats_merge.json"
    result = {
        "criteria_name": "tmdb_RU_quality",
        "country": "RU",
        "mode": "quality",
        "settings": {"criteria_name": "tmdb_RU_quality"},
        "candidates": [
            {
                "title": "Stats Merge Candidate",
                "year": 2021,
                "tmdb_id": 504,
                "tmdb_rating": 7.6,
                "tmdb_votes": 61,
            }
        ],
    }
    result_path.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")

    first_stats = tmdb_import.import_tmdb_result_to_common_pool(result_path)
    second_stats = tmdb_import.import_tmdb_result_to_common_pool(result_path)

    assert_check("Первый import читает 1 кандидата", first_stats["read"] == 1)
    assert_check("Первый import добавляет 1 кандидата", first_stats["added"] == 1)
    assert_check("Первый import без ошибок", first_stats["errors"] == 0)
    assert_check("Второй import читает 1 кандидата", second_stats["read"] == 1)
    assert_check("Второй import считает duplicate", second_stats["duplicates"] == 1)
    assert_check("Второй import заполняет skipped_duplicates alias", second_stats["skipped_duplicates"] == 1)
    assert_check("criteria_name в stats не ломается", second_stats["criteria_name"] == "tmdb_RU_quality")


def test_tmdb_candidate_pool_import_wrapper_delegates_to_new_module() -> None:
    """Проверяет, что legacy wrapper вызывает новый import_tmdb модуль."""
    print("\n19) Проверяем wrapper import в tmdb_candidate_pool")

    fake_stats = {
        "ok": True,
        "read": 1,
        "added": 1,
        "updated": 0,
        "watched_skipped": 0,
        "skipped_watched": 0,
        "duplicates": 0,
        "skipped_duplicates": 0,
        "errors": 0,
        "criteria_name": "tmdb_RU_quality",
        "pool_size_before": 0,
        "pool_size_after": 1,
        "pool_size": 1,
    }

    with patch("candidates.import_tmdb.import_tmdb_result_to_common_pool", return_value=fake_stats) as mocked:
        stats = tmdb_candidate_pool.import_tmdb_result_to_common_pool("dummy.json", criteria_name="tmdb_RU_quality")

    assert_check("Wrapper возвращает статистику нового модуля", stats == fake_stats)
    assert_check("Wrapper реально делегирует в новый модуль", mocked.called is True)


def test_tmdb_auto_import_helper_accepts_enter_as_yes() -> None:
    """Проверяет, что Enter считается подтверждением auto-import."""
    print("\n20) Проверяем auto-import helper: Enter = yes")

    calls = []
    output = []

    def fake_import(result_path, criteria_name=None):
        calls.append((result_path, criteria_name))
        return {
            "ok": True,
            "read": 3,
            "added": 2,
            "updated": 1,
            "watched_skipped": 0,
            "skipped_watched": 0,
            "duplicates": 0,
            "skipped_duplicates": 0,
            "errors": 0,
            "criteria_name": criteria_name,
            "pool_size_before": 10,
            "pool_size_after": 12,
            "pool_size": 12,
        }

    stats = interface_funcs.maybe_auto_import_tmdb_result(
        "saved_result.json",
        "tmdb_RU_quality",
        input_func=lambda _prompt: "",
        output_func=output.append,
        import_func=fake_import,
    )

    assert_check("Enter запускает import", calls == [("saved_result.json", "tmdb_RU_quality")])
    assert_check("Успешная статистика возвращается наружу", stats["ok"] is True)
    assert_check("Helper печатает заголовок успешного импорта", any("Импорт TMDb result завершён." in line for line in output))


def test_tmdb_auto_import_helper_n_cancels_import() -> None:
    """Проверяет, что n отменяет auto-import без вызова import."""
    print("\n21) Проверяем auto-import helper: n = cancel")

    output = []

    def fail_import(_result_path, criteria_name=None):
        raise AssertionError(f"Import не должен вызываться: {criteria_name}")

    stats = interface_funcs.maybe_auto_import_tmdb_result(
        "saved_result.json",
        "tmdb_RU_quality",
        input_func=lambda _prompt: "n",
        output_func=output.append,
        import_func=fail_import,
    )

    assert_check("При n helper возвращает None", stats is None)
    assert_check(
        "При n печатается сообщение об отмене",
        any("Импорт отменён. Result сохранён" in line for line in output),
    )


def test_tmdb_auto_import_helper_reports_error_without_crash() -> None:
    """Проверяет, что ошибка import не валит flow и snapshot остаётся сохранённым."""
    print("\n22) Проверяем auto-import helper: ошибка import")

    output = []
    stats = interface_funcs.maybe_auto_import_tmdb_result(
        "saved_result.json",
        "tmdb_RU_quality",
        input_func=lambda _prompt: "",
        output_func=output.append,
        import_func=lambda _path, criteria_name=None: {"ok": False, "error": "boom", "criteria_name": criteria_name},
    )

    assert_check("Helper не бросает исключение и возвращает stats ошибки", stats["ok"] is False)
    assert_check("Печатается понятная ошибка import", any("Авто-импорт не выполнен: boom" in line for line in output))
    assert_check("Печатается подсказка про сохранённый result", any("Result сохранён" in line for line in output))


def test_run_tmdb_candidate_pool_flow_calls_auto_import_after_save() -> None:
    """Проверяет, что обычный TMDb build предлагает auto-import после save."""
    print("\n23) Проверяем normal build -> auto-import prompt")

    db_path = Path(constant.DATA_DIR) / "fake_imdb.sqlite"
    db_path.write_text("ok", encoding="utf-8")
    build_result = {"stats": {"discover_total": 1}, "candidates": []}
    answers = iter(["", "", "", "", "", "", "", "", "", "", "", "y"])
    output = io.StringIO()

    with patch("builtins.input", side_effect=lambda _prompt: next(answers)):
        with patch("apis.imdb_sql.DEFAULT_DB_PATH", db_path):
            with patch("ui.console.interface_funcs.ui.clean_terminal", return_value=None):
                with patch("candidates.tmdb_candidate_pool.build_candidate_pool", return_value=build_result):
                    with patch(
                        "candidates.tmdb_candidate_pool.save_candidate_pool_result",
                        return_value=("data/candidate_pool/tmdb_result.json", "data/candidate_pool/tmdb_result.csv"),
                    ):
                        with patch("ui.console.interface_funcs._print_tmdb_candidate_stats", return_value=None):
                            with patch("ui.console.interface_funcs.maybe_auto_import_tmdb_result") as mocked_auto_import:
                                with contextlib.redirect_stdout(output):
                                    interface_funcs.run_tmdb_candidate_pool_flow(is_test_run=False)

    assert_check("После обычного save вызывается auto-import helper", mocked_auto_import.called is True)
    assert_check(
        "Auto-import helper получает путь к сохранённому result",
        mocked_auto_import.call_args.args == ("data/candidate_pool/tmdb_result.json", "tmdb_RU_quality"),
    )
    assert_check("UI сообщает, что TMDb result сохранён", "TMDb result сохранён: data/candidate_pool/tmdb_result.json" in output.getvalue())


def test_run_tmdb_candidate_pool_flow_test_run_skips_auto_import() -> None:
    """Проверяет, что test-run не предлагает auto-import."""
    print("\n24) Проверяем test-run без auto-import")

    db_path = Path(constant.DATA_DIR) / "fake_imdb_test.sqlite"
    db_path.write_text("ok", encoding="utf-8")
    build_result = {"stats": {"discover_total": 1}, "candidates": []}
    answers = iter(["", "", "", "", "", "", "", "", "", "", "y"])
    output = io.StringIO()

    with patch("builtins.input", side_effect=lambda _prompt: next(answers)):
        with patch("apis.imdb_sql.DEFAULT_DB_PATH", db_path):
            with patch("ui.console.interface_funcs.ui.clean_terminal", return_value=None):
                with patch("candidates.tmdb_candidate_pool.build_candidate_pool", return_value=build_result):
                    with patch(
                        "candidates.tmdb_candidate_pool.save_candidate_pool_test_result",
                        return_value=("data/candidate_pool/tmdb_test_result.json", "data/candidate_pool/tmdb_test_result.csv"),
                    ):
                        with patch("ui.console.interface_funcs._print_tmdb_candidate_stats", return_value=None):
                            with patch("ui.console.interface_funcs.maybe_auto_import_tmdb_result") as mocked_auto_import:
                                with contextlib.redirect_stdout(output):
                                    interface_funcs.run_tmdb_candidate_pool_flow(is_test_run=True)

    assert_check("В test-run auto-import helper не вызывается", mocked_auto_import.called is False)
    assert_check("В test-run нет сообщения про auto-import save prompt", "TMDb result сохранён:" not in output.getvalue())


def test_tmdb_genre_diagnostics_and_helpers() -> None:
    """Проверяет TMDb genre diagnostics и чистые genre helpers без изменения dataset."""
    print("\n15) Проверяем TMDb genre diagnostics")

    dataset = {
        "A": make_movie("A", user_score=8.0),
        "B": make_movie("B", user_score=7.0),
    }
    meta = {
        "A": {"tmdb_id": 1},
    }
    before_dataset = json.loads(json.dumps(dataset, ensure_ascii=False))

    def fake_details(tmdb_id):
        if tmdb_id == 1:
            return {
                "id": 1,
                "name": "A",
                "first_air_date": "2020-01-01",
                "genres": [{"name": "Drama"}, {"name": "Crime"}],
                "external_ids": {},
                "credits": {},
            }
        raise RuntimeError("unexpected tmdb_id")

    progress_events = []
    report = tmdb_candidate_pool.build_tmdb_genre_distribution_report(
        dataset,
        meta,
        details_fetcher=fake_details,
        search_func=lambda _title: [],
        progress_callback=progress_events.append,
    )
    report_path = tmdb_candidate_pool.save_tmdb_genre_distribution_report(
        report,
        output_dir=Path(constant.DATA_DIR) / "diagnostics",
    )
    genres = tmdb_candidate_pool.collect_candidate_genres({
        "genres": [{"name": "Драма"}, "драма"],
        "imdb_genres": [" Crime "],
        "genres_tmdb": ["Mystery"],
        "genre_ids": [80],
    })

    assert_check("Diagnostics считает Drama", report["genre_counts"]["Drama"] == 1)
    assert_check("Diagnostics считает Crime", report["genre_counts"]["Crime"] == 1)
    assert_check("Unmatched попадает в unmatched_items", report["unmatched_items"] == [{"title": "B", "year": 2024}])
    assert_check("Diagnostics JSON создаётся", report_path.exists())
    assert_check("Dataset не меняется при diagnostics", dataset == before_dataset)
    assert_check("Diagnostics progress вызывается по каждой записи", len(progress_events) == 4)
    assert_check("collect_candidate_genres нормализует и убирает дубли", genres == {"драма", "crime", "mystery"})

    error_dataset = {
        "A": make_movie("A", user_score=8.0),
        "B": make_movie("B", user_score=7.0),
        "C": make_movie("C", user_score=6.0),
    }
    error_events = []
    stopped_report = tmdb_candidate_pool.build_tmdb_genre_distribution_report(
        error_dataset,
        {},
        search_func=lambda _title: (_ for _ in ()).throw(RuntimeError("offline")),
        progress_callback=error_events.append,
        max_consecutive_errors=2,
    )
    assert_check("Diagnostics останавливается после серии ошибок", stopped_report["stopped_early"] is True)
    assert_check("Diagnostics не обходит весь dataset при сетевых ошибках", stopped_report["processed"] == 2)
    assert_check("Diagnostics сообщает stop progress event", error_events[-1]["status"] == "stopped")


def run_tests() -> None:
    """Запускает все тесты проекта."""
    temp_dir, old_paths = setup_temp_project()
    try:
        print("=== Тесты проекта recommended ===")
        test_build_watched_movie_card()
        test_export_watched_movies_json()
        test_watched_movies_export_is_read_only()
        test_files_created()
        test_add_new_movie()
        test_meta_overrides_raw_scores()
        test_duplicate_rejected()
        test_feature_formatting()
        test_excel_row_supports_genre()
        test_excel_title_guard()
        test_excel_export_overwrite_refreshes_dataset()
        test_validation()
        test_title_punctuation_validation()
        test_scores_menu_structure_and_safe_update()
        test_linear_distribution_draft()
        test_apply_rating_order_draft()
        test_apply_rating_order_draft_stops_on_stale_or_missing_data()
        test_apply_rating_order_draft_cancel_keeps_dataset()
        test_tag_compatibility()
        test_training()
        test_noise_experiment_uses_loo_metrics()
        test_feature_ablation_helpers()
        test_feature_ablation_alpha_selection()
        test_feature_ablation_subset_weights()
        test_feature_ablation_error_rows()
        test_feature_ablation_collect_report_is_read_only()
        test_feature_ablation_console_report()
        test_genre_markup_efficiency_coverage_and_conclusion()
        test_genre_markup_efficiency_report_is_read_only()
        test_genre_markup_efficiency_console_report()
        test_biggest_error_cards_in_report()
        test_backup_restore()
        test_api_fallback_to_secondary_token()
        test_sql_title_search()
        test_sql_title_search_aliases()
        test_build_api_defaults_from_raw_movie()
        test_merge_defaults_prefers_api_and_keeps_sql()
        test_build_genre_defaults_ignores_unknown()
        test_build_genre_defaults_maps_mystery_to_detective()
        test_split_known_genres_uses_shared_raw_mapper()
        test_manual_add_defaults_when_lookup_fails()
        test_add_defaults_rejects_sql_api_identity_mismatch()
        test_add_defaults_accepts_sql_when_imdb_id_matches()
        test_add_defaults_accepts_sql_when_title_year_match_without_imdb_id()
        test_add_defaults_rejects_sql_when_year_differs_more_than_one()
        test_add_defaults_keeps_sql_only_flow()
        test_add_record_country_selection()
        test_add_resolver_second_pass_sql_after_identity_mismatch()
        test_add_resolver_skips_second_pass_when_api_has_imdb()
        test_add_resolver_rejects_second_pass_sql_mismatch()
        test_add_resolver_accepted_first_pass_skips_second_pass()
        test_add_resolver_sql_only_skips_second_pass()
        test_add_resolver_prioritizes_sql_and_kp()
        test_add_resolver_uses_tmdb_when_kp_fails()
        test_add_resolver_offline_without_sql_is_manual()
        test_candidate_pool_cross_criteria_keys_survive_save_load()
        test_candidate_pool_same_criteria_duplicates_keep_best()
        test_candidate_pool_duplicate_scripts_helpers()
        test_load_candidate_pool_is_read_only()
        test_read_path_keeps_watched_in_json()
        test_write_path_purges_watched_from_json()
        test_get_pool_stats_reports_raw_storage_watched_ready()
        test_build_prediction_filter_defaults_from_saved_criteria()
        test_prediction_filters_apply_saved_criteria_defaults()
        test_genre_normalization_runtime_filters()
        test_canonical_runtime_filters_use_genre_keys_and_country_codes()
        test_prediction_numeric_string_filters_are_safe()
        test_prediction_filter_accepts_numeric_strings()
        test_prediction_filter_rejects_invalid_numeric_strings()
        test_coerce_candidate_number_handles_runtime_values()
        test_prediction_filter_does_not_mutate_input_candidates()
        test_top_prediction_title_identity_dedupe()
        test_contributions_readiness_gate()
        test_candidate_service_read_only_facade()
        test_candidate_service_top_prediction_view()
        test_top_prediction_ui_read_only_helpers()
        test_candidate_service_prediction_filter_defaults_view()
        test_candidate_service_mark_watched_in_pool()
        test_candidate_service_retry_kp_enrichment()
        test_candidate_service_tmdb_import_result()
        test_candidate_service_tmdb_build_and_auto_import()
        test_candidate_service_polish_flows()
        test_candidate_schema_normalizes_legacy_complete_record()
        test_candidate_schema_marks_missing_kp()
        test_candidate_schema_preserves_specific_kp_status()
        test_candidate_schema_ready_for_predict_requires_kp_and_imdb()
        test_kp_country_alias_matching()
        test_kp_lookup_kr_country_aliases()
        test_kp_country_from_iso2_mapping()
        test_kp_series_type_filter()
        test_kp_tmdb_build_debug_traces_rejection_details()
        test_kp_tmdb_build_debug_session_save()
        test_retry_kp_enrichment_makes_candidate_complete()
        test_retry_kp_uses_candidate_country_when_criteria_missing()
        test_candidate_schema_keeps_unknown_fields()
        test_candidate_genre_schema_normalization()
        test_build_candidate_features_uses_genre_keys()
        test_candidate_genre_keys_to_dataset_genres_maps_supported_keys()
        test_candidate_genre_keys_to_dataset_genres_maps_mystery_to_detective()
        test_candidate_genre_keys_to_dataset_genres_reports_unmapped()
        test_candidate_genre_keys_to_dataset_genres_missing()
        test_raw_genres_to_dataset_genres_maps_mystery_to_detective()
        test_raw_genres_to_dataset_genres_maps_en_ru_supported()
        test_raw_genres_to_dataset_genres_reports_unmapped_raw()
        test_raw_genres_to_dataset_genres_missing()
        test_candidate_transfer_payload_uses_genre_keys_mapper()
        test_candidate_transfer_payload_falls_back_to_raw_genres_without_genre_keys()
        test_candidate_transfer_payload_does_not_mutate_candidate()
        test_candidate_genre_transfer_preview_maps_genre_keys()
        test_candidate_genre_transfer_preview_reports_partial_unmapped()
        test_candidate_genre_transfer_preview_falls_back_to_raw_genres()
        test_candidate_transfer_payload_falls_back_to_imdb_genres()
        test_candidate_genre_transfer_preview_falls_back_to_imdb_genres()
        test_candidate_genre_transfer_preview_warns_when_all_zero_with_raw_signals()
        test_candidate_genre_transfer_preview_does_not_mutate_candidate()
        test_mark_candidate_as_watched_prints_genre_transfer_preview()
        test_candidate_country_schema_normalization()
        test_remove_candidate_from_pool()
        test_candidate_pool_genre_filters()
        test_tmdb_candidate_pool_criteria_name()
        test_tmdb_candidate_pool_network_error_skip_gate()
        test_tmdb_candidate_pool_discover_genre_filters()
        test_tmdb_import_keeps_cross_criteria_entries()
        test_import_tmdb_module_normalizes_schema_and_keeps_unknown_fields()
        test_import_tmdb_module_skips_watched_by_title_identity()
        test_import_tmdb_preserves_existing_criteria_filters()
        test_import_tmdb_creates_new_criteria_with_safe_defaults()
        test_import_tmdb_repeated_import_keeps_custom_unknown_fields()
        test_import_tmdb_stats_stay_stable_with_criteria_merge()
        test_tmdb_candidate_pool_import_wrapper_delegates_to_new_module()
        test_tmdb_auto_import_helper_accepts_enter_as_yes()
        test_tmdb_auto_import_helper_n_cancels_import()
        test_tmdb_auto_import_helper_reports_error_without_crash()
        test_run_tmdb_candidate_pool_flow_calls_auto_import_after_save()
        test_run_tmdb_candidate_pool_flow_test_run_skips_auto_import()
        test_tmdb_genre_diagnostics_and_helpers()
        print("\nВсе проверки пройдены: True")
    finally:
        restore_project_paths(temp_dir, old_paths)


if __name__ == "__main__":
    run_tests()
